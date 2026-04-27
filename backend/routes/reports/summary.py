# summary, settlement receipt, advance report, dan bulk export endpoints

import re
import logging
from datetime import datetime, date
from io import BytesIO
from typing import Optional, List, Dict, Any, Tuple, Union
from flask import jsonify, send_file, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase.pdfmetrics import stringWidth

# Setup logging
logger = logging.getLogger(__name__)

from sqlalchemy.orm import joinedload
from models import User, Expense, Category, Settlement, db, Advance
from . import reports_bp
from .helpers import _default_report_year, _parse_iso_date, _year_date_bounds

COMMON_SUBCATEGORY_ORDER = [
    'Rental Tool',
    'Sales',
    'Gaji',
    'Pembuatan Alat',
    'Allowance',
    'Data Processing',
    'Project Operation',
    'Sparepart',
    'Maintenance',
    'Software License',
    'Operation',
    'Sewa Ruangan',
    'Modal Kerja',
    'Team Building',
    'Biaya Bank',
]


def _validate_date(date_input: Any) -> Optional[datetime]:
    """Validate date format and return datetime object. Handles strings and date/datetime objects."""
    if not date_input:
        return None

    # If it's already a datetime object
    if isinstance(date_input, datetime):
        return date_input

    # If it's a date object (but not datetime)
    if isinstance(date_input, date):
        return datetime.combine(date_input, datetime.min.time())

    try:
        # If it's a string
        return datetime.strptime(str(date_input), '%Y-%m-%d')
    except (ValueError, TypeError):
        return None


def _display_settlement_status(status: str) -> str:
    """Convert settlement status to display format."""
    return 'approved' if status == 'completed' else status


def _get_summary_approved_expenses(
    start_date: Optional[Union[str, date, datetime]] = None,
    end_date: Optional[Union[str, date, datetime]] = None,
    year: Optional[int] = None,
    mode: str = 'report'
) -> List[Expense]:
    """Get approved expenses with optional year/mode filtering."""
    query = Expense.query.filter(Expense.status == 'approved')

    # ✅ PRIORITAS: Filter berdasarkan mode report (report_year)
    if mode == 'report' and year:
        query = query.join(Settlement).filter(Settlement.report_year == year)
    else:
        # Mode 'actual' atau 'range': Filter berdasarkan tanggal transaksi
        if start_date:
            validated = _validate_date(start_date)
            if validated:
                query = query.filter(Expense.date >= validated.date())

        if end_date:
            validated = _validate_date(end_date)
            if validated:
                # Jika end_date berasal dari _year_date_bounds, itu adalah Jan 1 tahun depan (exclusive)
                # Tapi di sini kita asumsikan end_date adalah inclusive batas atas
                query = query.filter(Expense.date <= validated.date())

    return query.all()


def _expense_amount_for_summary(expense: Expense) -> float:
    """Get expense amount in IDR."""
    return expense.idr_amount or 0.0


def _get_display_subcategory(expense: Expense) -> str:
    """Get the consolidated subcategory label from the Expense model or description."""
    label = (expense.combined_subcategory_label or '').strip()
    if label:
        return label

    # Jika label kosong, coba ekstrak dari deskripsi format [Subkategori]
    if expense.description:
        match = re.match(r'^\[(.*?)\]', expense.description.strip())
        if match:
            return match.group(1).strip()

    return ''


def _summary_display_category(expense, category_by_id, children_by_parent_name):
    category = category_by_id.get(expense.category_id)
    if not category:
        return None, None

    # Cari root (kategori paling atas)
    root = category
    path = [category]
    while root.parent_id and category_by_id.get(root.parent_id):
        root = category_by_id[root.parent_id]
        path.append(root)

    root_name = root.name or '-'

    # 1. PRIORITAS: Jika expense.category_id itu sendiri adalah subkategori (punya parent)
    # dan parent-nya adalah root, maka itu adalah anak langsung.
    if category.parent_id == root.id:
        return root, category

    # 2. Jika category_id adalah level yang lebih dalam (cucu), cari anak langsung dari root.
    if len(path) > 2: # [category, child, root] -> length 3
        direct_child = path[-2]
        if direct_child.parent_id == root.id:
            return root, direct_child

    # 3. FALLBACK: Coba cocokkan berdasarkan label teks (inferred_subcategory)
    inferred_subcategory_name = _get_display_subcategory(expense)
    if inferred_subcategory_name:
        children_map = children_by_parent_name.get(root_name, {})
        key_search = inferred_subcategory_name.lower().strip()

        # Hilangkan prefix kode jika ada di key_search (misal 'A11-Sales' -> 'sales')
        if '-' in key_search:
            key_search = key_search.split('-', 1)[1].strip()

        for child_name, child_obj in children_map.items():
            c_name_low = child_name.lower().strip()
            # Hilangkan prefix kode di nama kategori database juga untuk perbandingan
            pure_child_name = c_name_low
            if '-' in pure_child_name:
                pure_child_name = pure_child_name.split('-', 1)[1].strip()

            # Cocokkan jika nama murni sama, atau salah satu mengandung yang lain
            if (key_search == pure_child_name or
                key_search == c_name_low or
                key_search in pure_child_name or
                pure_child_name in key_search):
                return root, child_obj
    # Jika expense dicatat langsung di kategori Induk (category_id == root_id)
    return root, None


def _build_summary_payload(expenses):
    all_categories = Category.query.order_by(db.func.length(Category.code).asc(), Category.code.asc()).all()
    # Ambil kategori akar (parent_id is None) - Urutkan alami (A, B, C...)
    root_categories = [cat for cat in all_categories if cat.parent_id is None]
    category_by_id = {cat.id: cat for cat in all_categories}
    children_by_parent_id = {}
    for cat in all_categories:
        if cat.parent_id is not None:
            children_by_parent_id.setdefault(cat.parent_id, []).append(cat)

    children_by_parent_name = {}
    for root in root_categories:
        children = children_by_parent_id.get(root.id, [])
        ordered_children = []
        used_ids = set()
        for child_name in COMMON_SUBCATEGORY_ORDER:
            match = next((child for child in children if child.name == child_name), None)
            if match:
                ordered_children.append(match)
                used_ids.add(match.id)
        for child in children:
            if child.id not in used_ids:
                ordered_children.append(child)
        children_by_parent_name[root.name] = {child.name: child for child in ordered_children}

    summary_data = []
    grand_total = 0
    expense_map_by_root = {
        root.id: {'root_expenses': [], 'child_expense_map': {}}
        for root in root_categories
    }

    for expense in expenses:
        mapped_root, mapped_child = _summary_display_category(
            expense,
            category_by_id,
            children_by_parent_name,
        )
        if not mapped_root:
            continue
        root_bucket = expense_map_by_root.setdefault(
            mapped_root.id,
            {'root_expenses': [], 'child_expense_map': {}},
        )
        if mapped_child:
            root_bucket['child_expense_map'].setdefault(mapped_child.id, []).append(expense)
        else:
            root_bucket['root_expenses'].append(expense)

    for root in root_categories:
        root_bucket = expense_map_by_root.get(root.id, {})
        root_expenses = root_bucket.get('root_expenses', [])
        child_expense_map = root_bucket.get('child_expense_map', {})

        children = list(children_by_parent_name.get(root.name, {}).values())
        children_list = []

        # 1. Proses Subkategori Resmi
        for child in children:
            child_expenses = child_expense_map.get(child.id, [])
            monthly = {str(i): 0 for i in range(1, 13)}
            yearly_total = 0
            for e in child_expenses:
                month_str = str(e.date.month)
                amount = _expense_amount_for_summary(e)
                monthly[month_str] += amount
                yearly_total += amount

            # Tambahkan ke grand_total hanya di level ini
            grand_total += yearly_total

            children_list.append({
                'category': f"{child.code} - {child.name}",
                'monthly': monthly,
                'yearly_total': yearly_total,
                'is_parent': False,
                'level': 1
            })

        # 2. Proses Baris Sisa (Yang masuk langsung ke Root)
        if root_expenses:
            combine_monthly = {str(i): 0 for i in range(1, 13)}
            combine_yearly_total = 0
            labels = set()
            for e in root_expenses:
                month_str = str(e.date.month)
                amount = _expense_amount_for_summary(e)
                combine_monthly[month_str] += amount
                combine_yearly_total += amount

                lbl = _get_display_subcategory(e)
                if lbl: labels.add(lbl)

            grand_total += combine_yearly_total

            label_str = " - ".join(sorted(list(labels))) if labels else "Uncategorized"
            # Jangan gunakan kode dummy '0', gunakan kode Root saja agar tidak membingungkan
            children_list.insert(0, {
                'category': f"{root.code} (Other) - {label_str}",
                'monthly': combine_monthly,
                'yearly_total': combine_yearly_total,
                'is_parent': False,
                'level': 1
            })
        # 3. Hitung Total Kategori Induk (dari semua baris di children_list)
        parent_monthly = {str(i): 0 for i in range(1, 13)}
        parent_yearly_total = 0.0

        for child_row in children_list:
            for i in range(1, 13):
                month_key = str(i)
                parent_monthly[month_key] += child_row['monthly'].get(month_key, 0) or 0
            parent_yearly_total += child_row.get('yearly_total', 0) or 0

        # Masukkan baris INDUK
        summary_data.append({
            'category': f"{root.code} - {root.name}",
            'monthly': parent_monthly,
            'yearly_total': parent_yearly_total,
            'is_parent': True,
            'level': 0
        })

        # Masukkan semua anak (termasuk baris combine tadi)
        summary_data.extend(children_list)

    return summary_data, grand_total


@reports_bp.route('/summary', methods=['GET'])
@jwt_required()
def get_summary_report():
    user = User.query.get(int(get_jwt_identity()))
    print(f'[SUMMARY_API] User={user.username}, role={user.role}')

    if user.role != 'manager':
        return jsonify({'error': 'Akses ditolak'}), 403

    year = request.args.get('year', type=int)
    mode = request.args.get('mode', default='report')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    # Jika mode bukan range, gunakan bounds tahun
    if mode != 'range' and year:
        # ✅ FIX: _year_date_bounds hanya butuh 1 argumen
        start_date, end_date = _year_date_bounds(year)
    else:
        start_date = _parse_iso_date(start_date_str)
        end_date = _parse_iso_date(end_date_str)

    # ✅ Update call dengan parameter baru
    expenses = _get_summary_approved_expenses(
        start_date=start_date,
        end_date=end_date,
        year=year,
        mode=mode
    )
    summary_data, grand_total = _build_summary_payload(expenses)

    return jsonify({
        'summary': summary_data,
        'grand_total': grand_total,
        'year': year,
        'mode': mode,
        'start_date': start_date.isoformat() if start_date else None,
        'end_date': end_date.isoformat() if end_date else None
    }), 200


@reports_bp.route('/summary/pdf', methods=['GET'])
@jwt_required()
def export_summary_pdf():
    user = User.query.get(int(get_jwt_identity()))
    if user.role != 'manager':
        return jsonify({'error': 'Akses ditolak'}), 403

    year = request.args.get('year', type=int)
    mode = request.args.get('mode', default='report')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if mode != 'range' and year:
        # ✅ FIX: _year_date_bounds hanya butuh 1 argumen
        start_date, end_date = _year_date_bounds(year)
    else:
        start_date = _parse_iso_date(start_date_str)
        end_date = _parse_iso_date(end_date_str)

    # ✅ Update call dengan parameter baru
    expenses = _get_summary_approved_expenses(
        start_date=start_date,
        end_date=end_date,
        year=year,
        mode=mode
    )
    summary_data, grand_total = _build_summary_payload(expenses)

    buffer = BytesIO()
    # Perkecil margin agar tabel punya ruang lebih luas (792 total lebar landscape)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=20, rightMargin=20, topMargin=30, bottom=30)
    styles = getSampleStyleSheet()
    elements = []

    title = "Laporan Summary Pengeluaran"
    if start_date or end_date:
        period_text = f"Periode: {(start_date.isoformat() if start_date else '-')} s/d {(end_date.isoformat() if end_date else '-')}"
    else:
        period_text = "Periode: Semua Data"

    elements.append(Paragraph(title, styles['Heading1']))
    elements.append(Paragraph(period_text, styles['Normal']))
    elements.append(Spacer(1, 10))

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agt', 'Sep', 'Okt', 'Nov', 'Des']
    table_data = [["Kategori"] + months + ["Total"]]

    # --- HITUNG LEBAR DINAMIS ---
    # Gunakan font sedikit lebih kecil agar angka besar (milyaran) tidak bertabrakan
    fontSize = 6.5
    max_cat_name_width = 110 # Minimum awal
    for row in summary_data:
        cat_name = str(row.get('category', '-'))
        level = row.get('level', 0)
        indent = level * 10
        w = stringWidth(cat_name, 'Helvetica-Bold' if row.get('is_parent') else 'Helvetica', fontSize) + indent + 12
        if w > max_cat_name_width:
            max_cat_name_width = w

    # Cap category width agar kolom angka tetap proporsional
    if max_cat_name_width > 180:
        max_cat_name_width = 180

    month_totals = [0] * 12
    for row in summary_data:
        monthly = row.get('monthly', {})
        is_parent = row.get('is_parent', False)
        level = row.get('level', 0)

        values = []
        for i in range(1, 13):
            val = float(monthly.get(str(i), 0) or 0)
            month_totals[i - 1] += val
            values.append(f"{val:,.0f}" if val > 0 else "-")

        cat_name = row.get('category', '-')
        p_style = ParagraphStyle('TableText', parent=styles['Normal'], fontSize=fontSize, leading=8)
        if is_parent:
            display_name = Paragraph(f"<b>{cat_name}</b>", p_style)
        else:
            p_style.leftIndent = level * 10
            display_name = Paragraph(cat_name, p_style)

        table_data.append([display_name] + values + [f"{float(row.get('yearly_total', 0) or 0):,.0f}"])

    table_data.append([Paragraph("<b>GRAND TOTAL</b>", ParagraphStyle('Total', parent=styles['Normal'], fontSize=fontSize))] +
                      [f"{v:,.0f}" if v > 0 else "-" for v in month_totals] + [f"{float(grand_total):,.0f}"])

    # Atur lebar kolom: month_width ditingkatkan ke 45, total_width ke 75
    month_col_width = 45
    last_col_width = 75
    col_widths = [max_cat_name_width] + [month_col_width] * 12 + [last_col_width]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), fontSize),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'), ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6F0FF')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buffer, as_attachment=True, download_name=f"summary_{timestamp}.pdf", mimetype='application/pdf')


@reports_bp.route('/excel', methods=['GET'])
@jwt_required()
def generate_excel_report():
    user = User.query.get(int(get_jwt_identity()))
    if user.role != 'manager':
        return jsonify({'error': 'Akses ditolak'}), 403

    year = request.args.get('year', _default_report_year(), type=int)
    mode = request.args.get('mode', default='report')
    start_date = _parse_iso_date(request.args.get('start_date'))
    end_date = _parse_iso_date(request.args.get('end_date'))

    # ✅ Update call dengan parameter baru dan perbaikan jumlah argumen
    expenses = _get_summary_approved_expenses(
        start_date=start_date,
        end_date=end_date,
        year=year,
        mode=mode
    )
    summary_data, grand_total_val = _build_summary_payload(expenses)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Summary Bulanan"

    # --- HEADER SECTION ---
    # Row 1: Title
    ws['A1'] = "Daily Report - Balance Calculation Operation Cost"
    ws['A1'].font = Font(bold=True, size=14)

    # Row 2: Project
    ws['A2'] = "Project"
    ws['B2'] = ":"
    ws['C2'] = f"REVENUE vs OPERATION COST Tahun {year}"
    ws['A2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)

    # Row 3: Date
    ws['A3'] = "Date"
    ws['B3'] = ":"
    if start_date and end_date:
        ws['C3'] = f"{start_date.strftime('%d %B')} - {end_date.strftime('%d %B %Y')}"
    else:
        ws['C3'] = f"Januari - Desember {year}"
    ws['A3'].font = Font(bold=True)
    ws['C3'].font = Font(bold=True)

    # --- LOGO SECTION ---
    logo_path = r"D:\2. Organize\1. Projects\MiniProjectKPI_EWI\frontend\assets\images\sheet1.png"
    if os.path.exists(logo_path):
        img = OpenpyxlImage(logo_path)
        # Sesuai analisis screenshot: Width 6.69cm (~253px), Height 2.22cm (~84px)
        img.width = 253
        img.height = 84
        ws.add_image(img, 'N1') # Anchor di N1 agar sejajar ke arah kolom O

    # --- TABLE SECTION ---
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agt', 'Sep', 'Okt', 'Nov', 'Des']
    headers = ["Kategori"] + months + ["Total"]

    start_row = 6
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=i)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid') # Blue header
        cell.alignment = Alignment(horizontal='center')

    current_row = start_row + 1
    for row in summary_data:
        is_parent = row.get('is_parent', False)
        level = row.get('level', 0)
        cat_name = row.get('category', '-')

        display_name = ("  " * level) + cat_name
        ws.cell(row=current_row, column=1).value = display_name

        if is_parent:
            ws.cell(row=current_row, column=1).font = Font(bold=True)

        monthly = row.get('monthly', {})
        for i in range(1, 13):
            val = monthly.get(str(i), 0)
            cell = ws.cell(row=current_row, column=i+1)
            cell.value = val
            cell.number_format = '#,##0'
            if is_parent:
                cell.font = Font(bold=True)

        total_val = row.get('yearly_total', 0)
        total_cell = ws.cell(row=current_row, column=14)
        total_cell.value = total_val
        total_cell.number_format = '#,##0'
        if is_parent:
            total_cell.font = Font(bold=True)

        current_row += 1

    # Grand Total row
    ws.cell(row=current_row, column=1).value = "GRAND TOTAL"
    ws.cell(row=current_row, column=1).font = Font(bold=True)

    for i in range(1, 13):
        total_m = sum(r['monthly'].get(str(i), 0) for r in summary_data if r.get('level') == 0) # Only sum top level to avoid double count
        cell = ws.cell(row=current_row, column=i+1)
        cell.value = total_m
        cell.number_format = '#,##0'
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')

    gt_cell = ws.cell(row=current_row, column=14)
    gt_cell.value = grand_total_val
    gt_cell.number_format = '#,##0'
    gt_cell.font = Font(bold=True)
    gt_cell.fill = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')

    # Column widths
    ws.column_dimensions['A'].width = 40
    for i in range(2, 14):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions['N'].width = 18

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buffer, as_attachment=True, download_name=f"summary_{timestamp}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/excel_advance', methods=['GET'])
@jwt_required()
def generate_excel_advance_report():
    # Natural Sort: panjang kode -> isi kode (A1, A2... A10)
    # Sertakan semua kategori agar data di kategori induk ikut muncul
    categories = Category.query.order_by(db.func.length(Category.code).asc(), Category.code.asc()).all()

    category_map = {cat.id: cat for cat in categories}
    cat_col_map = {cat.name: i + 4 for i, cat in enumerate(categories)}
    from models import Advance

    advance_id = request.args.get('advance_id', type=int)
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    query = Advance.query.filter_by(id=advance_id) if advance_id else Advance.query.filter_by(status='approved')

    if start_date_str:
        try:
            query = query.filter(Advance.approved_at >= datetime.strptime(start_date_str, '%Y-%m-%d'))
        except ValueError as e:
            logger.debug(f'Invalid start_date format: {start_date_str} - {e}')

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(Advance.approved_at <= end_date)
        except ValueError as e:
            logger.debug(f'Invalid end_date format: {end_date_str} - {e}')

    approved_advances = query.order_by(Advance.approved_at.asc()).all()
    wb = Workbook(); ws = wb.active; ws.title = "Laporan Perencanaan Kasbon"
    headers = ["Tanggal Disetujui", "Deskripsi Kasbon", "Karyawan", "Judul Kasbon"] + [cat.name for cat in categories]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)
    for advance in approved_advances:
        row_data = [None] * len(headers)
        row_data[0] = advance.approved_at if advance.approved_at else None
        row_data[1] = advance.description; row_data[2] = advance.requester.full_name if advance.requester else ''; row_data[3] = advance.title
        cat_sums = {}
        for item in advance.items:
            cn = category_map[item.category_id].name if item.category_id in category_map else None
            if cn: cat_sums[cn] = cat_sums.get(cn, 0) + (item.idr_amount or 0)
        for cn, amt in cat_sums.items():
            if cn in cat_col_map: row_data[cat_col_map[cn]] = amt
        ws.append(row_data)
    for i, col_name in enumerate(headers, 1):
        w = len(col_name) + 4
        if w > 35: w = 35
        ws.column_dimensions[get_column_letter(i)].width = w
    for col, w in {'A':18,'B':35,'C':20,'D':30}.items(): ws.column_dimensions[col].width = w
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buffer, as_attachment=True, download_name=f"laporan_kasbon_{timestamp}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/advance/<int:advance_id>/pdf', methods=['GET'])
@jwt_required()
def generate_pdf_advance_report(advance_id):
    user_id = int(get_jwt_identity()); user = User.query.get(user_id)
    from models import Advance
    advance = Advance.query.get_or_404(advance_id)
    if user.role != 'manager' and advance.user_id != user_id:
        return jsonify({'error': 'Akses ditolak'}), 403
    buffer = BytesIO(); doc = SimpleDocTemplate(buffer, pagesize=letter); elements = []; styles = getSampleStyleSheet()
    title_style = styles['Heading1']; title_style.alignment = 1
    elements.append(Paragraph("Kasbon (Advance Request)", title_style)); elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Judul:</b> {advance.title}", styles['Normal']))
    elements.append(Paragraph(f"<b>Deskripsi:</b> {advance.description}", styles['Normal']))
    elements.append(Paragraph(f"<b>Pemohon:</b> {advance.requester.full_name}", styles['Normal']))
    elements.append(Paragraph(f"<b>Status:</b> {advance.status.upper()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Estimasi:</b> Rp {advance.total_amount:,.0f}", styles['Normal']))
    elements.append(Paragraph(f"<b>Disetujui:</b> Rp {advance.approved_amount:,.0f}", styles['Normal']))
    elements.append(Spacer(1, 12))
    data = [["Kategori", "Deskripsi", "Mata Uang", "Amount", "Kurs", "Estimasi (IDR)"]]
    for item in advance.items:
        currency = item.currency or 'IDR'
        amount = item.estimated_amount or 0
        kurs = item.currency_exchange or 1
        idr = item.idr_amount or 0
        data.append([
            item.category.name if item.category else "-",
            item.description,
            currency,
            f"{amount:,.0f}",
            f"{kurs:,.0f}",
            f"{idr:,.0f}",
        ])
    t = Table(data, colWidths=[120, 180, 40, 60, 50, 80])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'LEFT'),('ALIGN',(3,0),(5,-1),'RIGHT'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,0),12),('BACKGROUND',(0,1),(-1,-1),colors.beige),('GRID',(0,0),(-1,-1),1,colors.black)]))
    elements.append(t); doc.build(elements); buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buffer, as_attachment=True, download_name=f"kasbon_{advance_id}_{timestamp}.pdf", mimetype='application/pdf')


@reports_bp.route('/settlement/<int:settlement_id>/receipt', methods=['GET'])
@jwt_required()
def generate_settlement_receipt(settlement_id):
    user_id = int(get_jwt_identity()); user = User.query.get(user_id)
    settlement = Settlement.query.get_or_404(settlement_id)
    if user.role != 'manager' and settlement.user_id != user_id:
        return jsonify({'error': 'Akses ditolak'}), 403
    buffer = BytesIO(); doc = SimpleDocTemplate(buffer, pagesize=letter); elements = []; styles = getSampleStyleSheet()
    title_style = styles['Heading1']; title_style.alignment = 1
    elements.append(Paragraph("Settlement Receipt", title_style)); elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Judul:</b> {settlement.title}", styles['Normal']))
    elements.append(Paragraph(f"<b>Deskripsi:</b> {settlement.description}", styles['Normal']))
    elements.append(Paragraph(f"<b>Karyawan:</b> {settlement.creator.full_name}", styles['Normal']))
    elements.append(
        Paragraph(
            f"<b>Status:</b> {_display_settlement_status(settlement.status).upper()}",
            styles['Normal'],
        )
    )
    elements.append(Paragraph(f"<b>Total Pengajuan:</b> Rp {settlement.total_amount:,.0f}", styles['Normal']))
    elements.append(Paragraph(f"<b>Disetujui:</b> Rp {settlement.approved_amount:,.0f}", styles['Normal']))
    elements.append(Spacer(1, 12))
    data = [["No","Tanggal","Kategori","Deskripsi","Mata Uang","Amount","Kurs","Ekivalen (IDR)","Status"]]
    for i, exp in enumerate(settlement.expenses, 1):
        data.append([str(i),exp.date.strftime('%Y-%m-%d'),exp.category.name if exp.category else "-",exp.description,exp.currency or 'IDR',f"{exp.amount:,.0f}",f"{exp.currency_exchange or 1:,.0f}",f"{exp.idr_amount or 0:,.0f}",exp.status.upper()])
    t = Table(data, colWidths=[20,60,80,110,40,50,40,70,50])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'LEFT'),('ALIGN',(5,0),(7,-1),'RIGHT'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,0),8),('GRID',(0,0),(-1,-1),1,colors.black)]))
    elements.append(t); doc.build(elements); buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buffer, as_attachment=True, download_name=f"receipt_{settlement_id}_{timestamp}.pdf", mimetype='application/pdf')


@reports_bp.route('/settlements/pdf', methods=['GET'])
@jwt_required()
def generate_bulk_settlements_pdf():
    user = User.query.get(int(get_jwt_identity()))
    if user.role != 'manager': return jsonify({'error': 'Akses ditolak'}), 403
    status = request.args.get('status'); start_date_str = request.args.get('start_date'); end_date_str = request.args.get('end_date'); report_year = request.args.get('report_year', type=int)
    query = Settlement.query
    if status in ('approved', 'completed'):
        query = query.filter(Settlement.status.in_(('approved', 'completed')))
    elif status and status != 'null':
        query = query.filter_by(status=status)
    if report_year is not None:
        query = query.join(Expense, Expense.settlement_id == Settlement.id).filter(
            db.extract('year', Expense.date) == report_year
        ).distinct()
    if start_date_str: query = query.filter(Settlement.created_at >= datetime.strptime(start_date_str, '%Y-%m-%d').date())
    if end_date_str: query = query.filter(Settlement.created_at <= datetime.strptime(end_date_str, '%Y-%m-%d').date())
    settlements = query.order_by(Settlement.created_at.desc()).all()
    buffer = BytesIO(); doc = SimpleDocTemplate(buffer, pagesize=letter); elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("Laporan Bulk Settlement", styles['Heading1'])); elements.append(Spacer(1, 12))
    data = [["ID","Judul","Karyawan","Status","Total (Rp)"]]
    for s in settlements:
        data.append([
            str(s.id),
            s.title,
            s.creator.full_name,
            _display_settlement_status(s.status).upper(),
            f"{s.total_amount:,.0f}",
        ])
    t = Table(data, colWidths=[40,180,120,80,80])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'LEFT'),('GRID',(0,0),(-1,-1),1,colors.black)]))
    elements.append(t); doc.build(elements); buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buffer, as_attachment=True, download_name=f"bulk_settlement_{timestamp}.pdf", mimetype='application/pdf')


@reports_bp.route('/advances/pdf', methods=['GET'])
@jwt_required()
def generate_bulk_advances_pdf():
    user = User.query.get(int(get_jwt_identity()))
    if user.role != 'manager': return jsonify({'error': 'Akses ditolak'}), 403
    from models import Advance
    start_date_str = request.args.get('start_date'); end_date_str = request.args.get('end_date')
    query = Advance.query
    if start_date_str: query = query.filter(Advance.created_at >= datetime.strptime(start_date_str, '%Y-%m-%d').date())
    if end_date_str: query = query.filter(Advance.created_at <= datetime.strptime(end_date_str, '%Y-%m-%d').date())
    advances = query.order_by(Advance.created_at.desc()).all()
    buffer = BytesIO(); doc = SimpleDocTemplate(buffer, pagesize=letter); elements = []; styles = getSampleStyleSheet()
    elements.append(Paragraph("Laporan Bulk Kasbon (Advances)", styles['Heading1'])); elements.append(Spacer(1, 12))
    data = [["ID","Judul","Pemohon","Status","Total (Rp)"]]
    for a in advances: data.append([str(a.id),a.title,a.requester.full_name,a.status.upper(),f"{a.total_amount:,.0f}" if a.total_amount else "0"])
    t = Table(data, colWidths=[40,180,120,80,80])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('ALIGN',(0,0),(-1,-1),'LEFT'),('GRID',(0,0),(-1,-1),1,colors.black)]))
    elements.append(t); doc.build(elements); buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buffer, as_attachment=True, download_name=f"bulk_kasbon_{timestamp}.pdf", mimetype='application/pdf')
