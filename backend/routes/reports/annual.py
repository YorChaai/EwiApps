# annual laporan endpoints dan helpers (pdf, excel)

import os
import re
import json
import logging
from datetime import datetime, date
from io import BytesIO
from copy import copy
from collections import OrderedDict
from typing import List, Dict, Optional, Any, Tuple
from flask import jsonify, send_file, current_app, request
from flask_jwt_extended import jwt_required, get_jwt_identity
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from sqlalchemy import text as sql_text
from sqlalchemy.orm import joinedload

from models import User, Expense, Category, Settlement, ManualCombineGroup, db
from . import reports_bp
from .helpers import (
    _default_report_year, _safe_set_cell, _safe_set_number, _get_top_left_cell,
    _safe_set_cell_with_merge, _merge_description_cell,
    _normalize_external_formula_refs, _clear_range, _clear_range_force, _clear_data_keep_formulas,
    _set_rows_hidden, _safe_text, _extract_imported_row, _extract_imported_sheet_row,
    _is_template_detail_data_row, _map_expense_category_index_from_name,
    _pick_template_formula_col, _get_expense_blocks, _parse_iso_date,
    _to_float, _as_iso_date, _idr_from_currency, _shorten, _map_expense_column,
    _is_batch_settlement, _clean_settlement_title, _group_annual_expenses,
    _format_date_dd_mmm_yy, _set_formula_with_format, _set_date_with_format,
    _year_date_bounds,
)

# Setup logging
logger = logging.getLogger(__name__)

# Column width constants
COLUMN_WIDTH_CURRENCY = 5
COLUMN_WIDTH_EXCHANGE = 10
COLUMN_WIDTH_CATEGORY = 18
ROW_HEIGHT_HEADER_VERTICAL = 60
ROW_HEIGHT_DIVIDEND = 20

# Definisi Lebar Kolom Minimum (MIN_WIDTHS)
MIN_WIDTHS = {
    'A': 0.6328125, 'B': 11.81640625, 'C': 4.0, 'D': 56.26953125,
    'E': 4.6328125, 'F': 13.36328125, 'G': 5.0, 'H': 10.0,
    'I': 18.0, 'J': 18.0, 'K': 18.0, 'L': 18.0, 'M': 18.0,
    'N': 18.0, 'O': 18.0, 'P': 18.0, 'Q': 18.0
}

# Row constants
EXPENSE_HEADER_ROW = 40
EXPENSE_START_ROW = 41
REVENUE_START_ROW = 8
TAX_START_ROW = 26

# Template row limits
REVENUE_TEMPLATE_END = 22
TAX_TEMPLATE_END = 37

PARENT_CATEGORY_NAMES = {
    'biaya operasi',
    'biaya research (r&d)',
    'biaya sewa peralatan',
    'biaya interpretasi log data',
    'administrasi',
    'pembelian barang',
    'sewa kantor',
    'kesehatan',
    'bisnis dev',
}

# Green style for category headers (matching template)
GREEN_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')

# Border style for all cells (thin black border)
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Blue fill for batch expense headers
BLUE_FILL = PatternFill(fill_type='solid', fgColor='92CDDC')




def _extract_subcategory_from_description(description):
    """Ekstrak sub-kategori dari deskripsi dengan format [SubCategory] Description"""
    if not description:
        return ''
    match = re.match(r'^\[(.*?)\]\s*', description.strip())
    if match:
        return match.group(1).strip()
    return ''


def _root_category_info(category_id, category_by_id):
    cat = category_by_id.get(category_id)
    if not cat: return '-', '-'
    subcategory_name = cat.name
    while cat.parent_id and category_by_id.get(cat.parent_id):
        cat = category_by_id[cat.parent_id]
    return cat.name or '-', subcategory_name or '-'


def _subcategory_sort_key(label):
    text = _safe_text(label).strip()
    if not text or text == '-':
        return (1, '')
    return (0, text.lower())


def _write_dynamic_category_headers(ws, root_cats, header_row=40, start_col=9, template_end_col=30):
    from openpyxl.styles import PatternFill, Border, Side
    no_fill = PatternFill(fill_type=None)
    no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

    last_cat_col = start_col + len(root_cats) - 1

    # Reset styling for all potential ghost columns
    for col in range(start_col, template_end_col + 1):
        cell = ws.cell(row=header_row, column=col)
        if isinstance(cell, MergedCell):
            continue
        if col > last_cat_col:
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border

    for offset, category in enumerate(root_cats):
        col = start_col + offset
        cell = ws.cell(row=header_row, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.value = category.name

        # ✅ APPLY GREEN STYLE TO ALL CATEGORIES (OLD + NEW)
        cell.fill = GREEN_FILL
        # ✅ FIX: Use Arial Narrow 12 instead of Aptos Narrow 10
        cell.font = Font(bold=True, size=12, name='Arial Narrow')

        # ✅ APPLY BORDER TO CATEGORY HEADERS
        cell.border = THIN_BORDER

        # ✅ FIX: wrap_text=True supaya nama kategori panjang wrap ke bawah
        alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.alignment = alignment


def _clone_row_format(ws, source_row, target_row, start_col=1, end_col=17):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(start_col, end_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        if isinstance(source_cell, MergedCell) or isinstance(target_cell, MergedCell):
            continue
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format


def _replace_text_in_sheet(ws, search_text, replace_text):
    """
    Search and replace text in all cells of a worksheet.
    Used for updating headers like 'TAHUN 2024' -> 'TAHUN 2030'.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if search_text in cell.value:
                    cell.value = cell.value.replace(search_text, replace_text)


def _apply_standard_cell_style(cell, style_type='data', bold=False, align='left', fill=None):
    """
    Helper pusat untuk merampingkan styling sel Excel.
    style_type: 'header', 'data', 'total', 'final'
    """
    # 1. Font
    font_size = 12 if style_type == 'header' else 11
    font_name = 'Arial Narrow'
    cell.font = Font(bold=bold or (style_type in ('header', 'total', 'final')),
                     size=font_size, name=font_name)

    # 2. Alignment
    h_align = 'center' if style_type == 'header' else align
    cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=(style_type == 'header'))

    # 3. Border
    s_thin = Side(style='thin', color='000000')
    s_double = Side(style='double', color='000000')

    top_s = s_thin if style_type in ('total', 'final') else s_thin
    bottom_s = s_double if style_type == 'final' else (s_thin if style_type == 'total' else s_thin)

    cell.border = Border(left=s_thin, right=s_thin, top=top_s, bottom=bottom_s)

    # 4. Fill
    if fill:
        cell.fill = fill
    elif style_type == 'header':
        cell.fill = GREEN_FILL


def _copy_template_row(ws, source_row, target_row, start_col=2, end_col=17, include_values=True):
    if source_row == target_row:
        return

    _clear_merged_ranges_in_region(ws, target_row, target_row, start_col, end_col)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(start_col, end_col + 1):
        target_cell = ws.cell(row=target_row, column=col)
        if isinstance(target_cell, MergedCell):
            continue
        target_cell.value = None

    for col in range(start_col, end_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        if isinstance(source_cell, MergedCell) or isinstance(target_cell, MergedCell):
            continue
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.value = copy(source_cell.value) if include_values else None

    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row != source_row or max_row != source_row:
            continue
        if min_col < start_col or max_col > end_col:
            continue
        try:
            ws.merge_cells(
                start_row=target_row,
                start_column=min_col,
                end_row=target_row,
                end_column=max_col,
            )
        except Exception:
            pass


def _sum_sheet_column_formula(sheet_ref, column_letter, start_row, end_row):
    return f'=SUM({sheet_ref}!{column_letter}{start_row}:{column_letter}{end_row})'


def _to_float(value, default=0.0):
    """Converts a value to float safely."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def _add_image_to_sheet(ws, img_path, anchor, width=None, height=None):
    """Safely add an image to an openpyxl worksheet."""
    if not os.path.exists(img_path):
        logger.warning(f"Image not found for Excel: {img_path}")
        return
    try:
        img = OpenpyxlImage(img_path)
        if width: img.width = width
        if height: img.height = height
        ws.add_image(img, anchor)
    except Exception as e:
        logger.error(f"Failed to add image {img_path} to Excel: {e}")


def _apply_autofit_sheet1(ws, last_col_idx):
    """
    Logika AutoFit Dinamis (Hanya Sheet 1):
    - Kolom A-Q: DIKUNCI MATI (Fixed) sesuai daftar MIN_WIDTHS agar presisi.
    - Kolom > Q (Kategori tambahan): Baru menggunakan AutoFit dinamis.
    """
    # Pastikan kita memproses minimal sampai kolom Q (17) untuk Table 1 & 2
    max_process_col = max(last_col_idx, 17)

    for col_idx in range(1, max_process_col + 1):
        col_letter = get_column_letter(col_idx)

        # 1. JIKA KOLOM TERDAFTAR DI MIN_WIDTHS (A-Q), PAKSA NILAINYA (FIXED)
        if col_letter in MIN_WIDTHS:
            ws.column_dimensions[col_letter].width = MIN_WIDTHS[col_letter]
            continue

        # 2. UNTUK KOLOM DINAMIS (> Q), BARU GUNAKAN LOGIKA AUTOFIT
        max_length = 0
        # Scan baris 6 sampai max_row
        for row in range(6, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)

            # Skip if cell value is None or it's a merged cell (but not the top-left one)
            if cell.value is None or isinstance(cell, MergedCell):
                continue

            # Abaikan teks vertikal (seperti di Tabel 3 header Source, Currency, Rate)
            if cell.alignment and cell.alignment.text_rotation == 90:
                continue

            # Penanganan tanggal (visual format 11-Jan-24)
            if isinstance(cell.value, (datetime, date)):
                length = 11
            else:
                length = len(str(cell.value))

            if length > max_length:
                max_length = length

        # Hitung lebar akhir dengan 18.0 sebagai batas bawah untuk kategori tambahan
        ws.column_dimensions[col_letter].width = max(18.0, max_length * 1.1)

    # Kolom A selalu pake MIN_WIDTHS
    if 'A' in MIN_WIDTHS:
        ws.column_dimensions['A'].width = MIN_WIDTHS['A']

def _is_true(value):
    if value is None:
        return False
    return str(value).strip().lower() in ('1', 'true', 'yes', 'y')


def _set_dividend_row_border(ws, row, col_b='B', col_c='C', col_d='D', col_e='E'):
    """
    ✅ Helper function to set border for dividend row with NO vertical line between D and E
    Reusable for header, data rows, and last row
    """
    # Column B: full border
    ws.cell(row=row, column=2).border = THIN_BORDER
    # Column C: full border
    ws.cell(row=row, column=3).border = THIN_BORDER
    # Column D: border without right side
    ws.cell(row=row, column=4).border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style=None),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    # Column E: border without left side
    ws.cell(row=row, column=5).border = Border(
        left=Side(style=None),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )


def _annual_cache_paths(year):
    cache_dir = os.path.abspath(
        os.path.join(current_app.root_path, '..', 'exports', 'annual_cache')
    )
    os.makedirs(cache_dir, exist_ok=True)
    return {
        'json': os.path.join(cache_dir, f'annual_{year}.json'),
        'pdf': os.path.join(cache_dir, f'annual_{year}.pdf'),
        'excel': os.path.join(cache_dir, f'annual_{year}.xlsx'),
    }


def _load_annual_payload_cache(year):
    json_path = _annual_cache_paths(year)['json']
    if not os.path.exists(json_path):
        return None
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _save_annual_payload_cache(year, payload):
    data = dict(payload)
    data['cache_generated_at'] = datetime.now().isoformat()
    json_path = _annual_cache_paths(year)['json']
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return data


def _save_annual_pdf_cache(year, pdf_bytes):
    pdf_path = _annual_cache_paths(year)['pdf']
    with open(pdf_path, 'wb') as f:
        f.write(pdf_bytes)
    return pdf_path


def _tagged_ids_for_year(table_name, report_year):
    try:
        rows = db.session.execute(
            sql_text("""SELECT row_id FROM report_entry_tags
                WHERE table_name = :table_name AND report_year = :report_year"""),
            {'table_name': table_name, 'report_year': int(report_year)},
        ).fetchall()
        return [int(r[0]) for r in rows]
    except Exception:
        db.session.rollback()
        return []


def _has_any_report_tags():
    try:
        row = db.session.execute(sql_text("SELECT 1 FROM report_entry_tags LIMIT 1")).first()
        return row is not None
    except Exception:
        db.session.rollback()
        return False


def _compute_dividend_distribution(revenues, expenses, profit_retained, recipient_count):
    revenue_total = sum((r.idr_amount_received or 0) for r in revenues)
    pph23_total = sum((r.pph_23 or 0) for r in revenues)
    total_cost = sum((e.idr_amount or 0) for e in expenses)
    profit_before_tax = revenue_total - total_cost
    tax_corporate = profit_before_tax * 0.22 * 0.5 if profit_before_tax > 0 else 0.0
    profit_after_tax = profit_before_tax - (tax_corporate - pph23_total)
    dividend_distributed = max(profit_after_tax - profit_retained, 0.0)
    dividend_per_person = (
        dividend_distributed / recipient_count if recipient_count > 0 else 0.0
    )
    return {
        'revenue_total': revenue_total,
        'pph23_total': pph23_total,
        'total_cost': total_cost,
        'profit_before_tax': profit_before_tax,
        'tax_corporate': tax_corporate,
        'profit_after_tax': profit_after_tax,
        'profit_retained': profit_retained,
        'dividend_distributed': dividend_distributed,
        'dividend_per_person': dividend_per_person,
    }

def _build_annual_payload_from_db(year: int) -> Dict[str, Any]:
    """Build annual report payload from database."""
    from models import Dividend, DividendSetting, Revenue, Tax

    logger.debug(f'Building payload for year={year}')
    year_start, next_year_start = _year_date_bounds(year)

    has_report_tags = _has_any_report_tags()
    tagged_revenue_ids = _tagged_ids_for_year('revenues', year)
    tagged_tax_ids = _tagged_ids_for_year('taxes', year)

    logger.debug(f'Report tags: has={has_report_tags}, revenue_ids={len(tagged_revenue_ids)}, tax_ids={len(tagged_tax_ids)}')

    def revenue_sort_key(r):
        rtype = str(getattr(r, 'revenue_type', '') or 'pendapatan_langsung').strip().lower()
        weight = 1 if rtype == 'pendapatan_lain_lain' else 0
        return (weight, r.receive_date or r.invoice_date or datetime.min.date(), r.id)

    if has_report_tags:
        if tagged_revenue_ids:
            revenues_raw = Revenue.query.filter(
                Revenue.id.in_(tagged_revenue_ids),
                Revenue.invoice_date >= year_start,
                Revenue.invoice_date < next_year_start,
            ).all()
            # ✅ Filter tagged revenue by year too - handle None invoice_date
            logger.debug(f'Filtered {len(revenues_raw)} tagged revenues for year {year}')
            # ✅ Urutkan berdasarkan tipe (langsung vs lain-lain) dan receive_date
            revenues = sorted(revenues_raw, key=revenue_sort_key)
        else:
            revenues = []
    elif tagged_revenue_ids:
        revenues = Revenue.query.filter(
            Revenue.id.in_(tagged_revenue_ids),
            Revenue.invoice_date >= year_start,
            Revenue.invoice_date < next_year_start,
        ).all()
        # ✅ Filter tagged revenue by year too - handle None invoice_date
        logger.debug(f'Filtered {len(revenues)} tagged revenues for year {year}')
        # ✅ Urutkan berdasarkan tipe dan receive_date
        revenues.sort(key=revenue_sort_key)
    else:
        # ✅ FIX: Use Database filter instead of pulling all rows into memory
        revenues = Revenue.query.filter(
            Revenue.invoice_date >= year_start,
            Revenue.invoice_date < next_year_start,
        ).all()
        revenues.sort(key=revenue_sort_key)

    logger.debug(f'Loaded {len(revenues)} revenues for year {year}')
    if revenues:
        logger.debug(f'Revenue date range: {revenues[0].invoice_date} to {revenues[-1].invoice_date}')

    if has_report_tags:
        if tagged_tax_ids:
            taxes = Tax.query.filter(
                Tax.id.in_(tagged_tax_ids),
                Tax.date >= year_start,
                Tax.date < next_year_start,
            ).all()
            # ✅ Filter tagged tax by year too
            tax_order = {tid: idx for idx, tid in enumerate(tagged_tax_ids)}
            taxes.sort(key=lambda t: tax_order.get(t.id, 10**9))
        else:
            taxes = []
    elif tagged_tax_ids:
        taxes = Tax.query.filter(
            Tax.id.in_(tagged_tax_ids),
            Tax.date >= year_start,
            Tax.date < next_year_start,
        ).all()
        # ✅ Filter tagged tax by year too
        taxes.sort(key=lambda t: t.date)
    else:
        # ✅ FIX: Use Database filter instead of pulling all rows into memory
        taxes = Tax.query.filter(
            Tax.date >= year_start,
            Tax.date < next_year_start,
        ).order_by(Tax.date.asc(), Tax.id.asc()).all()
        logger.debug(f'Filtered {len(taxes)} taxes for year {year}')

    logger.debug(f'Loaded {len(taxes)} taxes for year {year}')
    if taxes:
        logger.debug(f'Tax date range: {taxes[0].date} to {taxes[-1].date}')

    # ✅ ADDED joinedload to avoid N+1 queries during to_dict() loop
    expenses_query = Expense.query.options(
        joinedload(Expense.category),
        joinedload(Expense.subcategories),
        joinedload(Expense.settlement)
    ).join(
        Settlement, Expense.settlement_id == Settlement.id
    ).filter(
        Settlement.status.in_(('approved', 'completed')),
        Expense.date >= year_start,
        Expense.date < next_year_start,
    )
    expenses = expenses_query.order_by(Expense.date.asc()).all()

    logger.debug(f'Loaded {len(expenses)} expenses for year {year}')
    if expenses:
        logger.debug(f'Expense date range: {expenses[0].date} to {expenses[-1].date}')

    dividends = Dividend.query.filter(
        Dividend.date >= year_start,
        Dividend.date < next_year_start,
    ).order_by(Dividend.date.asc(), Dividend.id.asc()).all()
    dividend_setting = DividendSetting.query.filter_by(year=year).first()

    logger.debug(f'Loaded {len(dividends)} dividends for year {year}')

    revenue_data = [r.to_dict() for r in revenues]
    tax_data = [t.to_dict() for t in taxes]

    # ✅ VALIDATE DATA EARLY - before any processing or rendering
    _validate_revenue_data(revenue_data)

    dividend_calc = _compute_dividend_distribution(
        revenues,
        expenses,
        dividend_setting.profit_retained if dividend_setting else 0.0,
        len(dividends),
    )
    dividend_data = []
    for d in dividends:
        row = d.to_dict()
        row['dividend_per_person'] = dividend_calc['dividend_per_person']
        dividend_data.append(row)

    # ✅ Optimization: use joinedload for categories to avoid nested database hits
    all_categories = Category.query.options(joinedload(Category.parent)).all()
    category_by_id = {c.id: c for c in all_categories}
    expense_data = []
    for e in expenses:
        d = e.to_dict()

        # Ambil kode kategori asli (misal: A1, B2)
        full_code = '-'
        curr_cat = category_by_id.get(e.category_id)
        if curr_cat:
            full_code = curr_cat.code or '-'

            # Cari nama & kode kategori akar untuk header tabel (tetap perlu root_name)
            temp_cat = curr_cat
            while temp_cat.parent_id and category_by_id.get(temp_cat.parent_id):
                temp_cat = category_by_id[temp_cat.parent_id]
            root_name = temp_cat.name or '-'
            root_code = temp_cat.code or '-'
        else:
            root_name = '-'
            root_code = '-'

        d['category_name'] = root_name
        d['category_code'] = root_code # ✅ FIX: Gunakan kode root (A, B) bukan kode detail (A1, B1)
        d['subcategory_name'] = e.combined_subcategory_label
        d['settlement_id'] = e.settlement_id
        d['settlement_title'] = _clean_settlement_title(e.settlement.title if e.settlement else '-')
        d['settlement_type'] = e.settlement.settlement_type if e.settlement else 'single'
        d['settlement_description'] = e.settlement.description if e.settlement else ''
        expense_data.append(d)
    ordered_expense_data = []
    for group in _group_annual_expenses(expense_data, year):
        ordered_expense_data.extend(group)
    return {
        'year': year,
        'revenue': {'data': revenue_data, 'total_amount_received': sum(r.idr_amount_received for r in revenues),
                    'total_ppn': sum(r.ppn or 0 for r in revenues), 'total_pph23': sum(r.pph_23 or 0 for r in revenues)},
        'tax': {'data': tax_data, 'total_ppn': sum(t.ppn or 0 for t in taxes), 'total_pph21': sum(t.pph_21 or 0 for t in taxes),
                'total_pph23': sum(t.pph_23 or 0 for t in taxes), 'total_pph26': sum(t.pph_26 or 0 for t in taxes)},
        'dividend': {
            'data': dividend_data,
            'profit_after_tax': dividend_calc['profit_after_tax'],
            'profit_retained': dividend_calc['profit_retained'],
            'total_amount': dividend_calc['dividend_distributed'],
            'total_recipient_count': len(dividends),
            'dividend_per_person': dividend_calc['dividend_per_person'],
            'settings': dividend_setting.to_dict() if dividend_setting else {
                'year': year,
                'profit_retained': 0.0,
                'opening_cash_balance': 0.0,
                'accounts_receivable': 0.0,
                'prepaid_tax_pph23': 0.0,
                'prepaid_expenses': 0.0,
                'other_receivables': 0.0,
                'office_inventory': 0.0,
                'other_assets': 0.0,
                'accounts_payable': 0.0,
                'salary_payable': 0.0,
                'shareholder_payable': 0.0,
                'accrued_expenses': 0.0,
                'share_capital': 0.0,
                'retained_earnings_balance': 0.0,
            },
        },
        'operation_cost': {'data': ordered_expense_data, 'total_expenses': sum((e.idr_amount or 0) for e in expenses)},
        'generated_at': datetime.now().isoformat(),
    }

def _build_annual_pdf_bytes(payload):
    year = payload.get('year', datetime.now().year)
    revenues = payload.get('revenue', {}).get('data', [])
    taxes = payload.get('tax', {}).get('data', [])
    dividends = payload.get('dividend', {}).get('data', [])
    expenses = payload.get('operation_cost', {}).get('data', [])
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []; styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=0, fontSize=14)
    elements.append(Paragraph(f"Project: REVENUE vs OPERATION COST Tahun {year}", title_style))
    elements.append(Paragraph(f"Date: Januari - Desember {year}", styles['Normal']))
    elements.append(Spacer(1, 15))
    # tabel revenue & tax
    elements.append(Paragraph("REVENUE & TAX", styles['Heading3']))
    t1_headers = ['Invoice Date','#','Detail/Description','INVOICE VALUE','Curr','Rate','INVOICE Num','Client','Receive Date','AMT RECEIVED','PPn','PPH 23','TransferFee','Remark']

    # Group revenues by type
    grouped_revenues = _group_revenues_by_type(revenues)

    # Check if we have both groups with data
    has_direct = len(grouped_revenues['pendapatan_langsung']) > 0
    has_other = len(grouped_revenues['pendapatan_lain_lain']) > 0
    use_grouping = has_direct and has_other  # Only group if BOTH types exist

    t1_data = [t1_headers]
    total_inv_value = 0
    total_received = 0
    total_ppn = 0
    total_pph = 0
    row_idx = 1

    if use_grouping:
        # Render with grouping (both types exist)
        for group_name, group_revenues in grouped_revenues.items():
            if not group_revenues:
                continue

            # Add group header row
            t1_data.append([f'=== {group_name.upper().replace("_", " ")} ===', '', '', '', '', '', '', '', '', '', '', '', '', ''])

            # Add revenue items for this group
            for idx, r in enumerate(group_revenues, 1):
                inv_val = _idr_from_currency(r.get('invoice_value'), r.get('currency'), r.get('currency_exchange'))
                amt_rec = _idr_from_currency(r.get('amount_received'), r.get('currency'), r.get('currency_exchange'))
                ppn = _to_float(r.get('ppn'))
                pph23 = _to_float(r.get('pph_23'))
                transfer = _to_float(r.get('transfer_fee'))

                total_inv_value += inv_val
                total_received += amt_rec
                total_ppn += ppn
                total_pph += pph23

                t1_data.append([
                    _as_iso_date(r.get('invoice_date')),
                    str(row_idx),
                    _shorten(r.get('description'), 30),
                    f"{inv_val:,.0f}" if inv_val else '-',
                    _safe_text(r.get('currency')) or 'IDR',
                    f"{_to_float(r.get('currency_exchange'), 1):,.0f}" if r.get('currency_exchange') else '-',
                    _safe_text(r.get('invoice_number')),
                    _safe_text(r.get('client')),
                    _as_iso_date(r.get('receive_date')),
                    f"{amt_rec:,.0f}" if amt_rec else '-',
                    f"{ppn:,.0f}" if ppn else '-',
                    f"{pph23:,.0f}" if pph23 else '-',
                    f"{transfer:,.0f}" if transfer else '-',
                    _safe_text(r.get('remark'))
                ])
                row_idx += 1

            # Add subtotal row for this group
            group_total_inv = sum(_idr_from_currency(r.get('invoice_value'), r.get('currency'), r.get('currency_exchange')) for r in group_revenues)
            group_total_rec = sum(_idr_from_currency(r.get('amount_received'), r.get('currency'), r.get('currency_exchange')) for r in group_revenues)
            group_total_ppn = sum(_to_float(r.get('ppn')) for r in group_revenues)
            group_total_pph = sum(_to_float(r.get('pph_23')) for r in group_revenues)

            t1_data.append([
                f"Subtotal {group_name.replace('_', ' ').title()}",
                '',
                '',
                f"{group_total_inv:,.0f}",
                '',
                '',
                '',
                '',
                '',
                f"{group_total_rec:,.0f}",
                f"{group_total_ppn:,.0f}",
                f"{group_total_pph:,.0f}",
                '',
                ''
            ])

            # Add blank row between groups
            if group_name == 'pendapatan_langsung' and grouped_revenues['pendapatan_lain_lain']:
                t1_data.append(['', '', '', '', '', '', '', '', '', '', '', '', '', ''])
    else:
        # Render WITHOUT grouping (only one type exists) - use original flat structure
        for idx, r in enumerate(revenues, 1):
            inv_val = _idr_from_currency(r.get('invoice_value'), r.get('currency'), r.get('currency_exchange'))
            amt_rec = _idr_from_currency(r.get('amount_received'), r.get('currency'), r.get('currency_exchange'))
            ppn = _to_float(r.get('ppn'))
            pph23 = _to_float(r.get('pph_23'))
            transfer = _to_float(r.get('transfer_fee'))

            total_inv_value += inv_val
            total_received += amt_rec
            total_ppn += ppn
            total_pph += pph23

            t1_data.append([
                _as_iso_date(r.get('invoice_date')),
                str(idx),
                _shorten(r.get('description'), 30),
                f"{inv_val:,.0f}" if inv_val else '-',
                _safe_text(r.get('currency')) or 'IDR',
                f"{_to_float(r.get('currency_exchange'), 1):,.0f}" if r.get('currency_exchange') else '-',
                _safe_text(r.get('invoice_number')),
                _safe_text(r.get('client')),
                _as_iso_date(r.get('receive_date')),
                f"{amt_rec:,.0f}" if amt_rec else '-',
                f"{ppn:,.0f}" if ppn else '-',
                f"{pph23:,.0f}" if pph23 else '-',
                f"{transfer:,.0f}" if transfer else '-',
                _safe_text(r.get('remark'))
            ])

    # Add grand total row
    t1_data.append(["REVENUE (IDR)", "", "", f"{total_inv_value:,.0f}", "", "", "", "", "", f"{total_received:,.0f}", f"{total_ppn:,.0f}", f"{total_pph:,.0f}", "-", "-"])
    t1 = Table(t1_data, colWidths=[60,20,120,70,30,35,60,60,60,70,50,50,50,50])
    t1.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2E7D32')),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(3,1),(5,-1),'RIGHT'),
        ('ALIGN',(9,1),(12,-1),'RIGHT'),('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#E8F5E9')),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]))
    elements.append(t1); elements.append(Spacer(1, 15))
    # tabel tax
    elements.append(Paragraph("PAJAK PENGELUARAN", styles['Heading3']))
    t2_headers = ['Date','#','Detail/Description','Trans Value','Curr','Rate','PPN','PPh 21','PPh 23','PPh 26']
    t2_data = [t2_headers]; sum_ppn = 0; sum_pph21 = 0; sum_pph23 = 0; sum_pph26 = 0
    for idx, t in enumerate(taxes, 1):
        ppn = _to_float(t.get('ppn')); pph21 = _to_float(t.get('pph_21')); pph23 = _to_float(t.get('pph_23')); pph26 = _to_float(t.get('pph_26'))
        trans_val = _idr_from_currency(t.get('transaction_value'), t.get('currency'), t.get('currency_exchange'))
        sum_ppn += ppn; sum_pph21 += pph21; sum_pph23 += pph23; sum_pph26 += pph26
        t2_data.append([_as_iso_date(t.get('date')),str(idx),_shorten(t.get('description'),40),
            f"{trans_val:,.0f}" if trans_val else '-',_safe_text(t.get('currency')) or 'IDR',
            f"{_to_float(t.get('currency_exchange'),1):,.0f}" if t.get('currency_exchange') else '-',
            f"{ppn:,.0f}" if ppn else '-',f"{pph21:,.0f}" if pph21 else '-',f"{pph23:,.0f}" if pph23 else '-',f"{pph26:,.0f}" if pph26 else '-'])
    t2_data.append(["PENERIMAAN NEGARA","","","","","",f"{sum_ppn:,.0f}",f"{sum_pph21:,.0f}",f"{sum_pph23:,.0f}",f"{sum_pph26:,.0f}"])
    t2 = Table(t2_data, colWidths=[60,20,180,70,30,35,60,60,60,60])
    t2.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#D84315')),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(3,1),(-1,-1),'RIGHT'),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#FBE9E7')),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]))
    elements.append(t2); elements.append(Spacer(1, 15))
    if dividends:
        elements.append(Paragraph("DIVIDEN", styles['Heading3']))
        t_div_headers = ['Date', '#', 'Nama Penerima', 'Profit Ditahan', 'Dividen Dibagi', 'Dibagi/Orang']
        t_div_data = [t_div_headers]
        total_dividend = _to_float(payload.get('dividend', {}).get('total_amount'))
        profit_retained = _to_float(payload.get('dividend', {}).get('profit_retained'))
        dividend_per_person = _to_float(payload.get('dividend', {}).get('dividend_per_person'))
        for idx, item in enumerate(dividends, 1):
            t_div_data.append([
                _as_iso_date(item.get('date')),
                str(idx),
                _safe_text(item.get('name')),
                f"{profit_retained:,.0f}" if profit_retained else '-',
                f"{total_dividend:,.0f}" if total_dividend else '-',
                f"{dividend_per_person:,.0f}" if dividend_per_person else '-',
            ])
        t_div_data.append([
            'TOTAL',
            '',
            '',
            f"{profit_retained:,.0f}" if profit_retained else '-',
            f"{total_dividend:,.0f}" if total_dividend else '-',
            f"{dividend_per_person:,.0f}" if dividend_per_person else '-',
        ])
        t_div = Table(t_div_data, colWidths=[60, 20, 180, 90, 90, 90])
        t_div.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6A1B9A')),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(3,1),(-1,-1),'RIGHT'),
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#F3E5F5')),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]))
        elements.append(t_div); elements.append(Spacer(1, 15))
    # tabel expense
    elements.append(Paragraph("PENGELUARAN & OPERATION COST", styles['Heading3']))

    # ✅ Pre-calculate totals per root category to filter zero ones
    all_cats_sync = Category.query.all()
    cat_map = {c.id: c for c in all_cats_sync}
    root_totals = {}
    for e in expenses:
        nominal = _expense_amount_for_display(e)
        r_name, _ = _root_category_info(e.get('category_id'), cat_map)
        if r_name:
            root_totals[r_name] = root_totals.get(r_name, 0.0) + nominal

    root_cats_all = Category.query.filter_by(parent_id=None).order_by(Category.sort_order, Category.id).all()
    root_cats = [c for c in root_cats_all if root_totals.get(c.name, 0.0) > 0]
    if not root_cats and root_cats_all:
        root_cats = [root_cats_all[0]]

    cat_columns = [c.name for c in root_cats]
    cat_names = cat_columns

    t3_headers = ['Date','#','Activity (Desc)','Source','Jumlah (IDR)','Curr','Rate'] + [c[:10]+'.' for c in cat_columns]
    t3_data = [t3_headers]
    cat_totals = [0]*len(cat_columns); grand_total = 0

    # Separate single vs batch
    single_expenses = [e for e in expenses if not _is_batch_settlement(e.get('settlement_type'), e.get('settlement_title'))]
    batch_expenses = [e for e in expenses if _is_batch_settlement(e.get('settlement_type'), e.get('settlement_title'))]

    # Render Single
    if single_expenses:
        single_grouped = _group_expenses_by_subcategory(single_expenses)
        row_idx = 1
        for subcat, items in single_grouped['groups'].items():
            # Subcat header
            t3_data.append(['', '', subcat, '', '', '', ''] + [''] * len(cat_columns))
            for e in items:
                nominal_idr = _expense_amount_for_display(e)
                col_idx = _map_expense_column(e.get('category_name'), cat_names)
                row_cats = ['-']*len(cat_columns); row_cats[col_idx] = f"{nominal_idr:,.0f}" if nominal_idr else '-'
                cat_totals[col_idx] += nominal_idr; grand_total += nominal_idr
                t3_data.append([_as_iso_date(e.get('date')), str(row_idx), _shorten(e.get('description'),30), _shorten(e.get('source'),10),
                    f"{nominal_idr:,.0f}" if nominal_idr else '-', _safe_text(e.get('currency')) or 'IDR',
                    f"{_to_float(e.get('currency_exchange'),1):,.0f}" if e.get('currency_exchange') else '-'] + row_cats)
                row_idx += 1

    # Render Batch
    batch_by_settlement = {}
    for e in batch_expenses:
        key = f"{e.get('settlement_id')}:{e.get('settlement_title')}"
        batch_by_settlement.setdefault(key, []).append(e)

    sorted_keys = sorted(batch_by_settlement.keys(), key=lambda x: (int(x.split(':')[0]) if x.split(':')[0].isdigit() else 0))

    batch_counter = 0
    for key in sorted_keys:
        batch_counter += 1
        items = batch_by_settlement[key]
        title = _clean_settlement_title(items[0].get('settlement_title') or 'Tanpa Settlement')

        # Batch header (Expense#N) - Use Dark Navy Blue for "biru keiteman"
        t3_data.append([f'Expense#{batch_counter}', ':', title, '', '', '', ''] + [''] * len(cat_columns))

        batch_grouped = _group_expenses_by_subcategory(items)
        row_idx = 1
        for subcat, sub_items in batch_grouped['groups'].items():
            t3_data.append(['', '', subcat, '', '', '', ''] + [''] * len(cat_columns))
            for e in sub_items:
                nominal_idr = _expense_amount_for_display(e)
                col_idx = _map_expense_column(e.get('category_name'), cat_names)
                row_cats = ['-']*len(cat_columns); row_cats[col_idx] = f"{nominal_idr:,.0f}" if nominal_idr else '-'
                cat_totals[col_idx] += nominal_idr; grand_total += nominal_idr
                t3_data.append([_as_iso_date(e.get('date')), str(row_idx), _shorten(e.get('description'),30), _shorten(e.get('source'),10),
                    f"{nominal_idr:,.0f}" if nominal_idr else '-', _safe_text(e.get('currency')) or 'IDR',
                    f"{_to_float(e.get('currency_exchange'),1):,.0f}" if e.get('currency_exchange') else '-'] + row_cats)
                row_idx += 1

        t3_data.append(["TOTAL PENGELUARAN","","","",f"{grand_total:,.0f}","",""] + [f"{t:,.0f}" if t > 0 else "-" for t in cat_totals])
        cw = [55,20,110,50,60,25,25] + [45]*len(cat_columns)
        t3 = Table(t3_data, colWidths=cw)

    # Custom styling to apply different background for Expense# headers
    t3_style = [
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0D47A1')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),6),
        ('ALIGN',(4,1),(-1,-1),'RIGHT'),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#E3F2FD')),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')
    ]

    # Loop data to find 'Expense#' rows and subcategory headers to apply styling
    for i, row in enumerate(t3_data):
        if i == 0: continue # Skip main header
        if str(row[0]).startswith('Expense#'):
            t3_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#0D47A1'))) # Dark Navy Blue
            t3_style.append(('TEXTCOLOR', (0, i), (-1, i), colors.whitesmoke))
            t3_style.append(('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'))
        elif row[2] and not row[0] and not row[1] and i < len(t3_data)-1: # Subcat header
            t3_style.append(('FONTNAME', (2, i), (2, i), 'Helvetica-Bold'))
    t3.setStyle(TableStyle(t3_style))
    elements.append(t3); elements.append(Spacer(1, 14))
    elements.append(Paragraph("AREA DISPLAY / IMPORT (Kosong - diisi saat refresh)", styles['Heading3']))
    area_data = [['No','Keterangan','Nilai']]
    for idx in range(1, 6): area_data.append([str(idx),'',''])
    area_table = Table(area_data, colWidths=[30,410,180], rowHeights=[18]+[24]*5)
    area_table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#ECEFF1')),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    elements.append(area_table)
    doc.build(elements); buffer.seek(0)
    return buffer.read()

def _sheet_ref(name: str) -> str:
    return f"'{name}'" if " " in name or "-" in name else name


def _write_secondary_summary_sheets(wb, payload, year, main_sheet_name, expense_total_row=None, revenue_last_row=None):
    revenues = payload.get('revenue', {}).get('data', [])
    taxes = payload.get('tax', {}).get('data', [])
    expenses = payload.get('operation_cost', {}).get('data', [])
    monthly = {m: {'revenue': 0.0, 'tax': 0.0, 'expense': 0.0} for m in range(1, 13)}
    for r in revenues:
        dtv = _parse_iso_date(r.get('invoice_date'))
        if not dtv: continue
        monthly[dtv.month]['revenue'] += _idr_from_currency(r.get('amount_received'), r.get('currency'), r.get('currency_exchange'))
    for t in taxes:
        dtv = _parse_iso_date(t.get('date'))
        if not dtv: continue
        monthly[dtv.month]['tax'] += _to_float(t.get('ppn')) + _to_float(t.get('pph_21')) + _to_float(t.get('pph_23')) + _to_float(t.get('pph_26'))
    for e in expenses:
        dtv = _parse_iso_date(e.get('date'))
        if not dtv: continue
        monthly[dtv.month]['expense'] += _expense_amount_for_display(e)

    # ✅ CENTRALIZED: Use single source of truth for revenue row bounds
    if revenue_last_row is None:
        last_revenue_row, total_revenue_row = _get_revenue_bounds(revenues)
    else:
        last_revenue_row = revenue_last_row
        total_revenue_row = revenue_last_row + 1

    # ✅ NO FALLBACK - expense_total_row is required
    if expense_total_row is None:
        raise ValueError("expense_total_row is required - expense rendering must complete successfully")
    cost_totals_row = expense_total_row

    def _get_or_create(name):
        if name in wb.sheetnames:
            return wb[name]
        return wb.create_sheet(name)
    header_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
    month_labels = ['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Agu','Sep','Okt','Nov','Des']
    main_ref = _sheet_ref(main_sheet_name)
    # sheet 2: laba rugi
    ws_lr = _get_or_create(f'Laba rugi -{year}')
    ws_lr['A1'] = f'Laba Rugi Tahun {year}'; ws_lr['A1'].font = Font(bold=True, size=14); ws_lr.merge_cells('A1:E1')
    ws_lr['A3'] = 'Bulan'; ws_lr['B3'] = 'Revenue (IDR)'; ws_lr['C3'] = 'Tax Out (IDR)'; ws_lr['D3'] = 'Operation Cost (IDR)'; ws_lr['E3'] = 'Laba/Rugi (IDR)'
    for c in 'ABCDE':
        cell = ws_lr[f'{c}3']; cell.font = Font(bold=True); cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    row = 4
    for m in range(1, 13):
        ws_lr.cell(row=row, column=1, value=month_labels[m-1])
        ws_lr.cell(row=row, column=2, value=monthly[m]['revenue'])
        ws_lr.cell(row=row, column=3, value=monthly[m]['tax'])
        ws_lr.cell(row=row, column=4, value=monthly[m]['expense'])
        ws_lr.cell(row=row, column=5, value=f'=B{row}-C{row}-D{row}')
        row += 1
    ws_lr.cell(row=row, column=1, value='TOTAL')
    for c in range(2, 6): ws_lr.cell(row=row, column=c, value=f'=SUM({chr(64+c)}4:{chr(64+c)}{row-1})')
    for c in range(1, 6): ws_lr.cell(row=row, column=c).font = Font(bold=True)
    for col, width in {'A':14,'B':22,'C':20,'D':24,'E':24}.items(): ws_lr.column_dimensions[col].width = width
    # sheet 3: business summary
    ws_bs = _get_or_create('Business Summary')
    _safe_set_cell_with_merge(ws_bs, 1, 1, f'Business Summary {year}')
    ws_bs['A1'].font = Font(bold=True, size=14)
    # Fetch root categories dynamically from DB
    root_cats_all = Category.query.filter_by(parent_id=None).order_by(Category.sort_order).all()

    # ✅ Pre-calculate totals per root category to filter zero ones
    all_cats_sync = Category.query.all()
    cat_map = {c.id: c for c in all_cats_sync}
    root_totals = {}
    for e in expenses:
        nominal = _expense_amount_for_display(e)
        r_name, _ = _root_category_info(e.get('category_id'), cat_map)
        if r_name:
            root_totals[r_name] = root_totals.get(r_name, 0.0) + nominal

    # Filter root_cats to only those with data
    root_cats = [c for c in root_cats_all if root_totals.get(c.name, 0.0) > 0]
    if not root_cats and root_cats_all:
        root_cats = [root_cats_all[0]]

    cat_names = [c.name for c in root_cats]
    last_cat_col_letter = get_column_letter(9 + len(cat_names) - 1)

    # ✅ FIX: Use dynamic category range instead of hardcoded :Q
    rows = [
        ('Total Revenue (IDR)', _sum_sheet_column_formula(main_ref, get_column_letter(COL_AMOUNT_RECEIVED), REVENUE_START_ROW, last_revenue_row)),
        ('Total Tax Out (IDR)', _sum_sheet_column_formula(main_ref, get_column_letter(COL_PPH_23), REVENUE_START_ROW, last_revenue_row)),
        ('Total Operation Cost (IDR)', f'=SUM({main_ref}!I{cost_totals_row}:{last_cat_col_letter}{cost_totals_row})'),
        ('Net Profit/Loss (IDR)', '=B4-B5-B6'),
    ]
    ws_bs['A3'] = 'Keterangan'; ws_bs['B3'] = 'Nilai'
    for c in 'AB':
        cell = ws_bs[f'{c}3']; cell.font = Font(bold=True); cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    rr = 4
    for label, value in rows:
        ws_bs.cell(row=rr, column=1, value=label)
        ws_bs.cell(row=rr, column=2, value=value)
        rr += 1
    ws_bs['A9'] = 'Monthly Snapshot'; ws_bs['A9'].font = Font(bold=True)
    ws_bs['A10'] = 'Bulan'; ws_bs['B10'] = 'Revenue'; ws_bs['C10'] = 'Tax'; ws_bs['D10'] = 'Expense'; ws_bs['E10'] = 'Net'
    for c in 'ABCDE':
        cell = ws_bs[f'{c}10']; cell.font = Font(bold=True); cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    rr = 11
    for m in range(1, 13):
        ws_bs.cell(row=rr, column=1, value=month_labels[m-1])
        ws_bs.cell(row=rr, column=2, value=monthly[m]['revenue'])
        ws_bs.cell(row=rr, column=3, value=monthly[m]['tax'])
        ws_bs.cell(row=rr, column=4, value=monthly[m]['expense'])
        ws_bs.cell(row=rr, column=5, value=monthly[m]['revenue'] - monthly[m]['tax'] - monthly[m]['expense'])
        rr += 1
    for col, width in {'A':36,'B':22,'C':20,'D':24,'E':20}.items(): ws_bs.column_dimensions[col].width = width


def _expense_column_mapping_name(expense, category_by_id_map=None):
    category_name = _safe_text(expense.get('category_name')).strip()
    if category_name:
        return category_name

    category_id = expense.get('category_id')
    if category_id is None:
        return ''

    # Use provided map if available, otherwise fallback to query
    if category_by_id_map:
        category = category_by_id_map.get(category_id)
    else:
        category = Category.query.get(category_id)

    if not category:
        return ''

    while category.parent_id:
        if category_by_id_map:
            parent = category_by_id_map.get(category.parent_id)
        else:
            parent = Category.query.get(category.parent_id)
        if not parent:
            break
        category = parent
    return _safe_text(category.name).strip()


def _expense_amount_for_display(expense):
    nominal_idr = _to_float(expense.get('idr_amount'))
    if nominal_idr != 0:
        return nominal_idr
    return _idr_from_currency(
        expense.get('amount'),
        expense.get('currency'),
        expense.get('currency_exchange'),
    )


def _expense_subcategory_label(expense):
    """
    Extract subcategory from expense.
    Priority:
    1. subcategory_name field from database (set in payload as combined_subcategory_label)
    2. [SubCategory] prefix in description
    3. 'Subcategory: X' in notes
    4. Keyword matching from description (fallback)
    """
    label = ""
    # ✅ PRIORITY 1: Database value (populated as combined_subcategory_label in payload)
    subcategory_name = _safe_text(expense.get('subcategory_name')).strip()
    if subcategory_name:
        label = subcategory_name
    else:
        raw_desc = _safe_text(expense.get('description')).strip()

        # ✅ PRIORITY 2: Check [SubCategory] prefix in description (same as Flutter)
        prefixed = re.match(r'^\[(.*?)\]\s*(.*)$', raw_desc)
        if prefixed:
            prefix = _safe_text(prefixed.group(1)).strip()
            if prefix:
                label = prefix

        if not label:
            # ✅ PRIORITY 3: Check notes for "Subcategory: X" pattern (same as Flutter)
            notes = _safe_text(expense.get('notes')).strip()
            note_match = re.search(r'\bSubcategory:\s*([^|]+)', notes, flags=re.IGNORECASE)
            if note_match:
                note_subcategory = _safe_text(note_match.group(1)).strip()
                if note_subcategory:
                    label = note_subcategory

        if not label:
            # ✅ PRIORITY 4: Keyword matching from description (EXACTLY LIKE FLUTTER)
            desc = raw_desc.lower()

            # Match Flutter's keyword order exactly
            if 'rental tool' in desc: label = 'Rental Tool'
            elif 'sales' in desc: label = 'Sales'
            elif 'gaji' in desc or 'bonus' in desc: label = 'Gaji'
            elif 'pembuatan alat' in desc or 'mesin retort' in desc: label = 'Pembuatan Alat'
            elif 'thr' in desc or 'allowance' in desc: label = 'Allowance'
            elif 'data processing' in desc: label = 'Data Processing'
            elif 'moving slickline' in desc or 'project lampu' in desc: label = 'Project Operation'
            elif 'sampling tool' in desc or 'sparepart' in desc or 'ups biaya import' in desc: label = 'Sparepart'
            elif 'repair esor' in desc: label = 'Maintenance'
            elif 'licence' in desc or 'license' in desc: label = 'Software License'
            elif 'handphone operational' in desc: label = 'Operation'
            elif 'sewa ruangan' in desc or 'virtual office' in desc: label = 'Sewa Ruangan'
            elif 'modal kerja' in desc: label = 'Modal Kerja'
            elif 'team building' in desc: label = 'Team Building'
            elif 'biaya transaksi bank' in desc: label = 'Biaya Bank'

    # ✅ STEP 5: Strip internal suffixes like (A), (Q), (J), etc. for clean output
    # This regex matches a space followed by a bracketed uppercase string at the end of the label
    if label:
        label = re.sub(r'\s*\([A-Z]+\)$', '', label.strip())

    return label or ''

def _group_expenses_by_subcategory(expenses):
    """
    Group expenses by subcategory, same as frontend Flutter logic.
    Returns dict with sorted subcategories (A-Z) and uncategorized list.
    """
    groups = {}
    uncategorized = []

    for e in expenses:
        subcat = _expense_subcategory_label(e).strip()
        if not subcat:
            uncategorized.append(e)
        else:
            if subcat not in groups:
                groups[subcat] = []
            groups[subcat].append(e)

    # ✅ FIX: Sort subcategories - Multi-categories (with comma) FIRST, then alphabetical
    def _subcategory_sort_key(s):
        is_multiple = ',' in s
        seniority = 0 if is_multiple else 1
        return (seniority, s.lower())

    sorted_subcats = sorted(groups.keys(), key=_subcategory_sort_key)

    return {
        'groups': {subcat: groups[subcat] for subcat in sorted_subcats},
        'uncategorized': uncategorized,
    }


def _render_expense_section_from_data(
    ws,
    expenses: List[Dict],
    cat_names: List[str],
    category_by_id_map: Dict,
    year: int,
    start_row: int = 41
) -> int:
    """
    Render expense section from data.
    """
    logger.debug('Starting data-driven expense rendering')
    last_category_col = 9 + len(cat_names) - 1
    # Define a safe maximum column to clear template "ghost" data (up to column 30/AD)
    SAFE_MAX_COL = 30
    actual_last_col = max(last_category_col, SAFE_MAX_COL)

    from openpyxl.styles import PatternFill, Border, Side
    no_fill = PatternFill(fill_type=None)
    no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

    white_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
    green_fill = GREEN_FILL
    blue_fill = BLUE_FILL

    # ✅ STEP 1: Separate batch vs single expenses

    batch_expenses = [
        e for e in expenses
        if _is_batch_settlement(e.get('settlement_type'), e.get('settlement_title'))
    ]
    single_expenses = [
        e for e in expenses
        if not _is_batch_settlement(e.get('settlement_type'), e.get('settlement_title'))
    ]

    row_cursor = start_row
    seq_counter = 1

    # ✅ STEP 2: Render Single Expenses FIRST
    single_expenses.sort(key=lambda e: (
        _extract_imported_row(e.get('notes')) is None,
        _extract_imported_row(e.get('notes')) or 10**9,
        int(e.get('id') or 0),
    ))
    single_grouped = _group_expenses_by_subcategory(single_expenses)

    has_single_data = bool(single_grouped['groups']) or bool(single_grouped['uncategorized'])
    if has_single_data:
        for subcat, items in single_grouped['groups'].items():
            _clone_row_format(ws, start_row, row_cursor, start_col=2, end_col=actual_last_col)
            _clear_range(ws, row_cursor, row_cursor, 2, actual_last_col)
            for col in range(2, last_category_col + 1):
                cell = ws.cell(row=row_cursor, column=col)
                cell.fill = copy(white_fill)
                cell.border = THIN_BORDER
            _safe_set_cell(ws, row_cursor, 4, subcat)
            ws.cell(row=row_cursor, column=4).font = ws.cell(row=row_cursor, column=4).font.copy(bold=True)
            ws.cell(row=row_cursor, column=4).alignment = Alignment(horizontal='left', vertical='center')

            # Clear ghost borders for subcategory header in Single
            for col in range(last_category_col + 1, actual_last_col + 1):
                ws.cell(row=row_cursor, column=col).border = no_border
                ws.cell(row=row_cursor, column=col).fill = no_fill

            row_cursor += 1

            for expense in items:
                _clone_row_format(ws, start_row, row_cursor, start_col=2, end_col=actual_last_col)
                _clear_range(ws, row_cursor, row_cursor, 2, actual_last_col)
                for col in range(2, last_category_col + 1):
                    cell = ws.cell(row=row_cursor, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.fill = copy(white_fill)
                        cell.border = THIN_BORDER
                        cell.font = cell.font.copy(bold=False)

                clean_desc = re.sub(r'^\[.*?\]\s*', '', (expense.get('description') or '').strip()).strip()
                _set_date_with_format(ws, row_cursor, 2, expense.get('date'))
                ws.cell(row=row_cursor, column=2).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row_cursor, column=2).font = Font(size=11, name='Arial Narrow')
                _safe_set_cell(ws, row_cursor, 3, seq_counter)
                ws.cell(row=row_cursor, column=3).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_cursor, column=3).font = Font(size=11, name='Arial Narrow')
                _safe_set_cell(ws, row_cursor, 4, clean_desc or '-')
                ws.cell(row=row_cursor, column=4).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=row_cursor, column=4).font = Font(size=11, name='Arial Narrow')
                _safe_set_cell(ws, row_cursor, 5, expense.get('source') or '-')
                _safe_set_number(ws, row_cursor, 6, _expense_amount_for_display(expense))
                _safe_set_cell(ws, row_cursor, 7, expense.get('currency') or 'IDR')
                _safe_set_number(ws, row_cursor, 8, _to_float(expense.get('currency_exchange'), 1) or 1)
                ws.cell(row=row_cursor, column=7).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_cursor, column=8).alignment = Alignment(horizontal='center', vertical='center')

                root_name, _ = _root_category_info(expense.get('category_id'), category_by_id_map)
                cat_idx = _map_expense_category_index_from_name(root_name, cat_names)
                if cat_idx is not None:
                    _safe_set_number(ws, row_cursor, 9 + cat_idx, _expense_amount_for_display(expense))

                # ✅ CRITICAL: Reset borders for ALL ghost columns in this row
                for col in range(last_category_col + 1, actual_last_col + 1):
                    ws.cell(row=row_cursor, column=col).border = no_border
                    ws.cell(row=row_cursor, column=col).fill = no_fill

                row_cursor += 1
                seq_counter += 1

        for expense in single_grouped['uncategorized']:
            _clone_row_format(ws, start_row, row_cursor, start_col=2, end_col=actual_last_col)
            _clear_range(ws, row_cursor, row_cursor, 2, actual_last_col)
            for col in range(2, last_category_col + 1):
                cell = ws.cell(row=row_cursor, column=col)
                if not isinstance(cell, MergedCell):
                    cell.fill = copy(white_fill)
                    cell.border = THIN_BORDER
                    cell.font = cell.font.copy(bold=False)

            clean_desc = re.sub(r'^\[.*?\]\s*', '', (expense.get('description') or '').strip()).strip()
            _set_date_with_format(ws, row_cursor, 2, expense.get('date'))
            ws.cell(row=row_cursor, column=2).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row_cursor, column=2).font = Font(size=11, name='Arial Narrow')
            _safe_set_cell(ws, row_cursor, 3, seq_counter)
            ws.cell(row=row_cursor, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_cursor, column=3).font = Font(size=11, name='Arial Narrow')
            _safe_set_cell(ws, row_cursor, 4, clean_desc or '-')
            ws.cell(row=row_cursor, column=4).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_cursor, column=4).font = Font(size=11, name='Arial Narrow')
            _safe_set_number(ws, row_cursor, 6, _expense_amount_for_display(expense))
            _safe_set_cell(ws, row_cursor, 7, expense.get('currency') or 'IDR')
            _safe_set_number(ws, row_cursor, 8, _to_float(expense.get('currency_exchange'), 1) or 1)
            ws.cell(row=row_cursor, column=7).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_cursor, column=8).alignment = Alignment(horizontal='center', vertical='center')

            root_name, _ = _root_category_info(expense.get('category_id'), category_by_id_map)
            cat_idx = _map_expense_category_index_from_name(root_name, cat_names)
            if cat_idx is not None:
                _safe_set_number(ws, row_cursor, 9 + cat_idx, _expense_amount_for_display(expense))

            # Clear ghost borders for uncategorized in Single
            for col in range(last_category_col + 1, actual_last_col + 1):
                ws.cell(row=row_cursor, column=col).border = no_border
                ws.cell(row=row_cursor, column=col).fill = no_fill

            row_cursor += 1
            seq_counter += 1
    else:
        _clone_row_format(ws, start_row, row_cursor, start_col=2, end_col=actual_last_col)
        _clear_range(ws, row_cursor, row_cursor, 2, actual_last_col)
        for col in range(2, last_category_col + 1):
            cell = ws.cell(row=row_cursor, column=col)
            if not isinstance(cell, MergedCell):
                cell.border = THIN_BORDER
        _safe_set_cell(ws, row_cursor, 4, 'Belum ada data single pengeluaran')
        cell = ws.cell(row=row_cursor, column=4)
        cell.font = cell.font.copy(italic=True, color='808080')
        cell.alignment = cell.alignment.copy(horizontal='center')
        row_cursor += 1

    # ✅ STEP 3: Render separator (green fill)
    separator_row = row_cursor
    _clear_range(ws, separator_row, separator_row, 2, actual_last_col)
    for col in range(2, last_category_col + 1):
        cell = ws.cell(row=separator_row, column=col)
        cell.fill = green_fill
        cell.border = THIN_BORDER

    # Clear ghost borders for separator row
    for col in range(last_category_col + 1, actual_last_col + 1):
        ws.cell(row=separator_row, column=col).border = no_border
        ws.cell(row=separator_row, column=col).fill = no_fill

    _safe_set_cell(ws, separator_row, 4, 'OPERATION COST AND OFFICE - Expenses Report')
    ws.cell(row=separator_row, column=4).font = Font(bold=True)
    row_cursor += 1

    # ✅ STEP 4: Render Batch Expenses SECOND
    batch_by_settlement = {}
    for e in batch_expenses:
        settlement_id = e.get('settlement_id')
        settlement_title = e.get('settlement_title', 'Tanpa Settlement')
        key = f'{settlement_id}:{settlement_title}'
        if key not in batch_by_settlement:
            batch_by_settlement[key] = []
        batch_by_settlement[key].append(e)

    sorted_batches = sorted(
        batch_by_settlement.items(),
        key=lambda x: (int(x[0].split(':')[0]) if x[0].split(':')[0].isdigit() else 0, x[0])
    )

    if sorted_batches:
        batch_counter = 0
        for settlement_key, batch_items in sorted_batches:
            batch_counter += 1
            settlement_title = _clean_settlement_title(batch_items[0].get('settlement_title') or 'Tanpa Settlement')

            # Batch header (blue fill)
            batch_header_row = row_cursor
            _clear_range(ws, batch_header_row, batch_header_row, 2, actual_last_col)
            header_font = Font(bold=True, size=12, name='Arial Narrow', color='000000')

            for col in range(2, last_category_col + 1):
                cell = ws.cell(row=batch_header_row, column=col)
                cell.fill = blue_fill
                cell.border = THIN_BORDER
                cell.font = header_font
                cell.alignment = Alignment(vertical='center')

            # Clear ghost borders for batch header
            for col in range(last_category_col + 1, actual_last_col + 1):
                ws.cell(row=batch_header_row, column=col).border = no_border
                ws.cell(row=batch_header_row, column=col).fill = no_fill

            _safe_set_cell(ws, batch_header_row, 2, f'Expense#{batch_counter}')
            ws.cell(row=batch_header_row, column=2).alignment = Alignment(horizontal='left', vertical='center')

            # ✅ FIX COLON: Bold, Center, Arial Narrow 12
            _safe_set_cell(ws, batch_header_row, 3, ':')
            ws.cell(row=batch_header_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=batch_header_row, column=3).font = header_font

            _safe_set_cell(ws, batch_header_row, 4, settlement_title)
            ws.cell(row=batch_header_row, column=4).alignment = Alignment(horizontal='left', vertical='center')
            row_cursor += 1

            # Group batch items by subcategory
            batch_grouped = _group_expenses_by_subcategory(batch_items)
            for subcat, items in batch_grouped['groups'].items():
                _clear_range(ws, row_cursor, row_cursor, 2, actual_last_col)
                for col in range(2, last_category_col + 1):
                    cell = ws.cell(row=row_cursor, column=col)
                    cell.fill = copy(white_fill)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical='center')

                _safe_set_cell(ws, row_cursor, 4, subcat)
                ws.cell(row=row_cursor, column=4).font = Font(bold=True)
                ws.cell(row=row_cursor, column=4).alignment = Alignment(horizontal='left', vertical='center')
                # Clear ghost borders
                for col in range(last_category_col + 1, actual_last_col + 1):
                    ws.cell(row=row_cursor, column=col).border = no_border
                    ws.cell(row=row_cursor, column=col).fill = no_fill

                row_cursor += 1

                for expense in items:
                    _clone_row_format(ws, start_row, row_cursor, start_col=2, end_col=actual_last_col)
                    _clear_range(ws, row_cursor, row_cursor, 2, actual_last_col)
                    for col in range(2, last_category_col + 1):
                        cell = ws.cell(row=row_cursor, column=col)
                        if not isinstance(cell, MergedCell):
                            cell.fill = copy(white_fill)
                            cell.border = THIN_BORDER
                            cell.font = cell.font.copy(bold=False)

                    clean_desc = re.sub(r'^\[.*?\]\s*', '', (expense.get('description') or '').strip()).strip()
                    _set_date_with_format(ws, row_cursor, 2, expense.get('date'))
                    ws.cell(row=row_cursor, column=2).alignment = Alignment(horizontal='right', vertical='center')
                    ws.cell(row=row_cursor, column=2).font = Font(size=11, name='Arial Narrow')
                    _safe_set_cell(ws, row_cursor, 3, seq_counter)
                    ws.cell(row=row_cursor, column=3).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_cursor, column=3).font = Font(size=11, name='Arial Narrow')
                    _safe_set_cell(ws, row_cursor, 4, clean_desc or '-')
                    ws.cell(row=row_cursor, column=4).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(row=row_cursor, column=4).font = Font(size=11, name='Arial Narrow')
                    _safe_set_cell(ws, row_cursor, 5, expense.get('source') or '-')
                    _safe_set_number(ws, row_cursor, 6, _expense_amount_for_display(expense))
                    _safe_set_cell(ws, row_cursor, 7, expense.get('currency') or 'IDR')
                    _safe_set_number(ws, row_cursor, 8, _to_float(expense.get('currency_exchange'), 1) or 1)
                    ws.cell(row=row_cursor, column=7).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_cursor, column=8).alignment = Alignment(horizontal='center', vertical='center')

                    root_name, _ = _root_category_info(expense.get('category_id'), category_by_id_map)
                    cat_idx = _map_expense_category_index_from_name(root_name, cat_names)
                    if cat_idx is not None:
                        _safe_set_number(ws, row_cursor, 9 + cat_idx, _expense_amount_for_display(expense))

                    # Clear ghost columns values AND borders
                    for col in range(last_category_col + 1, actual_last_col + 1):
                        cell = ws.cell(row=row_cursor, column=col)
                        cell.value = None
                        cell.border = no_border
                        cell.fill = no_fill

                # Clear ghost borders
                for col in range(last_category_col + 1, actual_last_col + 1):
                    ws.cell(row=row_cursor, column=col).border = no_border
                    ws.cell(row=row_cursor, column=col).fill = no_fill

                row_cursor += 1
                seq_counter += 1

    # ✅ STEP 5: Render TOTAL row
    total_row = row_cursor
    _clear_range(ws, total_row, total_row, 2, actual_last_col)
    merge_range = f'B{total_row}:H{total_row}'
    try:
        ws.unmerge_cells(merge_range)
    except Exception: pass
    ws.merge_cells(merge_range)
    _safe_set_cell(ws, total_row, 2, 'TOTAL')
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right', vertical='center')

    cost_totals = _operation_cost_totals_by_column(expenses, cat_names, category_by_id_map)
    for i, total_val in enumerate(cost_totals):
        col = 9 + i
        # ✅ ROUND to avoid decimals in formula bar
        _safe_set_number(ws, total_row, col, round(total_val))
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).alignment = Alignment(horizontal='right', vertical='center')

    # Clear ghost borders for TOTAL row
    for col in range(last_category_col + 1, actual_last_col + 1):
        ws.cell(row=total_row, column=col).border = no_border
        ws.cell(row=total_row, column=col).fill = no_fill

    for col in range(2, last_category_col + 1):
        cell = ws.cell(row=total_row, column=col)
        if not isinstance(cell, MergedCell):
            cell.border = THIN_BORDER

    for row in range(start_row, total_row + 1):
        for col in range(2, last_category_col + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.border = THIN_BORDER

    _set_rows_hidden(ws, start_row, total_row, False)
    max_row = ws.max_row
    if total_row < max_row:
        _set_rows_hidden(ws, total_row + 1, max_row, True)

    logger.debug('Expense rendering completed successfully')
    return total_row



def _manual_combine_groups_by_table(table_name, year):
    groups = ManualCombineGroup.query.filter_by(
        table_name=table_name,
        report_year=int(year),
    ).order_by(ManualCombineGroup.group_date.asc(), ManualCombineGroup.id.asc()).all()
    return [group.to_dict() for group in groups]


def _set_merged_top_alignment(ws, row_num, col_num):
    cell = ws.cell(row=row_num, column=col_num)
    alignment = copy(cell.alignment)
    alignment.horizontal = 'center'
    alignment.vertical = 'center'
    cell.alignment = alignment


def _clear_merged_ranges_in_region(ws, start_row, end_row, start_col, end_col):
    ranges_to_clear = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        overlaps = not (
            max_row < start_row or min_row > end_row or
            max_col < start_col or min_col > end_col
        )
        if overlaps:
            ranges_to_clear.append(str(merged_range))

    for merge_range in ranges_to_clear:
        try:
            ws.unmerge_cells(merge_range)
        except Exception:
            pass


def _apply_manual_revenue_combine_groups(ws, revenues, combine_groups, start_row=8):
    _clear_merged_ranges_in_region(ws, start_row, start_row + max(len(revenues) - 1, 0), 11, 12)
    if not revenues or not combine_groups:
        return

    row_by_id = {
        int(revenue.get('id') or 0): start_row + idx
        for idx, revenue in enumerate(revenues)
        if revenue.get('id') is not None
    }
    revenue_by_id = {
        int(revenue.get('id') or 0): revenue
        for revenue in revenues
        if revenue.get('id') is not None
    }

    for group in combine_groups:
        group_ids = [int(row_id) for row_id in group.get('row_ids', []) if int(row_id) in row_by_id]
        if len(group_ids) < 2:
            continue

        row_numbers = [row_by_id[row_id] for row_id in group_ids]
        row_start = min(row_numbers)
        row_end = max(row_numbers)
        if row_end <= row_start:
            continue

        total_received = 0.0
        group_date = None
        for row_id in group_ids:
            revenue = revenue_by_id.get(row_id, {})
            total_received += _idr_from_currency(
                revenue.get('amount_received'),
                revenue.get('currency'),
                revenue.get('currency_exchange'),
            )
            if group_date is None:
                group_date = _parse_iso_date(revenue.get('receive_date'))

        for merge_range in (f'K{row_start}:K{row_end}', f'L{row_start}:L{row_end}'):
            try:
                ws.unmerge_cells(merge_range)
            except Exception:
                pass
            ws.merge_cells(merge_range)

        _safe_set_cell(ws, row_start, 11, group_date)
        _safe_set_cell(ws, row_start, 12, total_received)
        _set_merged_top_alignment(ws, row_start, 11)
        _set_merged_top_alignment(ws, row_start, 12)


def _apply_manual_tax_combine_groups(ws, taxes, combine_groups, start_row=27):
    # ✅ FIX: Clear merged ranges EXCEPT D:E (columns 4-5) which are for description
    # Only clear columns B (2) and F-Q (6-16) for grouping
    if taxes and combine_groups:
        # Clear column B (2) for date grouping
        _clear_merged_ranges_in_region(ws, start_row, start_row + max(len(taxes) - 1, 0), 2, 2)
        # Clear columns F-Q (6-16) for value grouping
        _clear_merged_ranges_in_region(ws, start_row, start_row + max(len(taxes) - 1, 0), 6, 16)

    if not taxes or not combine_groups:
        return

    row_by_id = {
        int(item.get('id') or 0): start_row + idx
        for idx, item in enumerate(taxes)
        if item.get('id') is not None
    }
    tax_by_id = {
        int(item.get('id') or 0): item
        for item in taxes
        if item.get('id') is not None
    }

    numeric_columns = {
        6: 'transaction_value',
        9: 'transaction_value',
        10: 'ppn',
        11: 'transaction_value',
        12: 'pph_21',
        13: 'transaction_value',
        14: 'pph_23',
        15: 'transaction_value',
        16: 'pph_26',
    }

    for group in combine_groups:
        group_ids = [int(row_id) for row_id in group.get('row_ids', []) if int(row_id) in row_by_id]
        if len(group_ids) < 2:
            continue

        row_numbers = [row_by_id[row_id] for row_id in group_ids]
        row_start = min(row_numbers)
        row_end = max(row_numbers)
        if row_end <= row_start:
            continue

        merged_date = None
        totals = {col: 0.0 for col in numeric_columns}
        for row_id in group_ids:
            item = tax_by_id.get(row_id, {})
            if merged_date is None:
                merged_date = _parse_iso_date(item.get('date'))
            for col_num, field_name in numeric_columns.items():
                value = _to_float(item.get(field_name))
                if col_num in (9, 11, 13, 15) and value <= 0:
                    continue
                totals[col_num] += value

        merge_range = f'B{row_start}:B{row_end}'
        try:
            ws.unmerge_cells(merge_range)
        except Exception:
            pass
        ws.merge_cells(merge_range)
        _safe_set_cell(ws, row_start, 2, merged_date)
        _set_merged_top_alignment(ws, row_start, 2)

        for col_num, total in totals.items():
            _safe_set_number(ws, row_start, col_num, total)


def _operation_cost_totals_by_column(expenses, cat_names, category_by_id_map=None):
    totals = [0.0] * len(cat_names)
    for expense in expenses:
        nominal_idr = _expense_amount_for_display(expense)
        col_idx = _map_expense_category_index_from_name(
            _expense_column_mapping_name(expense, category_by_id_map),
            cat_names
        )
        if col_idx is not None:
            totals[col_idx] += nominal_idr
    return totals


# ============================================================================
# CONSTANTS - TABLE ROW CONFIGURATION (NO HARDCODED ROWS!)
# ============================================================================
# Table 1: Revenue & Tax
REVENUE_START_ROW = 8
REVENUE_HEADER_ROW = 7
REVENUE_TEMPLATE_END = 22  # Include total row (data max 14 rows + 1 total row)
REVENUE_SEPARATOR_TEMPLATE_ROW = 23

# Table 2: Pajak Pengeluaran (+1 row separator from Table 1)
# Row 23: Separator (visible)
# Row 24: Section title "PAJAK PENGELUARAN"
# Row 25: Header 1 (Date, #, Detail/Description...)
# Row 26: Header 2 (Transaction, PPN, PPh...)
# Row 27+: Data rows
TAX_START_ROW = 27
TAX_HEADER_ROW = 26
TAX_TEMPLATE_END = 37
TAX_TITLE_TEMPLATE_ROW = 24
TAX_HEADER_TOP_TEMPLATE_ROW = 25
TAX_HEADER_BOTTOM_TEMPLATE_ROW = 26
TAX_DATA_TEMPLATE_ROW = 27
TAX_TOTAL_TEMPLATE_ROW = 37

# Table 3: Expenses
EXPENSE_START_ROW = 41
EXPENSE_HEADER_ROW = 40
EXPENSE_TEMPLATE_END = 700  # Template limit
EXPENSE_GAP_TEMPLATE_ROW = 38
EXPENSE_TITLE_TEMPLATE_ROW = 39
EXPENSE_HEADER_TEMPLATE_ROW = 40
EXPENSE_DATA_TEMPLATE_ROW = 41

# ============================================================================
# COLUMN CONSTANTS - DO NOT HARDCODE COLUMN LETTERS
# ============================================================================
# Revenue & Tax table columns (main sheet)
# ⚠️ NOTE: Column E (5) is SKIPPED in both tables per template design!

# Revenue Table Columns
COL_DATE = 2                # B: Invoice Date
COL_SEQ = 3                 # C: #
COL_DESCRIPTION = 4         # D: Detail/Description
# COLUMN E (5) = SKIPPED!
COL_INVOICE_VALUE = 6       # F: INVOICE VALUE
COL_CURRENCY = 7            # G: Currency (IDR)
COL_EXCHANGE_RATE = 8       # H: Currency Exchange
COL_INVOICE_NUMBER = 9      # I: INVOICE Number
COL_CLIENT = 10             # J: Client
COL_RECEIVE_DATE = 11       # K: Receive Date
COL_AMOUNT_RECEIVED = 12    # L: AMOUNT RECEIVED IDR
COL_PPN = 13                # M: PPn 11%
COL_PPH_23 = 14             # N: PPH (Pasal 23) 2%
COL_TRANSFER_FEE = 15       # O: Biaya transfer
COL_REMARK = 16             # P: Remark

# Expense category columns (start at 9, dynamic based on categories)
COL_EXPENSE_CATEGORY_START = 9

# Tax table columns (same structure, E also skipped)
COL_TAX_DATE = 2            # B: Date
COL_TAX_SEQ = 3             # C: #
COL_TAX_DESC = 4            # D: Description
# COLUMN E (5) = SKIPPED!
COL_TAX_TRANS_VALUE = 6     # F: Transaction Value
COL_TAX_CURRENCY = 7        # G: Currency (IDR)
COL_TAX_RATE = 8            # H: Rate
COL_TAX_DPP_PPN = 9         # I: DPP PPN
COL_TAX_VALUE_PPN = 10      # J: PPN Value
COL_TAX_DPP_PPH21 = 11      # K: DPP PPh 21
COL_TAX_VALUE_PPH21 = 12    # L: PPh 21 Value
COL_TAX_DPP_PPH23 = 13      # M: DPP PPh 23
COL_TAX_VALUE_PPH23 = 14    # N: PPh 23 Value
COL_TAX_DPP_PPH26 = 15      # O: DPP PPh 26
COL_TAX_VALUE_PPH26 = 16    # P: PPh 26 Value


def _get_revenue_bounds(revenues):
    """
    ✅ CENTRALIZED: Calculate revenue table row bounds.
    Returns (last_revenue_row, total_revenue_row) tuple.

    Handles empty data case by falling back to template row.
    """
    if revenues:
        last_revenue_row = REVENUE_START_ROW + len(revenues) - 1
        total_revenue_row = last_revenue_row + 1
    else:
        # No data: fall back to template row to ensure formulas still work
        last_revenue_row = REVENUE_TEMPLATE_END
        total_revenue_row = REVENUE_TEMPLATE_END
    return last_revenue_row, total_revenue_row


def _group_revenues_by_type(revenues):
    """
    Group revenues by revenue_type (pendapatan_langsung vs pendapatan_lain_lain).
    Returns dict with 'pendapatan_langsung' and 'pendapatan_lain_lain' keys.
    """
    grouped = {
        'pendapatan_langsung': [],
        'pendapatan_lain_lain': []
    }

    for r in revenues:
        r_type = (r.get('revenue_type') or 'pendapatan_langsung').strip().lower()
        if r_type == 'pendapatan_lain_lain':
            grouped['pendapatan_lain_lain'].append(r)
        else:
            grouped['pendapatan_langsung'].append(r)

    return grouped


def _validate_revenue_data(revenues):
    """✅ VALIDATION LAYER: Prevent type bugs before render"""
    for r in revenues:
        # CRITICAL: transfer_fee must be numeric, not string like "Pemungut"
        transfer_fee = r.get('transfer_fee')
        if isinstance(transfer_fee, str):
            print(f"[VALIDATION ERROR] transfer_fee is string at revenue ID {r.get('id')}: '{transfer_fee}'")
            logger.error(f"Validation error: transfer_fee is string at revenue ID {r.get('id')}")
            # Auto-fix: move string to remark, set transfer_fee to 0
            existing_remark = r.get('remark') or ''
            r['remark'] = f"{existing_remark} {transfer_fee}".strip()
            r['transfer_fee'] = 0
        # Ensure numeric types
        r['transfer_fee'] = _to_float(r.get('transfer_fee'), 0)
        r['ppn'] = _to_float(r.get('ppn'), 0)
        r['pph_23'] = _to_float(r.get('pph_23'), 0)
        r['amount_received'] = _to_float(r.get('amount_received'), 0)


def _sync_formatted_secondary_sheets(wb, payload, year, main_sheet_name, expense_total_row=None, revenue_last_row=None):
    """
    ✅ REDESIGNED: No hardcoded cell references (F22, N22, I750)
    All references are now dynamic based on actual table structure.

    Args:
        expense_total_row: Dynamic row where expense totals are stored (required, no fallback)
    """
    expenses = payload.get('operation_cost', {}).get('data', [])
    revenues = payload.get('revenue', {}).get('data', [])
    annual_settings = payload.get('dividend', {}).get('settings', {}) or {}

    # ✅ VALIDATION ALREADY DONE IN _build_annual_payload_from_db

    # ✅ CENTRALIZED: Use single source of truth for revenue row bounds
    if revenue_last_row is None:
        last_revenue_row, total_revenue_row = _get_revenue_bounds(revenues)
    else:
        last_revenue_row = revenue_last_row
        total_revenue_row = revenue_last_row + 1

    # ✅ NO FALLBACK - expense_total_row is required
    if expense_total_row is None:
        raise ValueError("expense_total_row is required - expense rendering must complete successfully")
    cost_totals_row = expense_total_row
    revenue_total = sum(
        _idr_from_currency(
            row.get('amount_received'),
            row.get('currency'),
            row.get('currency_exchange'),
        )
        for row in revenues
    )
    pph23_total = sum(_to_float(row.get('pph_23')) for row in revenues)

    # Fetch root categories dynamically from DB
    root_cats_all = Category.query.filter_by(parent_id=None).order_by(Category.sort_order).all()

    # ✅ Pre-calculate totals per root category to filter zero ones
    all_cats_sync = Category.query.all()
    category_by_id_map = {c.id: c for c in all_cats_sync}
    root_totals = {}
    for e in expenses:
        nominal = _expense_amount_for_display(e)
        r_name, _ = _root_category_info(e.get('category_id'), category_by_id_map)
        if r_name:
            root_totals[r_name] = root_totals.get(r_name, 0.0) + nominal

    # Filter root_cats to only those with data
    root_cats = [c for c in root_cats_all if root_totals.get(c.name, 0.0) > 0]
    if not root_cats and root_cats_all:
        root_cats = [root_cats_all[0]]

    cat_names = [c.name for c in root_cats]

    cost_totals = _operation_cost_totals_by_column(expenses, cat_names, category_by_id_map)
    total_cost = sum(cost_totals)

    profit_before_tax = revenue_total - total_cost
    tax_corporate = max(0.0, profit_before_tax * 0.11) if profit_before_tax > 0 else 0.0
    after_tax = _to_float(payload.get('dividend', {}).get('profit_after_tax'))
    if after_tax == 0 and profit_before_tax > 0:
        after_tax = profit_before_tax - max(0.0, tax_corporate - pph23_total)

    opening_cash_balance = _to_float(annual_settings.get('opening_cash_balance'))
    accounts_receivable = _to_float(annual_settings.get('accounts_receivable'))
    prepaid_tax_pph23 = _to_float(annual_settings.get('prepaid_tax_pph23'))
    prepaid_expenses = _to_float(annual_settings.get('prepaid_expenses'))
    other_receivables = _to_float(annual_settings.get('other_receivables'))
    office_inventory = _to_float(annual_settings.get('office_inventory'))
    other_assets = _to_float(annual_settings.get('other_assets'))
    accounts_payable = _to_float(annual_settings.get('accounts_payable'))
    salary_payable = _to_float(annual_settings.get('salary_payable'))
    shareholder_payable = _to_float(annual_settings.get('shareholder_payable'))
    accrued_expenses = _to_float(annual_settings.get('accrued_expenses'))
    share_capital = _to_float(annual_settings.get('share_capital'))
    retained_earnings_balance = _to_float(annual_settings.get('retained_earnings_balance'))

    main_ref = _sheet_ref(main_sheet_name)

    # ✅ MAIN SHEET: Don't overwrite cells - let Excel formulas work
    # Revenue total and PPh23 total are already calculated by Excel formulas in Table 1
    # Cost totals are written to template row for Laba Rugi references
    if main_sheet_name in wb.sheetnames:
        ws_main = wb[main_sheet_name]
        # ✅ Write cost totals to dynamic row (from render result)
        for idx, value in enumerate(cost_totals, start=9):
            ws_main.cell(row=cost_totals_row, column=idx, value=value)

    lr_name = f'Laba rugi -{year}'
    if lr_name in wb.sheetnames:
        ws_lr = wb[lr_name]

        # ✅ DISABLE GRIDLINES for a clean look
        ws_lr.sheet_view.showGridLines = False

        # ✅ FIX COLUMN WIDTHS to match Original Proportions (Gambar 1)
        ws_lr.column_dimensions['B'].width = 20
        ws_lr.column_dimensions['C'].width = 30
        ws_lr.column_dimensions['D'].width = 5
        ws_lr.column_dimensions['E'].width = 15
        ws_lr.column_dimensions['F'].width = 2  # ✅ NARROW GAP between tables
        ws_lr.column_dimensions['G'].width = 20
        ws_lr.column_dimensions['H'].width = 30
        ws_lr.column_dimensions['I'].width = 5
        ws_lr.column_dimensions['J'].width = 15

        # ✅ REPLACE YEAR IN HEADERS (B4 and F4 in template)
        _replace_text_in_sheet(ws_lr, '2024', str(year))

        # ✅ CRITICAL: PRE-CLEAR ONLY LABA RUGI SIDE
        # We only clear columns A to E (1-5) to preserve NERACA on the right (G-J)
        max_lr_row = max(ws_lr.max_row, 150)
        _clear_range_force(ws_lr, 8, max_lr_row, 1, 5, reset_style=True)

        # Surgically clear old signatures area on BOTH sides to prevent duplicates
        # Clear more columns (1 to 8) to ensure stray "Direktur" from template are removed
        _clear_range_force(ws_lr, 48, max_lr_row, 1, 8, reset_style=True)
        # Helper to apply clean border style to a row (B to E)
        def _apply_pl_row_style(r, style_type='data'):
            s_thin = Side(style='thin')
            s_none = Side(style=None)
            s_double = Side(style='double') if style_type == 'final' else s_none # Or double if needed

            # ✅ FIX: Set standard row height
            ws_lr.row_dimensions[r].height = 15

            # Determine Top/Bottom borders based on style
            top_s = s_thin if style_type in ('total', 'final', 'header') else s_none
            bottom_s = s_thin if style_type == 'total' else (s_double if style_type == 'final' else s_none)

            # ✅ UNIFIED BOX LOGIC: Tambahkan garis vertikal antara Deskripsi dan simbol Rp
            # Laba Rugi Side (B-E)
            ws_lr.cell(row=r, column=2).border = Border(left=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=3).border = Border(right=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=4).border = Border(left=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=5).border = Border(right=s_thin, top=top_s, bottom=bottom_s)

            # Neraca Side (G-J)
            ws_lr.cell(row=r, column=7).border = Border(left=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=8).border = Border(right=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=9).border = Border(left=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=10).border = Border(right=s_thin, top=top_s, bottom=bottom_s)

            # Alignment for Laba Rugi
            ws_lr.cell(row=r, column=2).alignment = Alignment(horizontal='left', vertical='center')
            ws_lr.cell(row=r, column=3).alignment = Alignment(horizontal='left', vertical='center')
            ws_lr.cell(row=r, column=4).alignment = Alignment(horizontal='right', vertical='center') # Rp closer to number
            ws_lr.cell(row=r, column=5).alignment = Alignment(horizontal='right', vertical='center')

            # Alignment for Neraca
            ws_lr.cell(row=r, column=7).alignment = Alignment(horizontal='left', vertical='center')
            ws_lr.cell(row=r, column=8).alignment = Alignment(horizontal='left', vertical='center')
            ws_lr.cell(row=r, column=9).alignment = Alignment(horizontal='right', vertical='center') # Rp closer to number
            ws_lr.cell(row=r, column=10).alignment = Alignment(horizontal='right', vertical='center')

            # Font - Header and Totals are Bold
            is_bold = style_type in ('header', 'total', 'final')
            font = Font(bold=is_bold, size=11, name='Arial Narrow')
            for col in list(range(2, 6)) + list(range(7, 11)):
                ws_lr.cell(row=r, column=col).font = font
                ws_lr.cell(row=r, column=col).fill = PatternFill(fill_type=None)

            # Number Format for Value Columns (E and J)
            ws_lr.cell(row=r, column=5).number_format = '#,##0'
            ws_lr.cell(row=r, column=10).number_format = '#,##0'

        # ✅ EXACT ORIGINAL PROPORTIONS: Match Gambar 2 (from 2024 backup template)
        # Row Heights for the header gap
        ws_lr.row_dimensions[1].height = 15.0
        ws_lr.row_dimensions[2].height = 37.25 # Logo area
        ws_lr.row_dimensions[3].height = 33.6  # PT. EXSPAN header
        ws_lr.row_dimensions[4].height = 20.45 # Title row
        ws_lr.row_dimensions[5].height = 6.0   # Small gap before table

        # Column Widths
        ws_lr.column_dimensions['A'].width = 1.13
        ws_lr.column_dimensions['B'].width = 5.53  # Indent col
        ws_lr.column_dimensions['C'].width = 38.33 # Label col
        ws_lr.column_dimensions['D'].width = 3.66  # Rp col
        ws_lr.column_dimensions['E'].width = 30.13 # Amount col
        ws_lr.column_dimensions['F'].width = 1.13  # Middle spacer
        ws_lr.column_dimensions['G'].width = 5.53  # Indent col
        ws_lr.column_dimensions['H'].width = 38.33 # Label col
        ws_lr.column_dimensions['I'].width = 3.66  # Rp col
        ws_lr.column_dimensions['J'].width = 30.13 # Amount col

        # Just clean up old metadata values if they somehow persist
        ws_lr['A3'] = None
        ws_lr['F3'] = None

        # bersihkan nilai template dummy spesifik di neraca
        if ws_lr['J10'].value in (161401093, '=161401093', 161401093.0):
            ws_lr['J10'] = 0

        # ✅ REVENUE TOTALS BY TYPE - Hitung dengan SUM langsung ke baris sesuai tipe
        direct_count = sum(1 for r in revenues if (r.get('revenue_type') or 'pendapatan_langsung').strip().lower() != 'pendapatan_lain_lain')
        other_count = len(revenues) - direct_count
        col_letter = get_column_letter(COL_AMOUNT_RECEIVED)
        start_row_ref = REVENUE_START_ROW

        # PENDAPATAN SECTION
        ws_lr['B8'] = 'PENDAPATAN'; ws_lr['B8'].font = Font(bold=True)
        _apply_pl_row_style(8, 'header')

        ws_lr['C9'] = 'PENDAPATAN LANGSUNG'
        ws_lr['D9'] = 'Rp'
        if direct_count > 0:
            ws_lr['E9'] = f'=SUM({main_ref}!{col_letter}{start_row_ref}:{col_letter}{start_row_ref + direct_count - 1})'
        else:
            ws_lr['E9'] = 0
        _apply_pl_row_style(9, 'data')

        ws_lr['C10'] = 'PENDAPATAN LAIN LAIN'
        ws_lr['D10'] = 'Rp'
        if other_count > 0:
            other_start = start_row_ref + direct_count
            ws_lr['E10'] = f'=SUM({main_ref}!{col_letter}{other_start}:{col_letter}{other_start + other_count - 1})'
        else:
            ws_lr['E10'] = 0
        _apply_pl_row_style(10, 'data')

        _apply_pl_row_style(11, 'data') # Gap baris 11

        ws_lr['B12'] = 'TOTAL PENDAPATAN'; ws_lr['B12'].font = Font(bold=True)
        ws_lr['D12'] = 'Rp'
        ws_lr['E12'] = '=SUM(E9:E11)'
        _apply_pl_row_style(12, 'total')

        _apply_pl_row_style(13, 'data') # Gap baris 13

        # ✅ Expense totals: Dynamic Generation based on Category.main_group
        langsung_cats = []
        admin_cats = []

        # We need the Category objects to check main_group
        category_obj_by_name = {c.name: c for c in root_cats}

        for idx, cat_name in enumerate(cat_names):
            col_letter = get_column_letter(9 + idx)
            cat_obj = category_obj_by_name.get(cat_name)

            # If main_group is 'BEBAN LANGSUNG', or if it's not set but matches keywords (fallback)
            is_direct = False
            if cat_obj and cat_obj.main_group:
                is_direct = (cat_obj.main_group == 'BEBAN LANGSUNG')
            else:
                # Fallback to keywords if main_group is missing
                direct_keywords = ['operasi', 'research', 'r&d', 'r & d', 'sewa peralatan', 'interpretasi', 'log data', 'project']
                is_direct = any(k in str(cat_name).lower() for k in direct_keywords)

            if is_direct:
                langsung_cats.append((cat_name, col_letter))
            else:
                admin_cats.append((cat_name, col_letter))

        # Clear old Laba Rugi area (A14 - E max_row) - ensure full cleanup of previous dynamic data
        max_lr_row = max(ws_lr.max_row, 150) # Increased buffer for safety
        # ✅ FIX: Use _clear_range_force to handle merged cells (e.g. from previous runs or template)
        # Clear columns A to E (1 to 5) to ensure Column A is also cleaned
        _clear_range_force(ws_lr, 14, max_lr_row, 1, 5, reset_style=True)

        row = 14
        ws_lr[f'B{row}'] = 'BEBAN LANGSUNG'; ws_lr[f'B{row}'].font = Font(bold=True)
        _apply_pl_row_style(row, 'header')
        row += 1

        start_bl = row
        for name, col_l in langsung_cats:
            ws_lr[f'C{row}'] = f'{name.upper()}'
            ws_lr[f'D{row}'] = 'Rp'
            ws_lr[f'E{row}'] = f'={main_ref}!{col_l}{cost_totals_row}'
            _apply_pl_row_style(row, 'data')
            row += 1
        end_bl = row - 1

        if end_bl < start_bl: end_bl = start_bl

        ws_lr[f'B{row}'] = 'TOTAL BIAYA LANGSUNG'; ws_lr[f'B{row}'].font = Font(bold=True)
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = f'=SUM(E{start_bl}:E{end_bl})'
        _apply_pl_row_style(row, 'total')
        tot_bl_row = row
        row += 1
        _apply_pl_row_style(row, 'data') # Gap row bergaris vertikal
        row += 1

        ws_lr[f'B{row}'] = 'BEBAN TIDAK LANGSUNG'; ws_lr[f'B{row}'].font = Font(bold=True)
        _apply_pl_row_style(row, 'header')
        row += 1
        start_btl = row
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = 0
        _apply_pl_row_style(row, 'data')
        end_btl = row

        row += 1
        ws_lr[f'B{row}'] = 'TOTAL BIAYA TIDAK LANGSUNG'; ws_lr[f'B{row}'].font = Font(bold=True)
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = f'=SUM(E{start_btl}:E{end_btl})'
        _apply_pl_row_style(row, 'total')
        tot_btl_row = row
        row += 1
        _apply_pl_row_style(row, 'data') # Gap row bergaris vertikal
        row += 1

        ws_lr[f'B{row}'] = 'BIAYA ADMINISTRASI DAN UMUM'; ws_lr[f'B{row}'].font = Font(bold=True)
        _apply_pl_row_style(row, 'header')
        row += 1
        start_ba = row
        for name, col_l in admin_cats:
            ws_lr[f'C{row}'] = f'{name.upper()}'
            ws_lr[f'D{row}'] = 'Rp'
            ws_lr[f'E{row}'] = f'={main_ref}!{col_l}{cost_totals_row}'
            _apply_pl_row_style(row, 'data')
            row += 1
        end_ba = row - 1

        if end_ba < start_ba: end_ba = start_ba

        ws_lr[f'B{row}'] = 'TOTAL BIAYA ADMINISTRASI DAN UMUM'; ws_lr[f'B{row}'].font = Font(bold=True)
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = f'=SUM(E{start_ba}:E{end_ba})'
        _apply_pl_row_style(row, 'total')
        tot_ba_row = row
        row += 1
        _apply_pl_row_style(row, 'data') # Gap row bergaris vertikal
        row += 1

        ws_lr[f'B{row}'] = 'TOTAL BIAYA'; ws_lr[f'B{row}'].font = Font(bold=True)
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = f'=E{tot_bl_row}+E{tot_btl_row}+E{tot_ba_row}'
        _apply_pl_row_style(row, 'total')
        tot_biaya_row = row
        row += 1
        _apply_pl_row_style(row, 'data') # Gap row bergaris vertikal
        row += 1

        ws_lr[f'B{row}'] = 'PENDAPATAN/BEBAN LAIN LAIN'; ws_lr[f'B{row}'].font = Font(bold=True)
        _apply_pl_row_style(row, 'header')
        row += 1
        ws_lr[f'C{row}'] = 'PENDAPATAN LAIN LAIN'
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = 0
        _apply_pl_row_style(row, 'data')
        pend_lain_row = row
        row += 1
        _apply_pl_row_style(row, 'data') # Gap row bergaris vertikal
        row += 1

        ws_lr[f'B{row}'] = 'LABA RUGI SEBELUM PAJAK'; ws_lr[f'B{row}'].font = Font(bold=True)
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = f'=E12-E{tot_biaya_row}'
        _apply_pl_row_style(row, 'total')
        laba_sblum_row = row
        row += 1
        _apply_pl_row_style(row, 'data') # Gap row bergaris vertikal
        row += 1

        ws_lr[f'B{row}'] = 'CORPORATE TAX (PPh 29)'; ws_lr[f'B{row}'].font = Font(bold=True)
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = 0
        _apply_pl_row_style(row, 'total')
        tax_row = row
        row += 1
        _apply_pl_row_style(row, 'data') # Gap row bergaris vertikal
        row += 1

        ws_lr[f'B{row}'] = 'LABA / RUGI BERSIH'; ws_lr[f'B{row}'].font = Font(bold=True)
        ws_lr[f'D{row}'] = 'Rp'
        ws_lr[f'E{row}'] = f'=E{laba_sblum_row}-E{tax_row}'
        _apply_pl_row_style(row, 'final')
        laba_bersih_row = row

        # ✅ Track last Laba Rugi row
        final_lr_row = row

        # ✅ Tentukan posisi tanda tangan: Jeda 3 baris dari tabel terpanjang
        final_neraca_row = 46
        gap_size = 3
        sig_name_row = max(final_lr_row, final_neraca_row) + gap_size
        sig_title_row = sig_name_row + 1

        # ✅ BERSIHKAN AREA GAP (Hapus garis gantung)
        # Scan dari baris setelah tabel sampai baris tanda tangan
        start_cleanup = min(final_lr_row, final_neraca_row) + 1
        _clear_range_force(ws_lr, start_cleanup, sig_name_row - 1, 1, 10, reset_style=True)

        # Apply data styling and labels to all Neraca rows from 8 to 46
        neraca_content = {
            8: ('AKTIVA', None), 9: ('AKTIVA LANCAR', None),
            10: ('  KAS DAN SETARA KAS TAHUN SEBELUMNYA', 'Rp'),
            11: ('  KAS DAN SETARA KAS TAHUN LAPORAN', 'Rp'),
            12: ('  PIUTANG USAHA', 'Rp'), 13: ('  PAJAK BAYAR DI MUKA (pph23)', 'Rp'),
            14: ('  BIAYA BAYAR DI MUKA', 'Rp'), 15: ('  PIUTANG LAIN LAIN', 'Rp'),
            16: ('TOTAL AKTIVA LANCAR', 'Rp'), 17: ('ACTIVA TETAP', None),
            20: ('  INVENTARIS KANTOR', 'Rp'), 22: ('TOTAL AKTIVA TETAP', 'Rp'),
            24: ('ACTIVA LAIN LAIN', None), 25: ('  TOTAL AKTIVA LAIN LAIN', 'Rp'),
            28: ('TOTAL AKTIVA', 'Rp'), 30: ('HUTANG DAN MODAL', None),
            32: ('HUTANG LANCAR', None), 33: ('  HUTANG USAHA', 'Rp'),
            34: ('  HUTANG GAJI', 'Rp'), 35: ('  HUTANG PEMEGANG SAHAM', 'Rp'),
            36: ('  BIAYA YANG MASIH HARUS DI BAYAR', 'Rp'),
            38: ('TOTAL HUTANG LANCAR', 'Rp'), 40: ('MODAL', None),
            41: ('  MODAL SAHAM', 'Rp'), 42: ('  LABA DI TAHAN', 'Rp'),
            43: ('  TOTAL LABA (RUGI) DITAHAN', 'Rp'), 44: ('TOTAL MODAL', 'Rp'),
            46: ('TOTAL HUTANG DAN MODAL', 'Rp')
        }

        for r in range(8, final_neraca_row + 1):
            s_type = 'header' if r in (8, 17, 24, 30, 32, 40) else ('total' if r in (16, 22, 25, 28, 38, 44) else ('final' if r == 46 else 'data'))

            # Apply labels
            if r in neraca_content:
                lbl, rp_sym = neraca_content[r]
                ws_lr[f'H{r}'] = lbl
                if rp_sym: ws_lr[f'I{r}'] = rp_sym

            # Helper to clear and apply Neraca side borders
            s_thin = Side(style='thin')
            top_s = s_thin if s_type in ('total', 'final', 'header') else Side(style=None)
            bottom_s = s_thin if s_type == 'total' else (Side(style='double') if s_type == 'final' else Side(style=None))

            ws_lr.cell(row=r, column=7).border = Border(left=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=8).border = Border(right=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=9).border = Border(left=s_thin, top=top_s, bottom=bottom_s)
            ws_lr.cell(row=r, column=10).border = Border(right=s_thin, top=top_s, bottom=bottom_s)

            # Standard Neraca Alignment & Font
            font_n = Font(bold=(s_type in ('header', 'total', 'final')), size=11, name='Arial Narrow')
            for c in range(7, 11):
                cell_n = ws_lr.cell(row=r, column=c)
                cell_n.font = font_n
                if c == 7: cell_n.alignment = Alignment(horizontal='left', vertical='center')
                elif c == 8: cell_n.alignment = Alignment(horizontal='left', vertical='center')
                elif c == 9: cell_n.alignment = Alignment(horizontal='right', vertical='center')
                elif c == 10:
                    cell_n.alignment = Alignment(horizontal='right', vertical='center')
                    cell_n.number_format = '#,##0'

        # ✅ ALIGNED BRANDING & MERGED SIGNATURES
        logo_path = os.path.abspath(os.path.join(current_app.root_path, '..', 'frontend', 'assets', 'images', 'sheet1.png'))
        logo_expsan_path = os.path.abspath(os.path.join(current_app.root_path, '..', 'frontend', 'assets', 'images', 'expsan excel.png'))
        profile_path = os.path.abspath(os.path.join(current_app.root_path, '..', 'frontend', 'assets', 'images', 'profil.png'))

        # Top Logos - Menggunakan expsan excel.png dengan ukuran presisi
        # Digeser ke E1 dan J1 agar sejajar dengan kolom nilai (Rp)
        ws_lr._images.clear()
        _add_image_to_sheet(ws_lr, logo_expsan_path, 'E1', width=220, height=65)
        _add_image_to_sheet(ws_lr, logo_expsan_path, 'J1', width=220, height=65)

        # Left Side Signature (Merge B-E and Center)
        ws_lr.merge_cells(f'B{sig_name_row}:E{sig_name_row}')
        ws_lr[f'B{sig_name_row}'] = 'Nama Lengkap'
        ws_lr[f'B{sig_name_row}'].font = Font(bold=True, underline='single')
        ws_lr[f'B{sig_name_row}'].alignment = Alignment(horizontal='center')

        ws_lr.merge_cells(f'B{sig_title_row}:E{sig_title_row}')
        ws_lr[f'B{sig_title_row}'] = 'Direktur'
        ws_lr[f'B{sig_title_row}'].font = Font(bold=True)
        ws_lr[f'B{sig_title_row}'].alignment = Alignment(horizontal='center')

        # Right Side Signature (Merge G-J and Center)
        ws_lr.merge_cells(f'G{sig_name_row}:J{sig_name_row}')
        ws_lr[f'G{sig_name_row}'] = 'Nama Lengkap'
        ws_lr[f'G{sig_name_row}'].font = Font(bold=True, underline='single')
        ws_lr[f'G{sig_name_row}'].alignment = Alignment(horizontal='center')

        ws_lr.merge_cells(f'G{sig_title_row}:J{sig_title_row}')
        ws_lr[f'G{sig_title_row}'] = 'Direktur'
        ws_lr[f'G{sig_title_row}'].font = Font(bold=True)
        ws_lr[f'G{sig_title_row}'].alignment = Alignment(horizontal='center')

        # Bottom Footers - Presisi sesuai screenshot (15.22cm x 1.25cm -> ~575px x 47px)
        footer_row = sig_title_row + 1
        _add_image_to_sheet(ws_lr, profile_path, f'B{footer_row}', width=575, height=47)
        _add_image_to_sheet(ws_lr, profile_path, f'G{footer_row}', width=575, height=47)

        # input dan formula neraca (ROUNDED)
        ws_lr['J10'] = round(opening_cash_balance)
        ws_lr['J11'] = f"=ROUND(E{laba_bersih_row}-'Business Summary'!E13, 0)"
        ws_lr['J12'] = round(accounts_receivable)
        ws_lr['J13'] = round(prepaid_tax_pph23)
        ws_lr['J14'] = round(prepaid_expenses)
        ws_lr['J15'] = round(other_receivables)
        ws_lr['J16'] = '=SUM(J10:J15)'
        ws_lr['J20'] = round(office_inventory)
        ws_lr['J22'] = '=J20'
        ws_lr['J25'] = round(other_assets)
        ws_lr['J26'] = '=SUM(J25:J25)'
        ws_lr['J28'] = '=J26+J22+J16'
        ws_lr['J33'] = round(accounts_payable)
        ws_lr['J34'] = round(salary_payable)
        ws_lr['J35'] = round(shareholder_payable)
        ws_lr['J36'] = round(accrued_expenses)
        ws_lr['J38'] = '=SUM(J33:J36)'
        ws_lr['J41'] = round(share_capital)
        ws_lr['J42'] = round(retained_earnings_balance)
        ws_lr['J43'] = '=J28-J38-J41-J42'
        ws_lr['J44'] = '=SUM(J41:J43)'
        ws_lr['J46'] = '=J44+J38'

    if 'Business Summary' in wb.sheetnames:
        ws_bs = wb['Business Summary']
        ws_bs['A4'] = None  # Clean up mistake
        ws_bs['B4'] = f'BUSINESS SUMMARY | TAHUN {year}'

        # ✅ REPLACE YEAR IN HEADERS
        _replace_text_in_sheet(ws_bs, '2024', str(year))

        # ✅ Clean background: Hide gridlines for a professional look
        ws_bs.sheet_view.showGridLines = False

        # ✅ Set column widths for Sheet 3
        ws_bs.column_dimensions['A'].width = 3.0
        ws_bs.column_dimensions['B'].width = 4.0
        ws_bs.column_dimensions['C'].width = 40.0
        ws_bs.column_dimensions['D'].width = 4.0
        ws_bs.column_dimensions['E'].width = 25.0
        ws_bs.column_dimensions['F'].width = 2.0  # Narrow separator as requested
        ws_bs.column_dimensions['G'].width = 15.0

        # FIX: Set consistent row height for main table (rows 6-16)
        for row_idx in range(6, 17):
            ws_bs.row_dimensions[row_idx].height = ROW_HEIGHT_DIVIDEND

        # ✅ FIX: Define variables to prevent NameErrors
        last_cat_col_letter_sync = get_column_letter(9 + len(cat_names) - 1)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Apply Custom Border to hide vertical line between D and E for rows 7-16 (NO MERGE)
        for r in range(7, 17):
            try:
                ws_bs.unmerge_cells(f'D{r}:E{r}')
            except Exception: pass
            _set_dividend_row_border(ws_bs, r) # Uses helper that hides D-E line
            ws_bs[f'D{r}'] = 'Rp'
            ws_bs[f'D{r}'].alignment = Alignment(horizontal='left')
            ws_bs[f'E{r}'].alignment = Alignment(horizontal='right')

        # Row 7: Item 1 - Revenue
        ws_bs['B7'] = 1; ws_bs['C7'] = 'Revenue'
        ws_bs['E7'] = _sum_sheet_column_formula(main_ref, get_column_letter(COL_AMOUNT_RECEIVED), REVENUE_START_ROW, last_revenue_row)
        ws_bs['E7'].number_format = '#,##0'

        # Row 8: Item 2 - Revenue Diterima (Setelah Pajak)
        ws_bs['B8'] = 2; ws_bs['C8'] = 'Revenue Diterima (Setelah Pajak)'
        ws_bs['E8'] = f'=E7-{_sum_sheet_column_formula(main_ref, get_column_letter(COL_PPH_23), REVENUE_START_ROW, last_revenue_row)[1:]}'
        ws_bs['E8'].number_format = '#,##0'

        # Row 9: Item 3 - Biaya
        ws_bs['B9'] = 3; ws_bs['C9'] = 'Biaya'
        ws_bs['E9'] = f'=SUM({main_ref}!I{cost_totals_row}:{last_cat_col_letter_sync}{cost_totals_row})'
        ws_bs['E9'].number_format = '#,##0'

        # Row 10: Item 4 - Profit SebelumTax
        ws_bs['B10'] = 4; ws_bs['C10'] = 'Profit SebelumTax'
        ws_bs['E10'] = '=E7-E9'
        ws_bs['E10'].number_format = '#,##0'

        # Row 11: Item 5 - PPh Badan (22%*50%)
        ws_bs['B11'] = 5; ws_bs['C11'] = 'PPh Badan (22%*50%)'
        ws_bs['E11'] = '=E10*22%*50%'
        ws_bs['E11'].number_format = '#,##0'

        # Row 12: Item 6 - PPh 23 Dipotong oleh client
        ws_bs['B12'] = 6; ws_bs['C12'] = 'PPh 23 Dipotong oleh client'
        ws_bs['E12'] = _sum_sheet_column_formula(main_ref, get_column_letter(COL_PPH_23), REVENUE_START_ROW, last_revenue_row)
        ws_bs['E12'].number_format = '#,##0'

        # Row 13: Item 7 - Kekurangan Bayar PPh Badan
        ws_bs['B13'] = 7; ws_bs['C13'] = 'Kekurangan Bayar PPh Badan'
        ws_bs['E13'] = '=E11-E12'
        ws_bs['E13'].number_format = '#,##0'
        ws_bs['E13'].fill = yellow_fill

        # Row 14: Item 8 - Profit After Tax
        ws_bs['B14'] = 8; ws_bs['C14'] = 'Profit After Tax'
        ws_bs['E14'] = '=E10-E13'
        ws_bs['E14'].number_format = '#,##0'

        dividends = payload.get('dividend', {}).get('data', [])
        profit_retained = _to_float(payload.get('dividend', {}).get('profit_retained'))

        # Row 15: Item 9 - Deviden Dibagi
        ws_bs['B15'] = 9; ws_bs['C15'] = 'Deviden Dibagi'
        ws_bs['E15'] = '=E14-E16'
        ws_bs['E15'].number_format = '#,##0'

        # Row 16: Item 10 - Profit Ditahan
        ws_bs['B16'] = 10; ws_bs['C16'] = 'Profit Ditahan'
        ws_bs['E16'] = profit_retained
        ws_bs['E16'].number_format = '#,##0'

        # ✅ DEVIDEN SECTION
        ws_bs['B18'] = 'DEVIDEN'
        ws_bs['B18'].font = Font(bold=True, size=14, name='Trebuchet MS')

        # Row 19 Header - REMOVED 'Pajak' text
        ws_bs['B19'] = '#'; ws_bs['C19'] = 'Nama Lengkap'; ws_bs['D19'] = 'Rp'; ws_bs['E19'] = 'Amount'; ws_bs['G19'] = ''
        for col in ('B', 'C', 'D', 'E', 'G'):
            cell = ws_bs[f'{col}19']
            cell.font = Font(bold=True, size=14, name='Trebuchet MS')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Hide middle border for header too
        try: ws_bs.unmerge_cells('D19:E19')
        except Exception: pass
        _set_dividend_row_border(ws_bs, 19)

        # Clear ALL potential dividend data rows (20-80) and REMOVE any existing borders
        for row in range(20, 81):
            for col in ('B', 'C', 'D', 'E', 'F', 'G'):
                try:
                    ws_bs[f'{col}{row}'] = None
                    # Set border to None for all cells to clean up
                    ws_bs[f'{col}{row}'].border = Border()
                except Exception: pass
            try: ws_bs.unmerge_cells(f'D{row}:E{row}')
            except Exception: pass

        # Fill actual dividend data
        num_dividends = len(dividends)
        last_row = 19 + num_dividends if num_dividends > 0 else 19

        for idx, item in enumerate(dividends, 1):
            row = 19 + idx
            ws_bs[f'B{row}'] = idx
            ws_bs[f'C{row}'] = item.get('name', '')
            ws_bs[f'D{row}'] = 'Rp'
            ws_bs[f'E{row}'] = f'=IF(C{row}="","",ROUND($E$15/COUNTA($C$20:$C${19 + num_dividends}), 0))'
            ws_bs[f'G{row}'] = 'Pajak 10%'

            # Styling for data rows
            for col in ('B', 'C', 'D', 'E'):
                cell = ws_bs[f'{col}{row}']
                cell.font = Font(size=14, name='Trebuchet MS')
                if col == 'B': cell.alignment = Alignment(horizontal='center')
                elif col == 'C': cell.alignment = Alignment(horizontal='left')
                elif col == 'D': cell.alignment = Alignment(horizontal='left')
                elif col == 'E': cell.alignment = Alignment(horizontal='right')

            ws_bs[f'E{row}'].number_format = '#,##0'
            # ✅ Apply border ONLY to the actual data row
            _set_dividend_row_border(ws_bs, row)

        # Set consistent row heights
        for r in range(19, last_row + 1):
            ws_bs.row_dimensions[r].height = ROW_HEIGHT_DIVIDEND

        # ✅ ADD BRANDING TO BUSINESS SUMMARY
        logo_path = os.path.abspath(os.path.join(current_app.root_path, '..', 'frontend', 'assets', 'images', 'sheet1.png'))
        logo_expsan_path = os.path.abspath(os.path.join(current_app.root_path, '..', 'frontend', 'assets', 'images', 'expsan excel.png'))
        profile_path = os.path.abspath(os.path.join(current_app.root_path, '..', 'frontend', 'assets', 'images', 'profil.png'))

        # Top Logo - Balanced size (match Sheet 2: 220px x 65px)
        ws_bs._images.clear()
        _add_image_to_sheet(ws_bs, logo_expsan_path, 'D1', width=220, height=65)

        # Footer Profile - Restored "Gap Panjang" (4 rows below last_row to stay at the bottom)
        footer_row = last_row + 4
        _add_image_to_sheet(ws_bs, profile_path, f'B{footer_row}', width=550, height=45)

def _clone_sheet_from_template(src_ws, target_wb, new_name):
    if new_name in target_wb.sheetnames: del target_wb[new_name]
    dst_ws = target_wb.create_sheet(new_name)
    dst_ws.sheet_format = copy(src_ws.sheet_format); dst_ws.sheet_properties = copy(src_ws.sheet_properties)
    dst_ws.page_margins = copy(src_ws.page_margins); dst_ws.page_setup = copy(src_ws.page_setup)
    dst_ws.print_options = copy(src_ws.print_options); dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.auto_filter = copy(src_ws.auto_filter)
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst = dst_ws.column_dimensions[col_letter]; dst.width = col_dim.width; dst.hidden = col_dim.hidden
        dst.outlineLevel = col_dim.outlineLevel; dst.bestFit = col_dim.bestFit
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst = dst_ws.row_dimensions[row_idx]; dst.height = row_dim.height; dst.hidden = row_dim.hidden; dst.outlineLevel = row_dim.outlineLevel
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, min_col=1, max_col=src_ws.max_column):
        for cell in row:
            dst = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst.font = copy(cell.font); dst.fill = copy(cell.fill); dst.border = copy(cell.border)
                dst.alignment = copy(cell.alignment); dst.number_format = copy(cell.number_format); dst.protection = copy(cell.protection)
    for merged in src_ws.merged_cells.ranges: dst_ws.merge_cells(str(merged))
    return dst_ws


def _ensure_formatted_secondary_sheets(wb, year, template_dir):
    target_lr = f'Laba rugi -{year}'; target_bs = 'Business Summary'

    # Hapus semua sheet Laba rugi yang lama (dari template)
    sheets_to_remove = [name for name in wb.sheetnames if name.startswith('Laba rugi -') and name != target_lr]
    for sheet_name in sheets_to_remove:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    if target_lr in wb.sheetnames and target_bs in wb.sheetnames: return True

    donor_candidates = [
        os.path.abspath(os.path.join(template_dir, '20250427_EWI Financial-Repport_2024.xlsx')),
        os.path.abspath(os.path.join(template_dir, '20250427_EWI Financial-Repport_2024_Submitted_Rev1.xlsx')),
        os.path.abspath(os.path.join(template_dir, 'Revenue-Cost_2024_cleaned.xlsx')),
    ]
    donor_path = next((p for p in donor_candidates if os.path.exists(p)), None)
    if not donor_path: return False

    donor_wb = load_workbook(donor_path, keep_links=False)

    # Cari sheet Laba rugi (bisa Laba rugi -2024 atau nama lain)
    donor_lr_name = next((name for name in donor_wb.sheetnames if name.startswith('Laba rugi')), None)
    donor_lr = donor_wb[donor_lr_name] if donor_lr_name else None

    # Cari sheet Business Summary
    donor_bs_name = next((name for name in donor_wb.sheetnames if name == 'Business Summary'), None)
    donor_bs = donor_wb[donor_bs_name] if donor_bs_name else None

    if donor_lr is None or donor_bs is None: return False

    _clone_sheet_from_template(donor_lr, wb, target_lr)
    _clone_sheet_from_template(donor_bs, wb, target_bs)

    # ✅ FIX: Override column widths from template for Business Summary
    # Template has wide columns, we want smaller consistent columns
    if target_bs in wb.sheetnames:
        ws_bs = wb[target_bs]
        ws_bs.column_dimensions['A'].width = 1       # # column - kecil
        ws_bs.column_dimensions['B'].width = 5       # # column - kecil
        ws_bs.column_dimensions['C'].width = 40      # Items / Nama Lengkap (diperlebar)
        ws_bs.column_dimensions['D'].width = 5       # Rp currency
        ws_bs.column_dimensions['E'].width = 25      # Amount (diperlebar)
        ws_bs.column_dimensions['F'].width = 4       # Empty separator
        ws_bs.column_dimensions['G'].width = 10      # Pajak

    return True


@reports_bp.route('/annual', methods=['GET'])
@jwt_required()
def get_annual_report():
    user = User.query.get(int(get_jwt_identity()))
    year = request.args.get('year', _default_report_year(), type=int)
    payload = _build_annual_payload_from_db(year)
    payload['cache_source'] = 'refresh'
    saved_payload = _save_annual_payload_cache(year, payload)
    return jsonify(saved_payload), 200


@reports_bp.route('/annual/pdf', methods=['GET'])
@jwt_required()
def get_annual_report_pdf():
    user = User.query.get(int(get_jwt_identity()))
    year = request.args.get('year', _default_report_year(), type=int)
    filename = f"Laporan_Tahunan_{year}.pdf"
    payload = _save_annual_payload_cache(year, _build_annual_payload_from_db(year))
    pdf_path = _save_annual_pdf_cache(year, _build_annual_pdf_bytes(payload))
    return send_file(pdf_path, as_attachment=True, download_name=filename, mimetype='application/pdf')


@reports_bp.route('/annual/excel', methods=['GET'])
@jwt_required()
def get_annual_report_excel():
    """Generate and export annual report as Excel file."""
    user = User.query.get(int(get_jwt_identity()))

    # Get and validate year parameter
    try:
        year = request.args.get('year', _default_report_year(), type=int)
        current_year = datetime.now().year
        # Allow year range from 2020 to 10 years in future (for planning/projection)
        if not (2020 <= year <= current_year + 10):
            error_msg = f'Year must be between 2020 and {current_year + 10}'
            logger.error(f'Invalid year parameter: {year}')
            return jsonify({'error': error_msg}), 400
    except (TypeError, ValueError) as e:
        logger.error(f'Invalid year parameter: {e}')
        return jsonify({'error': 'Invalid year parameter'}), 400

    logger.info(f'Starting Excel export for year={year}')

    # ✅ STEP 1: Check cache first
    cache_paths = _annual_cache_paths(year)
    excel_cache_path = cache_paths['excel']
    json_cache_path = cache_paths['json']
    force_refresh = request.args.get('refresh', '0') == '1'

    # Auto-refresh if JSON cache is newer than Excel cache
    if os.path.exists(excel_cache_path) and os.path.exists(json_cache_path):
        if os.path.getmtime(json_cache_path) > os.path.getmtime(excel_cache_path):
            logger.info(f'JSON cache is newer than Excel cache for year {year}, forcing refresh')
            force_refresh = True

    if not force_refresh and os.path.exists(excel_cache_path):
        logger.info(f'Serving Excel from cache for year {year}')
        return send_file(
            excel_cache_path,
            as_attachment=True,
            download_name=f"Laporan_Tahunan_{year}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    try:
        # Load data
        payload = _build_annual_payload_from_db(year)
        _save_annual_payload_cache(year, payload)
        revenues = payload.get('revenue', {}).get('data', [])
        taxes = payload.get('tax', {}).get('data', [])
        expenses = payload.get('operation_cost', {}).get('data', [])

        logger.debug(f'Data loaded: {len(revenues)} revenues, {len(taxes)} taxes, {len(expenses)} expenses')

        # Load manual combine groups
        revenue_combine_groups = _manual_combine_groups_by_table('revenues', year)
        tax_combine_groups = _manual_combine_groups_by_table('taxes', year)

        # Get imported source name
        imported_source_name = None
        try:
            imported_source_name = db.session.execute(sql_text(
                """SELECT source_excel FROM report_entry_tags
                   WHERE report_year = :year AND source_excel IS NOT NULL AND TRIM(source_excel) <> ''
                   ORDER BY imported_at DESC, id DESC LIMIT 1"""),
                {'year': int(year)}).scalar()
        except Exception:
            db.session.rollback()
            logger.debug('No imported source name found')

        template_dir = os.path.abspath(os.path.join(current_app.root_path, '..', 'excel'))

        template_candidates = []
        if imported_source_name:
            template_candidates.extend([
                os.path.abspath(os.path.join(template_dir, imported_source_name)),
                os.path.abspath(os.path.join(current_app.root_path, '..', 'data', imported_source_name)),
            ])
        template_candidates.extend([
            os.path.abspath(os.path.join(template_dir, 'Revenue-Cost_2024_cleaned_asli_cleaned.xlsx')),
            os.path.abspath(os.path.join(template_dir, 'Revenue-Cost_2024_cleaned.xlsx')),
        ])
        template_path = next((p for p in template_candidates if os.path.exists(p)), template_candidates[0])

        logger.debug(f'Template path: {template_path}')

        if not os.path.exists(template_path):
            logger.error(f'Template not found: {template_path}')
            return jsonify({'error': f'Template tidak ditemukan: {template_path}'}), 404

        wb = load_workbook(template_path, keep_links=False)
        _normalize_external_formula_refs(wb)

        logger.debug('Template loaded, processing sheets')
        logger.debug('Keeping Excel tables from template for proper Alt+A behavior')

        # Hapus semua sheet Laba rugi yang lama dari template
        sheets_to_remove = [name for name in wb.sheetnames if name.startswith('Laba rugi -')]
        for sheet_name in sheets_to_remove:
            del wb[sheet_name]

        # Cari dan rename sheet Revenue-Cost
        revenue_cost_sheet_name = next((name for name in wb.sheetnames if name.startswith('Revenue-Cost')), None)
        if revenue_cost_sheet_name:
            ws = wb[revenue_cost_sheet_name]
            ws.title = f'Revenue-Cost_{year}'
        else:
            ws = wb.active
            ws.title = f'Revenue-Cost_{year}'

        # ✅ HILANGKAN GRIDLINES agar tampilan bersih putih polos
        ws.sheet_view.showGridLines = False

        # Update header document
        _safe_set_cell(ws, 2, 4, f'REVENUE vs OPERATION COST Tahun {year}')
        _safe_set_cell(ws, 3, 4, f'Januari - Desember {year}')

        # VALIDATE DATA BEFORE RENDER
        _validate_revenue_data(revenues)

        # MERGE D & E for header "Detail/Description"
        header_desc_row = REVENUE_HEADER_ROW - 1
        try:
            ws.merge_cells(f'D{header_desc_row}:E{header_desc_row}')
        except Exception as e:
            logger.warning(f'Failed to merge header: {e}')

        # Set Q7 as example for Remark column
        try:
            ws.cell(row=7, column=17).value = 'con: inc.PPN11%'
            ws.cell(row=7, column=17).font = Font(italic=True, color='808080')
            ws.cell(row=7, column=17).alignment = Alignment(horizontal='center')
        except Exception as e:
            logger.warning(f'Failed to set Q7 example: {e}')

        # Clear data area - PRESERVE STYLE (Borders) for Sheet 1
        _clear_range_force(ws, REVENUE_START_ROW, REVENUE_TEMPLATE_END, 2, 17, reset_style=False)
        _clear_range_force(ws, REVENUE_TEMPLATE_END + 1, REVENUE_TEMPLATE_END + 1, 2, 17, reset_style=False)

        # Render revenue data
        row_cursor = REVENUE_START_ROW
        has_revenue_data = len(revenues) > 0
    except Exception as e:
        logger.error(f'Error during initial setup: {e}')
        raise

    if has_revenue_data:
        for idx, r in enumerate(revenues, 1):
            _clone_row_format(ws, REVENUE_START_ROW, row_cursor, start_col=2, end_col=17)

            invoice_value = _to_float(r.get('invoice_value'))
            currency = r.get('currency') or 'IDR'
            exchange = _to_float(r.get('currency_exchange'), 1) or 1
            amount_received = _to_float(r.get('amount_received'))
            p_ppn = _to_float(r.get('ppn'))
            p_pph23 = _to_float(r.get('pph_23'))
            transfer_fee = _to_float(r.get('transfer_fee'), 0)

            remark_from_db = _safe_text(r.get('remark'))
            if 'inc.PPN11%' in remark_from_db or 'inc. PPN11%' in remark_from_db:
                remark = ''
            else:
                remark = remark_from_db

            _set_date_with_format(ws, row_cursor, COL_DATE, r.get('invoice_date'))
            _safe_set_cell(ws, row_cursor, COL_SEQ, idx)

            description = r.get('description') or ''
            source = r.get('source') or ''
            combined_desc = f'{description} ({source})' if source and source.strip() else description
            _merge_description_cell(ws, row_cursor, combined_desc, col_start=4, col_end=5)

            _safe_set_number(ws, row_cursor, COL_INVOICE_VALUE, invoice_value)
            _safe_set_cell(ws, row_cursor, COL_CURRENCY, currency)
            _safe_set_number(ws, row_cursor, COL_EXCHANGE_RATE, exchange)

            invoice_num = r.get('invoice_number')
            if invoice_num and str(invoice_num).strip():
                _safe_set_cell(ws, row_cursor, COL_INVOICE_NUMBER, str(invoice_num))
            else:
                _safe_set_cell(ws, row_cursor, COL_INVOICE_NUMBER, '')

            _safe_set_cell(ws, row_cursor, COL_CLIENT, r.get('client') or '')
            _set_date_with_format(ws, row_cursor, COL_RECEIVE_DATE, r.get('receive_date'))
            _safe_set_number(ws, row_cursor, COL_AMOUNT_RECEIVED, amount_received)
            _safe_set_number(ws, row_cursor, COL_PPN, p_ppn)
            _safe_set_number(ws, row_cursor, COL_PPH_23, p_pph23)
            _safe_set_number(ws, row_cursor, COL_TRANSFER_FEE, transfer_fee)
            _safe_set_cell(ws, row_cursor, COL_REMARK, remark)
            row_cursor += 1
    else:
        # ✅ No revenue data - show empty state message at row 8
        # ✅ FIX: Clone style FIRST, then clear and write
        _clone_row_format(ws, REVENUE_START_ROW, row_cursor, start_col=2, end_col=17)

        _clear_range_force(ws, row_cursor, row_cursor, 2, 17, reset_style=False)

        # ✅ MERGE D:E for empty state message using helper
        _merge_description_cell(ws, row_cursor, 'Belum ada data revenue dan pajak', col_start=4, col_end=5)
        cell = ws.cell(row=row_cursor, column=4)
        cell.font = cell.font.copy(italic=True, color='808080')
        cell.alignment = cell.alignment.copy(horizontal='center')
        row_cursor += 1

    # ✅ Calculate last & total row - SIMPLE
    # CEGAH RUMUS TERBALIK SAAT DATA KOSONG
    visible_revenue_rows = max(1, len(revenues))
    total_revenue_row = REVENUE_START_ROW + visible_revenue_rows
    last_revenue_row = total_revenue_row - 1

    # ✅ MERGE B8:E8 UNTUK TEKS REVENUE (IDR)
    merge_range_revenue = f'B{total_revenue_row}:E{total_revenue_row}'
    try:
        ws.unmerge_cells(merge_range_revenue)
    except Exception:
        pass
    ws.merge_cells(merge_range_revenue)

    _safe_set_cell(ws, total_revenue_row, 2, 'REVENUE (IDR)')
    ws.cell(row=total_revenue_row, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=total_revenue_row, column=2).font = Font(bold=True)

    # ✅ Set total row with Excel formulas
    _set_formula_with_format(ws, total_revenue_row, COL_INVOICE_VALUE, f'=SUM(F{REVENUE_START_ROW}:F{last_revenue_row})')
    _set_formula_with_format(ws, total_revenue_row, COL_AMOUNT_RECEIVED, f'=SUM(L{REVENUE_START_ROW}:L{last_revenue_row})')
    _set_formula_with_format(ws, total_revenue_row, COL_PPN, f'=SUM(M{REVENUE_START_ROW}:M{last_revenue_row})')
    _set_formula_with_format(ws, total_revenue_row, COL_PPH_23, f'=SUM(N{REVENUE_START_ROW}:N{last_revenue_row})')
    _set_formula_with_format(ws, total_revenue_row, COL_TRANSFER_FEE, f'=SUM(O{REVENUE_START_ROW}:O{last_revenue_row})')

    # ✅ Style yellow for total row
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for col_idx in range(9, 15):
        ws.cell(row=total_revenue_row, column=col_idx).fill = yellow_fill

    # ✅ Hide unused rows between total and template end
    if total_revenue_row < REVENUE_TEMPLATE_END:
        _set_rows_hidden(ws, total_revenue_row + 1, REVENUE_TEMPLATE_END, True)

    # ✅ TABLE 2: PAJAK PENGELUARAN - ANCHOR TO STATC ROW 24
    # Calculate starting row for Table 2. If Table 1 is huge, we shift down.
    # Otherwise, we stay at TAX_TITLE_TEMPLATE_ROW (24).
    revenue_gap_row = total_revenue_row + 1
    tax_title_row = max(TAX_TITLE_TEMPLATE_ROW, total_revenue_row + 2)
    tax_header_top_row = tax_title_row + 1
    tax_header_row = tax_title_row + 2
    tax_start_row = tax_title_row + 3

    _copy_template_row(ws, REVENUE_SEPARATOR_TEMPLATE_ROW, revenue_gap_row, start_col=2, end_col=17, include_values=True)
    _clear_range(ws, revenue_gap_row, revenue_gap_row, 2, 17)
    _set_rows_hidden(ws, revenue_gap_row, revenue_gap_row, False)

    _copy_template_row(ws, TAX_TITLE_TEMPLATE_ROW, tax_title_row, start_col=2, end_col=17, include_values=True)
    _copy_template_row(ws, TAX_HEADER_TOP_TEMPLATE_ROW, tax_header_top_row, start_col=2, end_col=17, include_values=True)
    _copy_template_row(ws, TAX_HEADER_BOTTOM_TEMPLATE_ROW, tax_header_row, start_col=2, end_col=17, include_values=True)

    # ✅ FIX: Explicitly merge headers vertically (B, C, D:E)
    # This ensures headers like "Detail/Description" are correctly joined
    for col_range in [f'B{tax_header_top_row}:B{tax_header_row}',
                      f'C{tax_header_top_row}:C{tax_header_row}',
                      f'D{tax_header_top_row}:E{tax_header_row}']:
        try:
            ws.unmerge_cells(col_range)
        except Exception:
            pass
        ws.merge_cells(col_range)

        # Center align headers
        top_cell_ref = col_range.split(':')[0]
        ws[top_cell_ref].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    _safe_set_cell(ws, tax_title_row, 2, 'PAJAK PENGELUARAN')
    _set_rows_hidden(ws, tax_title_row, tax_header_row, False)

    # ✅ Clear data area - UNMERGE D:E FIRST (from template), then clear values
    # This is CRITICAL to remove existing merge from template before rendering
    # Always clear 10 rows max (even if no data, to remove template placeholders)
    for idx in range(10):
        row = tax_start_row + idx
        _copy_template_row(ws, TAX_DATA_TEMPLATE_ROW, row, start_col=2, end_col=17, include_values=False)
        # Unmerge D:E first (from template)
        try:
            ws.unmerge_cells(f'D{row}:E{row}')
        except Exception:
            pass
        # Clear all columns B-Q (2-17)
        for col in range(2, 18):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    visible_tax_rows = max(1, min(len(taxes), 10))

    # Render tax rows dynamically with merged D:E columns
    for idx, t in enumerate(taxes[:10]):
        row = tax_start_row + idx

        _set_date_with_format(ws, row, 2, t.get('date'))
        _safe_set_cell(ws, row, 3, idx + 1)

        # ✅ MERGE D:E for TAX description using helper
        description = t.get('description') or ''
        _merge_description_cell(ws, row, description, col_start=4, col_end=5)

        # Get tax values
        trans_val = _to_float(t.get('transaction_value'))
        ppn_val = _to_float(t.get('ppn'))
        pph_21_val = _to_float(t.get('pph_21'))
        pph_23_val = _to_float(t.get('pph_23'))
        pph_26_val = _to_float(t.get('pph_26'))

        _safe_set_number(ws, row, 6, trans_val)
        _safe_set_cell(ws, row, 7, t.get('currency') or 'IDR')
        _safe_set_number(ws, row, 8, _to_float(t.get('currency_exchange'), default=1) or 1)

        # ✅ Set DPP & Value ONLY for tax that has value (not all columns!)
        # Format all numeric columns with thousand separator
        if ppn_val > 0:
            _safe_set_number(ws, row, 9, trans_val)  # DPP PPN
            _safe_set_number(ws, row, 10, ppn_val)   # PPN Value

        if pph_21_val > 0:
            _safe_set_number(ws, row, 11, trans_val)  # DPP PPh 21
            _safe_set_number(ws, row, 12, pph_21_val) # PPh 21 Value

        if pph_23_val > 0:
            _safe_set_number(ws, row, 13, trans_val)  # DPP PPh 23
            _safe_set_number(ws, row, 14, pph_23_val) # PPh 23 Value

        if pph_26_val > 0:
            _safe_set_number(ws, row, 15, trans_val)  # DPP PPh 26
            _safe_set_number(ws, row, 16, pph_26_val) # PPh 26 Value

    # Apply merge cells if any (BEFORE hide)
    _apply_manual_tax_combine_groups(
        ws,
        taxes[:visible_tax_rows],
        tax_combine_groups,
        start_row=tax_start_row,
    )

    # Handle no tax data
    if not taxes:
        # ✅ FIX: Clone style FIRST, then write
        _copy_template_row(ws, TAX_DATA_TEMPLATE_ROW, tax_start_row, start_col=2, end_col=17, include_values=False)

        # ✅ MERGE D:E for empty state message using helper
        _merge_description_cell(ws, tax_start_row, 'Belum ada data pajak', col_start=4, col_end=5)
        cell = ws.cell(row=tax_start_row, column=4)
        cell.font = cell.font.copy(italic=True, color='808080')
        cell.alignment = cell.alignment.copy(horizontal='center')
        _safe_set_cell(ws, tax_start_row, 7, 'IDR')
        _safe_set_number(ws, tax_start_row, 8, 1)
        visible_tax_rows = 1

    # ✅ FIX #2: TOTAL ROW MUST BE ADJACENT TO DATA (not hardcoded row 37!)
    total_tax_row = tax_start_row + visible_tax_rows
    _copy_template_row(ws, TAX_TOTAL_TEMPLATE_ROW, total_tax_row, start_col=2, end_col=17, include_values=False)

    # ✅ FIX #3: Use EXCEL FORMULAS (not hardcoded values)
    last_tax_data_row = total_tax_row - 1

    # ✅ MERGE B:H for "PENERIMAAN NEGARA (TAX) (IDR)" text
    merge_range_tax = f'B{total_tax_row}:H{total_tax_row}'
    try:
        ws.unmerge_cells(merge_range_tax)
    except Exception:
        pass
    ws.merge_cells(merge_range_tax)
    _safe_set_cell(ws, total_tax_row, 2, 'PENERIMAAN NEGARA (TAX) (IDR)')
    ws.cell(row=total_tax_row, column=2).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=total_tax_row, column=2).font = Font(bold=True)

    _set_formula_with_format(ws, total_tax_row, 10, f'=SUM(J{tax_start_row}:J{last_tax_data_row})')  # PPN
    _set_formula_with_format(ws, total_tax_row, 12, f'=SUM(L{tax_start_row}:L{last_tax_data_row})')  # PPh 21
    _set_formula_with_format(ws, total_tax_row, 14, f'=SUM(N{tax_start_row}:N{last_tax_data_row})')  # PPh 23
    _set_formula_with_format(ws, total_tax_row, 16, f'=SUM(P{tax_start_row}:P{last_tax_data_row})')  # PPh 26
    ws.cell(row=total_tax_row, column=9).value = 'PPN'
    ws.cell(row=total_tax_row, column=11).value = 'PPh'
    ws.cell(row=total_tax_row, column=13).value = 'PPh'
    ws.cell(row=total_tax_row, column=15).value = 'PPh'
    ws.cell(row=total_tax_row, column=17).value = '-'

    # Bold total row
    for col in range(2, 18):
        ws.cell(row=total_tax_row, column=col).font = Font(bold=True)

    if total_tax_row < TAX_TEMPLATE_END:
        _set_rows_hidden(ws, total_tax_row + 1, TAX_TEMPLATE_END, True)

    # ✅ TABLE 3: PENGELUARAN - ANCHOR TO STATIC ROW 39
    expense_title_row = max(EXPENSE_TITLE_TEMPLATE_ROW, total_tax_row + 2)
    expense_header_row = expense_title_row + 1
    expense_start_row = expense_title_row + 2
    # Define gap row for Table 3
    expense_gap_row = expense_title_row - 1

    _copy_template_row(ws, EXPENSE_GAP_TEMPLATE_ROW, expense_gap_row, start_col=2, end_col=17, include_values=True)
    _clear_range(ws, expense_gap_row, expense_gap_row, 2, 17)
    _copy_template_row(ws, EXPENSE_TITLE_TEMPLATE_ROW, expense_title_row, start_col=2, end_col=17, include_values=True)
    _safe_set_cell(ws, expense_title_row, 2, 'PENGELUARAN')
    _copy_template_row(ws, EXPENSE_HEADER_TEMPLATE_ROW, expense_header_row, start_col=2, end_col=17, include_values=True)
    _copy_template_row(ws, EXPENSE_DATA_TEMPLATE_ROW, expense_start_row, start_col=2, end_col=17, include_values=False)
    _set_rows_hidden(ws, expense_gap_row, expense_start_row, False)

    # Fetch root categories dynamically from DB - ORDER BY sort_order (from Kategori Tabular)
    root_cats_all = Category.query.filter_by(parent_id=None).order_by(Category.sort_order).all()

    # ✅ Pre-calculate totals per root category to filter zero ones
    all_cats_sync = Category.query.all()
    category_by_id_map = {c.id: c for c in all_cats_sync}
    root_totals = {}
    for e in expenses:
        nominal = _expense_amount_for_display(e)
        r_name, _ = _root_category_info(e.get('category_id'), category_by_id_map)
        if r_name:
            root_totals[r_name] = root_totals.get(r_name, 0.0) + nominal

    # Filter root_cats to only those with data
    root_cats = [c for c in root_cats_all if root_totals.get(c.name, 0.0) > 0]

    # If NO categories have data, fallback to show at least one (or handle empty state)
    if not root_cats and root_cats_all:
        root_cats = [root_cats_all[0]]

    cat_columns = [c.name for c in root_cats]
    cat_names = cat_columns # for mapping

    _write_dynamic_category_headers(ws, root_cats, header_row=expense_header_row)

    # ✅ FIX: Override font for Expense header row (row 40) to Arial Narrow 12
    # Template uses Aptos Narrow 10, we want Arial Narrow 12
    # Only apply to columns B-H (non-category headers)
    for col in range(2, 9):  # Columns B to H only
        cell = ws.cell(row=expense_header_row, column=col)
        if not isinstance(cell, MergedCell):
            cell.font = Font(bold=True, size=12, name='Arial Narrow')
            # Default alignment: center
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # ✅ FIX: Set vertical text for Source (E=5), Currency (G=7), Currency Exchange (H=8)
    # These columns should have text rotated 90 degrees (vertical)
    for col_idx in (5, 7, 8):  # E, G, H
        cell = ws.cell(row=expense_header_row, column=col_idx)
        if not isinstance(cell, MergedCell):
            # ✅ FIX: wrap_text=True + text_rotation=90 supaya text vertical wrap ke bawah
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90, wrap_text=True)

    # FIX: Set row height for header row to accommodate vertical text and wrapped category headers
    ws.row_dimensions[expense_header_row].height = ROW_HEIGHT_HEADER_VERTICAL

    # Define last_category_col for use below
    last_category_col = 9 + len(cat_names) - 1

    # TABLE 3: PENGELUARAN & OPERATION COST
    logger.debug(f'Rendering Table 3 (Expenses) starting at row {expense_start_row}')
    total_row = _render_expense_section_from_data(ws, expenses, cat_names, category_by_id_map, year, start_row=expense_start_row)
    logger.debug(f'Table 3 rendered up to row {total_row}')

    # ✅ ALIGNED BRANDING & MERGED SIGNATURES
    logo_path = os.path.abspath(os.path.join(current_app.root_path, '..', 'frontend', 'assets', 'images', 'sheet1.png'))

    # Clear ALL images from template to start fresh
    ws._images.clear()

    # Top Logo - Presisi agar terlihat rata kanan di kolom Q
    # Kita ikat di 'P1' agar gambar membentang menutupi P dan berakhir di ujung Q
    _add_image_to_sheet(ws, logo_path, 'O1', width=308, height=78)

    # Signature at the bottom of Table 3
    sig_name_row = total_row + 4
    sig_title_row = total_row + 5

    ws.merge_cells(f'B{sig_name_row}:E{sig_name_row}')
    ws[f'B{sig_name_row}'] = 'Nama Lengkap'
    ws[f'B{sig_name_row}'].font = Font(bold=True, underline='single')
    ws[f'B{sig_name_row}'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'B{sig_title_row}:E{sig_title_row}')
    ws[f'B{sig_title_row}'] = 'Direktur'
    ws[f'B{sig_title_row}'].font = Font(bold=True)
    ws[f'B{sig_title_row}'].alignment = Alignment(horizontal='center')

    # ✅ FOOTER IMAGE REMOVED (As requested)

    # ✅ APPLY AUTOFIT TO SHEET 1
    # Use last_category_col to define the range of columns to autofit
    _apply_autofit_sheet1(ws, last_category_col)

    # DO NOT REMOVE TABLES - Keep template structure intact
    logger.debug('Keeping Excel tables for proper Alt+A behavior')

    final_main_sheet_name = f'Revenue-Cost_{year}'
    if ws.title != final_main_sheet_name:
        ws.title = final_main_sheet_name
        logger.debug(f'Renamed sheet to {final_main_sheet_name}')

    has_formatted_secondary = _ensure_formatted_secondary_sheets(wb, year, template_dir)
    if has_formatted_secondary:
        # ✅ CASE 1: Template found (Preferred) - Use sophisticated side-by-side logic
        _sync_formatted_secondary_sheets(
            wb,
            payload,
            year,
            final_main_sheet_name,
            expense_total_row=total_row,
            revenue_last_row=last_revenue_row,
        )
    else:
        # ✅ CASE 2: No template found - Fallback to basic monthly summary
        logger.warning('Excel template not found - falling back to basic summary sheets')
        _write_secondary_summary_sheets(
            wb,
            payload,
            year,
            final_main_sheet_name,
            expense_total_row=total_row,
            revenue_last_row=last_revenue_row,
        )

    try:
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception as e:
        logger.warning(f'Failed to set calculation mode: {e}')

    logger.info('Export completed, sending file')

    # CRITICAL: Error handling for Excel save and send
    try:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f'Revenue-Cost_{year}.xlsx'
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f'ERROR: Failed to save/export Excel: {e}')
        wb.close()
        raise
