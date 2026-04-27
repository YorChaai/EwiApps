from openpyxl import load_workbook
import shutil

excel_path = r'D:\2. Organize\1. Projects\MiniProjectKPI_EWI\backend\scratch\migrateexceltosql\database\20250427_EWI Financial-Repport_2024.xlsx'
backup_path = r'D:\2. Organize\1. Projects\MiniProjectKPI_EWI\backend\scratch\migrateexceltosql\database\20250427_EWI Financial-Repport_2024_backup.xlsx'

# Backup first
shutil.copy(excel_path, backup_path)

wb = load_workbook(excel_path)
ws = wb['Laba rugi -2024']

print('Formula Awal E9 (Pendapatan Langsung):', ws['E9'].value)
print('Formula Awal E10 (Pendapatan Lain-lain):', ws['E10'].value)

# Memperbaiki Formula ke Kolom F (INVOICE VALUE) dari sheet Revenue-Cost_2024
# Baris 7 sampai 18 adalah Pendapatan Langsung
ws['E9'] = "=SUM('Revenue-Cost_2024'!F7:F18)"

# Baris 19 adalah Pendapatan Lain-lain (Bunga Bank)
ws['E10'] = "=SUM('Revenue-Cost_2024'!F19:F21)"

wb.save(excel_path)
print('Berhasil diperbaiki ke Kolom F (Invoice Value)!')
