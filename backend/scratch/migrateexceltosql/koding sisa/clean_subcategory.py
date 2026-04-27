#!/usr/bin/env python3
import argparse
import openpyxl


MAPPING = {
    # transportasi
    "Accomadation & Transportation": "Transportation",
    "Air Tikect": "Transportation",
    "Airplane": "Transportation",
    "Land Transportation": "Transportation",
    "Train Ticket": "Transportation",
    "transportasi": "Transportation",
    "transportasi Darat": "Transportation",
    "Transportasi Tools": "Transportation",
    "Transportation": "Transportation",
    "PRABUMULIH Field-ALFA by SI": "Transportation",
    # akomodasi
    "Hotel": "Accommodation",
    "Hotel and Logistic": "Accommodation",
    "Hotel and Loundry": "Accommodation",
    "Hotel and Laundry": "Accommodation",
    # tunjangan
    "Allowance": "Allowance",
    "Aloowance": "Allowance",
    "Field Bonus": "Allowance",
    "FIELD BONUS": "Allowance",
    "Tunjangan lapangan untuk 5 days": "Allowance",
    "Perdiem Uang Makan": "Allowance",
    "Perdiem uang makan": "Allowance",
    # makan
    "Meal": "Meal",
    "Meal Allowance": "Meal",
    "Meal on field Site": "Meal",
    "Meals at wellsite": "Meal",
    "Meals on field Site": "Meal",
    "Entertaiment on field Site": "Meal",
    # logistik
    "Logictic": "Logistic",
    "Logistic": "Logistic",
    "Logistik": "Logistic",
    # pengiriman
    "Shipping": "Shipping",
    "JNE courier": "Shipping",
    "ATK dan dokumen sent": "Shipping",
    # laundry
    "Laundry": "Laundry",
    "Laundry (25 - 30 May 2024)": "Laundry",
    # operasi
    "Operation": "Operation",
    "OPERATION": "Operation",
    "Operation Need": "Operation",
    # perjalanan
    "Trip Alan to TGB-033 Pertamina Cirebon": "Trip",
    "Trip to KL - Benchmark DTR": "Trip",
    "trip to Meeting at Cirebon": "Trip",
    "Trip Zurailey to TGB-033 Pertamina Cirebon": "Trip",
    # pelatihan
    "Training": "Training",
    "Training Course": "Training",
    "Trip to Bandung - Upskilling event Elnusa": "Training",
    # peralatan tangan
    "Hand Tools": "Hand Tools",
    "buy spare part connector from Pei-Genesis": "Hand Tools",
    # layanan it
    "Google Domain and Email Services PT. Exspan Wireline Indonesia": "IT Services",
    # medis
    "Medical": "Medical",

    # gaji
    "gaji": "Gaji",
    "thr": "Allowance",
    "bonus": "Allowance",

    # sales
    "sales cost": "Sales",
    "sales fee": "Sales",

    # sewa alat
    "rental tool": "Rental Tool",

    # pembuatan alat
    "pembuatan alat": "Pembuatan Alat",

    # data processing
    "data processing": "Data Processing",
    "data proccesing": "Data Processing",

    # modal kerja
    "penambahan modal": "Modal Kerja",
    "modal biaya kerja": "Modal Kerja",

    # operasi proyek
    "wastafel": "Project Operation",
    "project lampu taman": "Project Operation",
    "moving slickline": "Project Operation",

    # team building 
    "team building": "Team Building",

    # pemeliharaan
    "repair esor": "Maintenance",

    # lisensi software
    "lisence sonoechometer": "Software License",
    "license sonoechometer": "Software License",

    # suku cadang
    "sparepart dari pei-genesis": "Sparepart",
    "downhole sampling tool": "Sparepart",
    "kekurangan (ppn ) ke pt garindo": "Sparepart",

    # sewa ruangan
    "sewa ruangan kantor": "Sewa Ruangan",
    "sewa virtual office": "Sewa Ruangan",
}

LOWER_MAPPING = {k.strip().lower(): v for k, v in MAPPING.items()}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Clean legacy subcategory labels inside Expense# blocks.")
    parser.add_argument("--input", required=True, help="Source xlsx path")
    parser.add_argument("--output", required=True, help="Output xlsx path")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Target sheet name. Default: active sheet",
    )
    return parser.parse_args()


def clean_subcategories(input_path: str, output_path: str, sheet_name: str | None = None) -> None:
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    expense_headers = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if isinstance(label, str) and label.strip().lower().startswith("expense#"):
            expense_headers.append(r)

    changes_made = 0
    unmapped = set()

    for i, header_row in enumerate(expense_headers):
        next_row = expense_headers[i + 1] if i + 1 < len(expense_headers) else ws.max_row + 1
        for r in range(header_row + 1, next_row):
            date_val = ws.cell(r, 2).value
            desc_cell = ws.cell(r, 4)
            desc_val = desc_cell.value
            amount_val = ws.cell(r, 6).value
            if not (desc_val and isinstance(desc_val, str) and desc_val.strip()):
                continue

            has_date = date_val is not None
            has_amount = False
            if amount_val is not None:
                try:
                    has_amount = float(amount_val) > 0
                except (TypeError, ValueError):
                    has_amount = False

            if has_date or has_amount:
                continue

            desc_text = desc_val.strip()
            mapped = LOWER_MAPPING.get(desc_text.lower())
            if mapped is None:
                unmapped.add(desc_text)
                continue
            if mapped != desc_val:
                desc_cell.value = mapped
                changes_made += 1

    wb.save(output_path)
    print(f"Sheet            : {ws.title}")
    print(f"Expense blocks   : {len(expense_headers)}")
    print(f"Rows normalized  : {changes_made}")
    if unmapped:
        print("Unmapped labels:")
        for item in sorted(unmapped):
            print(f"- {item}")


def main() -> None:
    args = parse_args()
    clean_subcategories(args.input, args.output, args.sheet)


if __name__ == "__main__":
    main()
