MIDDLE_PART = """    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            parsed = dt.datetime.strptime(text[:19], fmt)
            return parsed.isoformat(sep=" ", timespec="seconds")
        except ValueError:
            continue
    return None


def to_iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            parsed = dt.datetime.strptime(text[:19], fmt)
            return parsed.date().isoformat()
        except ValueError:
            continue
    return None


def to_num(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    text = text.replace(",", "")
    if text.startswith("="):
        expr = text[1:].strip()
        # dukung formula numerik sederhana dari template, contoh: "=7*100"
        if expr and re.fullmatch(r"[0-9\\.\\+\\-\\.*/\\(\\)\\s]+", expr):
            try:
                result = eval(expr, {"__builtins__": {}}, {})
                if isinstance(result, (int, float)):
                    return float(result)
            except Exception:
                return default
        return default
    try:
        return float(text)
    except ValueError:
        return default


_CELL_REF_RE = re.compile(r"\\$?([A-Z]{1,3})\\$?(\\d+)")


def _eval_formula_expr(
    expr: str,
    ws_formula,
    ws_values,
    _depth: int = 0,
) -> float | None:
    # evaluasi ekspresi aritmatika sederhana bergaya excel menjadi float.
    #
    # dukungan:
    # - +, -, *, /, (), %, desimal
    # - referensi cell seperti f8, $n$19 (di sheet yang sama)
    if _depth > 8:
        return None

    work = expr.strip()
    if not work:
        return None

    # konversi persentase, contoh: 11% -> (11/100)
    work = re.sub(
        r"(\\d+(?:\\.\\d+)?)\\s*%",
        lambda m: str(float(m.group(1)) / 100.0),
        work,
    )

    def _ref_repl(match: re.Match) -> str:
        col_letters = match.group(1)
        row_num = int(match.group(2))
        from openpyxl.utils.cell import column_index_from_string
        col_num = column_index_from_string(col_letters)
        val = to_num_cell(ws_formula, ws_values, row_num, col_num, 0.0, _depth=_depth + 1)
        return str(val)

    work = _CELL_REF_RE.sub(_ref_repl, work)
    work = work.replace("^", "**")

    # hanya aritmatika aman
    if not re.fullmatch(r"[0-9\\.\\+\\-\\*/\\(\\)\\s]+", work):
        return None
    try:
        result = eval(work, {"__builtins__": {}}, {})
    except Exception:
        return None
    if isinstance(result, (int, float)):
        return float(result)
    return None


def to_num_cell(
    ws_formula,
    ws_values,
    row_num: int,
    col_num: int,
    default: float = 0.0,
    _depth: int = 0,
) -> float:
    # baca nilai numerik dari cell worksheet, dengan fallback ke formula
    v_cached = ws_values.cell(row_num, col_num).value
    if isinstance(v_cached, (int, float)):
        return float(v_cached)

    v_formula = ws_formula.cell(row_num, col_num).value
    if isinstance(v_formula, (int, float)):
        return float(v_formula)

    text = to_text(v_formula)
    if not text:
        return default
    if text.startswith("="):
        parsed = _eval_formula_expr(text[1:], ws_formula, ws_values, _depth=_depth)
        return parsed if parsed is not None else default
    return to_num(text, default)


def to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_subcategory_alias_map() -> dict[str, str]:
    # mode cleaned: gunakan label subkategori standar secara langsung
    return {name.lower(): name for name in STANDARD_SUBCATEGORIES}


SUBCATEGORY_ALIAS_MAP = build_subcategory_alias_map()


def normalize_subcategory(value: str) -> str:
    text = to_text(value)
    if not text:
        return ""
    clean = text.lower()
    # Exact match from SUBCATEGORY_ALIAS_MAP
    exact = SUBCATEGORY_ALIAS_MAP.get(clean)
    if exact:
        return exact
    # Exact match from FULL_MAPPING
    mapped = FULL_MAPPING.get(clean)
    if mapped:
        return mapped
    # Partial/prefix match from FULL_MAPPING
    for keyword, mapped_val in FULL_MAPPING.items():
        if clean.startswith(keyword) or keyword in clean:
            return mapped_val
    return text


def ensure_schema(conn: sqlite3.Connection, with_excel_metadata: bool) -> None:
    conn.executescript(
        \"\"\"
        PRAGMA foreign_keys = ON;

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER NOT NULL,
            username VARCHAR(80) NOT NULL,
            password_hash VARCHAR(256) NOT NULL,
            full_name VARCHAR(150) NOT NULL,
            role VARCHAR(20) NOT NULL,
            created_at DATETIME,
            PRIMARY KEY (id),
            UNIQUE (username)
        );

        CREATE TABLE IF NOT EXISTS revenues (
            id INTEGER NOT NULL,
            invoice_date DATE NOT NULL,
            description VARCHAR(300) NOT NULL,
            invoice_value FLOAT NOT NULL,
            currency VARCHAR(10),
            currency_exchange FLOAT,
            invoice_number VARCHAR(50),
            client VARCHAR(150),
            receive_date DATE,
            amount_received FLOAT,
            ppn FLOAT,
            pph_23 FLOAT,
            transfer_fee FLOAT,
            remark TEXT,
            created_at DATETIME,
            PRIMARY KEY (id)
        );

        CREATE TABLE IF NOT EXISTS taxes (
            id INTEGER NOT NULL,
            date DATE NOT NULL,
            description VARCHAR(300) NOT NULL,
            transaction_value FLOAT NOT NULL,
            currency VARCHAR(10),
            currency_exchange FLOAT,
            ppn FLOAT,
            pph_21 FLOAT,
            pph_23 FLOAT,
            pph_26 FLOAT,
            created_at DATETIME,
            PRIMARY KEY (id)
        );

        CREATE TABLE IF NOT EXISTS dividends (
            id INTEGER NOT NULL,
            date DATE NOT NULL,
            name VARCHAR(150) NOT NULL,
            amount FLOAT NOT NULL,
            tax_percentage FLOAT,
            created_at DATETIME,
            PRIMARY KEY (id)
        );

        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER NOT NULL,
            name VARCHAR(100) NOT NULL,
            code VARCHAR(10) NOT NULL,
            parent_id INTEGER,
            status VARCHAR(20),
            created_by INTEGER,
            PRIMARY KEY (id),
            UNIQUE (code),
            FOREIGN KEY(parent_id) REFERENCES categories (id),
            FOREIGN KEY(created_by) REFERENCES users (id)
        );

        CREATE TABLE IF NOT EXISTS advances (
            id INTEGER NOT NULL,
            title VARCHAR(200) NOT NULL,
            description TEXT,
            user_id INTEGER NOT NULL,
            status VARCHAR(20),
            notes TEXT,
            created_at DATETIME,
            updated_at DATETIME,
            approved_at DATETIME,
            PRIMARY KEY (id),
            FOREIGN KEY(user_id) REFERENCES users (id)
        );

        CREATE TABLE IF NOT EXISTS advance_items (
            id INTEGER NOT NULL,
            advance_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            description VARCHAR(300) NOT NULL,
            estimated_amount FLOAT NOT NULL,
            evidence_path VARCHAR(500),
            evidence_filename VARCHAR(200),
            created_at DATETIME,
            PRIMARY KEY (id),
            FOREIGN KEY(advance_id) REFERENCES advances (id),
            FOREIGN KEY(category_id) REFERENCES categories (id)
        );

        CREATE TABLE IF NOT EXISTS settlements (
            id INTEGER NOT NULL,
            title VARCHAR(200) NOT NULL,
            description TEXT,
            user_id INTEGER NOT NULL,
            settlement_type VARCHAR(10),
            status VARCHAR(20),
            created_at DATETIME,
            updated_at DATETIME,
            completed_at DATETIME,
            advance_id INTEGER,
            PRIMARY KEY (id),
            FOREIGN KEY(user_id) REFERENCES users (id),
            FOREIGN KEY(advance_id) REFERENCES advances (id)
        );

        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER NOT NULL,
            settlement_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            description VARCHAR(300) NOT NULL,
            amount FLOAT NOT NULL,
            date DATE NOT NULL,
            source VARCHAR(50),
            currency VARCHAR(10),
            currency_exchange FLOAT,
            evidence_path VARCHAR(500),
            evidence_filename VARCHAR(200),
            status VARCHAR(20),
            notes TEXT,
            created_at DATETIME,
            PRIMARY KEY (id),
            FOREIGN KEY(settlement_id) REFERENCES settlements (id),
            FOREIGN KEY(category_id) REFERENCES categories (id)
        );

        CREATE TABLE IF NOT EXISTS report_entry_tags (
            id INTEGER PRIMARY KEY,
            table_name TEXT NOT NULL,
            row_id INTEGER NOT NULL,
            report_year INTEGER NOT NULL,
            source_excel TEXT,
            imported_at DATETIME,
            UNIQUE(table_name, row_id, report_year)
        );
        \"\"\"
    )
    if with_excel_metadata:
        conn.executescript(
            \"\"\"
            CREATE TABLE IF NOT EXISTS excel_sheets (
                id INTEGER PRIMARY KEY,
                sheet_name TEXT NOT NULL,
                max_row INTEGER NOT NULL,
                max_col INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS excel_merged_ranges (
                id INTEGER PRIMARY KEY,
                sheet_id INTEGER NOT NULL,
                cell_range TEXT NOT NULL,
                FOREIGN KEY(sheet_id) REFERENCES excel_sheets(id)
            );

            CREATE TABLE IF NOT EXISTS excel_cells (
                id INTEGER PRIMARY KEY,
                sheet_id INTEGER NOT NULL,
                row_num INTEGER NOT NULL,
                col_num INTEGER NOT NULL,
                coordinate TEXT NOT NULL,
                raw_value TEXT,
                cached_value TEXT,
                value_type TEXT,
                is_formula INTEGER NOT NULL DEFAULT 0,
                formula TEXT,
                number_format TEXT,
                target_table TEXT,
                target_pk INTEGER,
                target_column TEXT,
                FOREIGN KEY(sheet_id) REFERENCES excel_sheets(id)
            );
            \"\"\"
        )
    conn.commit()


def tag_report_row(
    conn: sqlite3.Connection,
    table_name: str,
    row_id: int,
    report_year: int,
    source_excel: str,
) -> None:
    conn.execute(
        \"\"\"
        INSERT OR IGNORE INTO report_entry_tags
            (table_name, row_id, report_year, source_excel, imported_at)
        VALUES (?, ?, ?, ?, ?)
        \"\"\",
        (
            table_name,
            row_id,
            int(report_year),
            source_excel,
            dt.datetime.now().isoformat(sep=" ", timespec="seconds"),
        ),
    )


def ensure_reference_data(conn: sqlite3.Connection, with_excel_metadata: bool) -> dict[str, int]:
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    if with_excel_metadata:
        conn.execute("DELETE FROM excel_cells")
        conn.execute("DELETE FROM excel_merged_ranges")
        conn.execute("DELETE FROM excel_sheets")
    else:
        conn.execute("DROP TABLE IF EXISTS excel_cells")
        conn.execute("DROP TABLE IF EXISTS excel_merged_ranges")
        conn.execute("DROP TABLE IF EXISTS excel_sheets")

    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if user_count == 0:
        users = [
            (1, "manager1", "manager12345", "Manager", "manager"),
            (2, "staff1", "staff12345", "Staff 1", "staff"),
            (3, "staff2", "staff67890", "Staff 2", "staff"),
            (4, "mitra1", "mitra12345", "Mitra Eksternal 1", "mitra_eks"),
        ]
        for user_id, username, plain_password, full_name, role in users:
            conn.execute(
                \"\"\"
                INSERT INTO users (id, username, password_hash, full_name, role, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                \"\"\",
                (
                    user_id,
                    username,
                    generate_password_hash(plain_password),
                    full_name,
                    role,
                    now,
                ),
            )

    existing_categories = conn.execute(
        "SELECT id, name, code FROM categories ORDER BY id"
    ).fetchall()
    if existing_categories:
        category_ids_by_code = {code: cat_id for cat_id, _, code in existing_categories}
        category_ids_by_name = {name.lower(): cat_id for cat_id, name, _ in existing_categories}
        return category_ids_by_code, category_ids_by_name

    initial_data = {
        "Biaya Operasi": [
            "Transportation", "Accommodation", "Allowance", "Meal",
            "Shipping", "Laundry", "Operation", "Trip", "Training",
            "Gaji", "Sales", "Project Operation", "Team Building", "Maintenance",
        ],
        "Biaya Research (R&D)": [
            "Pembuatan Alat",
        ],
        "Biaya Sewa Peralatan": [
            "Rental Tool",
        ],
        "Biaya Interpretasi Log Data": [
            "Data Processing", "Software License",
        ],
        "Administrasi": [
            "IT Services",
            "Biaya Bank",
        ],
        "Pembelian Barang": [
            "Logistic", "Hand Tools", "Sparepart",
        ],
        "Sewa Kantor": [
            "Sewa Ruangan",
        ],
        "Kesehatan": [
            "Medical",
        ],
        "Bisnis Dev": [
            "Modal Kerja",
        ],
    }

    import string
    letters = string.ascii_uppercase
    
    categories = []
    cat_id = 1
    parent_idx = 0
    
    for parent_name, subs in initial_data.items():
        if parent_idx < len(letters):
            code = letters[parent_idx]
        else:
            code = f"Z{parent_idx - len(letters) + 1}"
            
        parent_id = cat_id
        categories.append((parent_id, parent_name, code, None))
        cat_id += 1
        
        for i, sub_name in enumerate(subs, 1):
            sub_code = f"{code}{i}"
            categories.append((cat_id, sub_name, sub_code, parent_id))
            cat_id += 1
            
        parent_idx += 1
    conn.executemany(
        \"\"\"
        INSERT INTO categories (id, name, code, parent_id, status, created_by)
        VALUES (?, ?, ?, ?, 'approved', 1)
        \"\"\",
        categories,
    )
    conn.commit()

    category_ids_by_code = {code: cat_id for cat_id, _, code, _ in categories}
    category_ids_by_name = {name.lower(): cat_id for cat_id, name, _, _ in categories}
    return category_ids_by_code, category_ids_by_name


def purge_year_data(conn: sqlite3.Connection, year: int) -> None:
    # pembersihan yang disarankan: berdasarkan tag dataset (report_year), independen dari tahun transaksi
    try:
        tagged = conn.execute(
            \"\"\"
            SELECT table_name, row_id
            FROM report_entry_tags
            WHERE report_year = ?
            ORDER BY id
            \"\"\",
            (int(year),),
        ).fetchall()
    except sqlite3.Error:
        tagged = []

    if tagged:
        ids_by_table: dict[str, set[int]] = {}
        for table_name, row_id in tagged:
            ids_by_table.setdefault(table_name, set()).add(int(row_id))

        # hapus tabel child lebih dulu untuk memenuhi konstrain fk
        delete_order = [
            "expenses",
            "advance_items",
            "revenues",
            "taxes",
            "dividends",
            "settlements",
            "advances",
            "users",
            "categories",
        ]
        ordered_tables = [t for t in delete_order if t in ids_by_table] + [
            t for t in ids_by_table.keys() if t not in delete_order
        ]

        for table_name in ordered_tables:
            ids = ids_by_table.get(table_name, set())
            if not ids:
                continue
            placeholders = ",".join("?" for _ in ids)
            conn.execute(
                f"DELETE FROM {table_name} WHERE id IN ({placeholders})",
                tuple(ids),
            )

        conn.execute("DELETE FROM report_entry_tags WHERE report_year = ?", (int(year),))
        conn.commit()
        return

    # fallback legacy: berdasarkan tahun transaksi
    year_text = str(year)
    imported_settlement_ids = [
        row[0]
        for row in conn.execute(
            \"\"\"
            SELECT DISTINCT e.settlement_id
            FROM expenses e
            JOIN settlements s ON s.id = e.settlement_id
            WHERE strftime('%Y', e.date) = ?
              AND s.description LIKE 'Imported from Sheet1 row %'
            \"\"\",
            (year_text,),
        ).fetchall()
    ]

    conn.execute("DELETE FROM revenues WHERE strftime('%Y', invoice_date) = ?", (year_text,))
    conn.execute("DELETE FROM taxes WHERE strftime('%Y', date) = ?", (year_text,))
    conn.execute("DELETE FROM dividends WHERE strftime('%Y', date) = ?", (year_text,))
    conn.execute("DELETE FROM expenses WHERE strftime('%Y', date) = ?", (year_text,))

    for settlement_id in imported_settlement_ids:
        still_used = conn.execute(
            "SELECT 1 FROM expenses WHERE settlement_id = ? LIMIT 1",
            (settlement_id,),
        ).fetchone()
        if not still_used:
            conn.execute("DELETE FROM settlements WHERE id = ?", (settlement_id,))

    conn.commit()


def store_excel_structure(conn: sqlite3.Connection, ws_formula, ws_values, with_excel_metadata: bool) -> int | None:
    if not with_excel_metadata:
        return None

    cur = conn.execute(
        "INSERT INTO excel_sheets (sheet_name, max_row, max_col) VALUES (?, ?, ?)",
        (ws_formula.title, ws_formula.max_row, ws_formula.max_column),
    )
    sheet_id = int(cur.lastrowid)

    conn.executemany(
        "INSERT INTO excel_merged_ranges (sheet_id, cell_range) VALUES (?, ?)",
        [(sheet_id, str(rng)) for rng in ws_formula.merged_cells.ranges],
    )

    cell_rows = []
    for r in range(1, ws_formula.max_row + 1):
        for c in range(1, ws_formula.max_column + 1):
            cell_f = ws_formula.cell(r, c)
            cell_v = ws_values.cell(r, c)
            raw_value = cell_f.value
            cached_value = cell_v.value
            is_formula = int(isinstance(raw_value, str) and raw_value.startswith("="))
            formula = raw_value if is_formula else None
            cell_rows.append(
                (
                    sheet_id,
                    r,
                    c,
                    cell_f.coordinate,
                    None if raw_value is None else str(raw_value),
                    None if cached_value is None else str(cached_value),
                    type(raw_value).__name__ if raw_value is not None else "NoneType",
                    is_formula,
                    formula,
                    cell_f.number_format,
                    None,
                    None,
                    None,
                )
            )

    conn.executemany(
        \"\"\"
        INSERT INTO excel_cells (
            sheet_id, row_num, col_num, coordinate, raw_value, cached_value,
            value_type, is_formula, formula, number_format,
            target_table, target_pk, target_column
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        \"\"\",
        cell_rows,
    )
    conn.commit()
    return sheet_id


def extract_year_from_header(ws_formula) -> int:
    header_text = to_text(ws_formula.cell(2, 4).value)
    m = re.search(r"(\\d{4})", header_text)
    if m:
        return int(m.group(1))
    return dt.datetime.now().year


def tag_cell_mapping(
    conn: sqlite3.Connection,
    sheet_id: int | None,
    row_num: int,
    col_num: int,
    target_table: str,
    target_pk: int,
    target_column: str,
) -> None:
    if sheet_id is None:
        return
    conn.execute(
        \"\"\"
        UPDATE excel_cells
        SET target_table = ?, target_pk = ?, target_column = ?
        WHERE sheet_id = ? AND row_num = ? AND col_num = ?
        \"\"\",
        (target_table, target_pk, target_column, sheet_id, row_num, col_num),
    )


def import_revenues(
    conn: sqlite3.Connection,
    sheet_id: int | None,
    ws_formula,
    ws_values,
    report_year: int,
    source_excel: str,
) -> int:
    count = 0
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    for r in range(REVENUE_ROW_START, REVENUE_ROW_END + 1):
        invoice_date = to_iso_date(ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value)
        description = to_text(ws_formula.cell(r, 4).value)
        invoice_value = to_num_cell(ws_formula, ws_values, r, 6)
        if not invoice_date or not description:
            continue
        currency = to_text(ws_formula.cell(r, 7).value) or "IDR"
        exchange = to_num_cell(ws_formula, ws_values, r, 8, 1.0) or 1.0
        invoice_number = to_text(ws_formula.cell(r, 9).value) or None
        client = to_text(ws_formula.cell(r, 10).value) or None
        receive_date = to_iso_date(ws_values.cell(r, 11).value or ws_formula.cell(r, 11).value)
        amount_received = to_num_cell(ws_formula, ws_values, r, 12)
        ppn = to_num_cell(ws_formula, ws_values, r, 13)
        pph_23 = to_num_cell(ws_formula, ws_values, r, 14)
        transfer_fee = to_num_cell(ws_formula, ws_values, r, 15)
        remark = to_text(ws_formula.cell(r, 16).value) or None

        cur = conn.execute(
            \"\"\"
            INSERT INTO revenues (
                invoice_date, description, invoice_value, currency, currency_exchange,
                invoice_number, client, receive_date, amount_received,
                ppn, pph_23, transfer_fee, remark, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            \"\"\",
            (
                invoice_date,
                description,
                invoice_value,
                currency,
                exchange,
                invoice_number,
                client,
                receive_date,
                amount_received,
                ppn,
                pph_23,
                transfer_fee,
                remark,
                now,
            ),
        )
        revenue_id = int(cur.lastrowid)
        tag_cell_mapping(conn, sheet_id, r, 2, "revenues", revenue_id, "invoice_date")
        tag_cell_mapping(conn, sheet_id, r, 4, "revenues", revenue_id, "description")
        tag_cell_mapping(conn, sheet_id, r, 6, "revenues", revenue_id, "invoice_value")
        tag_cell_mapping(conn, sheet_id, r, 13, "revenues", revenue_id, "ppn")
        tag_cell_mapping(conn, sheet_id, r, 14, "revenues", revenue_id, "pph_23")
        tag_cell_mapping(conn, sheet_id, r, 15, "revenues", revenue_id, "transfer_fee")
        tag_report_row(conn, "revenues", revenue_id, report_year, source_excel)
        count += 1

    conn.commit()
    return count


def import_taxes(
    conn: sqlite3.Connection,
    sheet_id: int | None,
    ws_formula,
    ws_values,
    report_year: int,
    source_excel: str,
) -> int:
    count = 0
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    for r in range(TAX_ROW_START, TAX_ROW_END + 1):
        tax_date = to_iso_date(ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value)
        description = to_text(ws_formula.cell(r, 4).value)
        transaction_value = to_num_cell(ws_formula, ws_values, r, 6)
        if not tax_date or not description:
            continue
        currency = to_text(ws_formula.cell(r, 7).value) or "IDR"
        exchange = to_num_cell(ws_formula, ws_values, r, 8, 1.0) or 1.0
        ppn = to_num_cell(ws_formula, ws_values, r, 10)
        pph_21 = to_num_cell(ws_formula, ws_values, r, 12)
        pph_23 = to_num_cell(ws_formula, ws_values, r, 14)
        pph_26 = to_num_cell(ws_formula, ws_values, r, 16)

        cur = conn.execute(
            \"\"\"
            INSERT INTO taxes (
                date, description, transaction_value, currency, currency_exchange,
                ppn, pph_21, pph_23, pph_26, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            \"\"\",
            (
                tax_date,
                description,
                transaction_value,
                currency,
                exchange,
                ppn,
                pph_21,
                pph_23,
                pph_26,
                now,
            ),
        )
        tax_id = int(cur.lastrowid)
        tag_cell_mapping(conn, sheet_id, r, 2, "taxes", tax_id, "date")
        tag_cell_mapping(conn, sheet_id, r, 4, "taxes", tax_id, "description")
        tag_cell_mapping(conn, sheet_id, r, 6, "taxes", tax_id, "transaction_value")
        tag_cell_mapping(conn, sheet_id, r, 10, "taxes", tax_id, "ppn")
        tag_cell_mapping(conn, sheet_id, r, 12, "taxes", tax_id, "pph_21")
        tag_cell_mapping(conn, sheet_id, r, 14, "taxes", tax_id, "pph_23")
        tag_cell_mapping(conn, sheet_id, r, 16, "taxes", tax_id, "pph_26")
        tag_report_row(conn, "taxes", tax_id, report_year, source_excel)
        count += 1
    conn.commit()
    return count


def import_dividends(
    conn: sqlite3.Connection,
    wb_values,
    report_year: int,
    source_excel: str,
) -> int:
    count = 0
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    
    if "Business Summary" not in wb_values.sheetnames:
        return 0
        
    ws_bs = wb_values["Business Summary"]
    
    for r in range(19, 100):
        name = to_text(ws_bs.cell(r, 2).value)
        if not name:
            break
            
        amount = to_num(ws_bs.cell(r, 5).value, 0.0)
        if amount <= 0:
            continue
            
        tax_text = to_text(ws_bs.cell(r, 7).value)
        tax_pct = 10.0
        if tax_text:
            m = re.search(r'([\\d\\.]+)', tax_text)
            if m:
                tax_pct = float(m.group(1))
                
        div_date = f"{report_year}-12-31"
        
        cur = conn.execute(
            \"\"\"
            INSERT INTO dividends (date, name, amount, tax_percentage, created_at)
            VALUES (?, ?, ?, ?, ?)
            \"\"\",
            (div_date, name, amount, tax_pct, now),
        )
        div_id = int(cur.lastrowid)
        tag_report_row(conn, "dividends", div_id, report_year, source_excel)
        count += 1
        
    conn.commit()
    return count



def build_expense_blocks(ws_formula, ws_values) -> dict:
    # kembalikan {'standalone_rows': [...], 'batch_blocks': [(seq, header_row, start_row, end_row), ...]}
    headers: list[tuple[int, int]] = []
    for r in range(1, ws_formula.max_row + 1):
        label = ws_formula.cell(r, 2).value
        if isinstance(label, str) and label.strip().lower().startswith("expense#"):
            digits = "".join(ch for ch in label if ch.isdigit())
            seq = int(digits) if digits else len(headers) + 1
            headers.append((seq, r))

    batch_blocks: list[tuple[int, int, int, int]] = []
    for i, (seq, header_row) in enumerate(headers):
        next_row = headers[i + 1][1] if i + 1 < len(headers) else ws_formula.max_row + 1
        batch_blocks.append((seq, header_row, header_row + 1, next_row - 1))

    # baris mandiri: antara akhir bagian tax dan header expense# pertama
    first_expense_row = headers[0][1] if headers else ws_formula.max_row + 1
    standalone_start = TAX_ROW_END + 1
    standalone_rows: list[int] = []
    for r in range(standalone_start, first_expense_row):
        date_val = ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value
        desc_val = to_text(ws_formula.cell(r, 4).value)
        amount_val = to_num_cell(ws_formula, ws_values, r, 6)
        if date_val and desc_val and amount_val > 0:
            standalone_rows.append(r)

    return {'standalone_rows': standalone_rows, 'batch_blocks': batch_blocks}


def detect_category_id(
    ws_formula,
    ws_values,
    row_num: int,
    category_ids_by_code: dict[str, int],
    category_ids_by_name: dict[str, int],
    description: str,
    subcategory: str = "",
) -> int:
"""

import os
with open("d:/2. Organize/1. Projects/MiniProjectKPI_EWI/backend/scripts/restore.py", "w", encoding="utf-8") as f:
    f.write('''import os

with open("d:/2. Organize/1. Projects/MiniProjectKPI_EWI/backend/scripts/excel_to_app_db.py", "r", encoding="utf-8") as f:
    lines = f.read().splitlines()

# bagian atas
top = "\\n".join(lines[:203]) + "\\n"

# bagian tengah di script adalah MIDDLE_PART

# bagian bawah dimulai di mana detect_category_id berlanjut
# saya akan mencari bagian bawah di file
bottom_idx = 0
for i, line in enumerate(lines):
    if "description_lower = description.lower()" in line:
        bottom_idx = i
        break

bottom = "\\n".join(lines[bottom_idx:]) + "\\n"

# tunggu, bagaimana jika `excel_to_app_db.py` tidak memiliki bagian bawah `detect_category_id` yang baru?
# dari langkah 488, Ze7G.py memiliki:
# 797: def detect_category_id(
# 798:     ws_formula,
# 799:     ws_values,
# 800:     row_num: int,

# jadi bottom_idx dari Ze7G akan menemukannya.
''')
    f.write("MIDDLE = ")
    f.write('"""')
    f.write(MIDDLE_PART)
    f.write('"""\n')
    f.write("""
try:
    with open("d:/2. Organize/1. Projects/MiniProjectKPI_EWI/backend/scripts/excel_to_app_db_restored.py", "w", encoding="utf-8") as out:
        out.write(top)
        out.write(MIDDLE)
        out.write(bottom)
    print("RESTORE SCRIPT GENERATED")
except Exception as e:
    print(e)
""")
