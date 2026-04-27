import sqlite3
import os
import sys

from collections import OrderedDict
import re

sys.path.append('d:/2. Organize/1. Projects/MiniProjectKPI_EWI/backend')

from openpyxl import load_workbook
from routes.reports.annual import (
    _get_expense_blocks, _extract_imported_sheet_row, _normalize_subtitle,
    _extract_expense_subcategory
)

template_path = 'd:/2. Organize/1. Projects/MiniProjectKPI_EWI/excel/Revenue-Cost_2024_cleaned_asli_cleaned.xlsx'
wb = load_workbook(template_path, keep_links=False)
ws = wb['Revenue-Cost_2024'] if 'Revenue-Cost_2024' in wb.sheetnames else wb.active

expense_blocks = _get_expense_blocks(ws)

print(f"Expense blocks found in template: {len(expense_blocks)}")
for _, header_row, start_row, end_row in expense_blocks[:2]:
    title = ws.cell(row=header_row, column=4).value
    print(f"  Header row {header_row}, title: {title}")

# ambil batch settlement user dari db
db_path = 'd:/2. Organize/1. Projects/MiniProjectKPI_EWI/backend/database.db'
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row

batch_settlements = conn.execute("""
    SELECT s.id, s.title, s.description
    FROM settlements s 
    WHERE s.settlement_type='batch'
""").fetchall()

print(f"\nBatch settlements in DB: {len(batch_settlements)}")

group_by_header_row = {}
groups_without_header = []

for s in batch_settlements:
    hr = _extract_imported_sheet_row(s['description'])
    if hr is None:
        groups_without_header.append(s)
    else:
        group_by_header_row[hr] = s

print(f"Groups mapped by header row id: {list(group_by_header_row.keys())[:10]}")

# perilaku modifikasi user
base_summary_end = 96
first_expense_header_row = expense_blocks[0][1] if expense_blocks else (base_summary_end + 1)
extra_rows_needed = 25 # contoh berdasarkan item summary

print(f"User script modifying header_rows with extra_rows_needed: {extra_rows_needed}")
new_group_by_header_row = {}
for hr, group in group_by_header_row.items():
    if extra_rows_needed > 0 and hr >= first_expense_header_row:
        hr += extra_rows_needed
    new_group_by_header_row[hr] = group

print(f"New mapped header rows: {list(new_group_by_header_row.keys())[:10]}")

for _, header_row, start_row, end_row in expense_blocks[:2]:
    block = new_group_by_header_row.get(header_row)
    print(f"\nBlock at {header_row} -> Mapped: {block is not None}")

conn.close()
