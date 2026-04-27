"""
Test script to verify Excel export fix for missing expense items
Usage: python scripts/test_export_fix.py --year 2024

This script:
1. Counts expenses in the database
2. Counts expense rows in the exported Excel file
3. Compares to identify any discrepancies
"""
import os
import sys
import argparse
from openpyxl import load_workbook

# Add backend to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

def count_expenses_in_db(year):
    """Count expenses in database by settlement type"""
    from app import create_app
    from models import db, Expense, Settlement

    app = create_app()

    with app.app_context():
        # Count all expenses for the year
        expenses = Expense.query.join(Settlement).filter(
            Settlement.status.in_(('approved', 'completed')),
            db.extract('year', Expense.date) == year
        ).all()

        print(f"\n{'='*60}")
        print(f"DATABASE EXPENSES FOR {year}")
        print(f"{'='*60}")
        print(f"Total expenses in DB: {len(expenses)}")

        # Group by settlement type
        from collections import defaultdict
        groups = defaultdict(list)
        for e in expenses:
            is_batch = e.settlement.settlement_type == 'batch' if e.settlement else False
            key = 'Batch' if is_batch else 'Single'
            groups[key].append(e)

        print(f"\nBreakdown:")
        for key, items in groups.items():
            print(f"  {key}: {len(items)} expenses")

        # Group by settlement for batch expenses
        batch_groups = defaultdict(list)
        for e in expenses:
            if e.settlement and e.settlement.settlement_type == 'batch':
                batch_groups[e.settlement_id].append(e)

        if batch_groups:
            print(f"\nBatch expense groups:")
            for settlement_id, items in sorted(batch_groups.items()):
                print(f"  Settlement #{settlement_id}: {len(items)} expenses")

        return len(expenses), groups

    return 0, {}


def count_expenses_in_excel(excel_path, year):
    """Count expense rows in exported Excel file"""
    print(f"\n{'='*60}")
    print(f"EXCEL FILE ANALYSIS: {excel_path}")
    print(f"{'='*60}")

    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        return 0, {}

    wb = load_workbook(excel_path, data_only=True)

    # Find the Revenue-Cost sheet
    sheet_name = next((name for name in wb.sheetnames if f'Revenue-Cost_{year}' in name), None)
    if not sheet_name:
        print(f"❌ Sheet 'Revenue-Cost_{year}' not found")
        print(f"Available sheets: {wb.sheetnames}")
        return 0, {}

    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print(f"Max row: {ws.max_row}")

    # Find expense rows by scanning for date pattern in column B
    expense_rows = []
    current_batch = None
    batch_expenses = defaultdict(int)

    for row in range(100, ws.max_row + 1):  # Start from row 100 (after headers)
        # Skip hidden rows
        if ws.row_dimensions[row].hidden:
            continue

        col_b = ws.cell(row=row, column=2).value

        # Check for batch header
        col_b_str = str(col_b) if col_b else ''
        if col_b_str.lower().startswith('expense#'):
            current_batch = col_b_str
            continue

        # Check for date-like value in column B
        if col_b and str(col_b).strip():
            # Check if it's a date or looks like a date
            if isinstance(col_b, str) and ('-' in col_b or '/' in col_b):
                expense_rows.append(row)
                if current_batch:
                    batch_expenses[current_batch] += 1
            elif hasattr(col_b, 'year'):  # datetime object
                expense_rows.append(row)
                if current_batch:
                    batch_expenses[current_batch] += 1

    print(f"\nTotal expense rows found: {len(expense_rows)}")

    if batch_expenses:
        print(f"\nBatch breakdown:")
        for batch, count in sorted(batch_expenses.items()):
            print(f"  {batch}: {count} expenses")

    return len(expense_rows), batch_expenses


def verify_export(excel_path, year):
    """Main verification function"""
    print(f"\n{'#'*60}")
    print(f"# EXCEL EXPORT VERIFICATION FOR {year}")
    print(f"{'#'*60}")

    # Count in DB
    db_total, db_groups = count_expenses_in_db(year)

    # Count in Excel
    excel_total, excel_batches = count_expenses_in_excel(excel_path, year)

    # Compare
    print(f"\n{'='*60}")
    print(f"COMPARISON")
    print(f"{'='*60}")
    print(f"Database: {db_total} expenses")
    print(f"Excel:    {excel_total} expenses")

    if db_total == excel_total:
        print(f"\n✅ SUCCESS: All expenses exported correctly!")
        return True
    else:
        diff = db_total - excel_total
        print(f"\n❌ MISMATCH: {diff} expenses missing in Excel!")
        print(f"\nPossible causes:")
        print(f"  1. Row detection issue (_is_template_detail_data_row)")
        print(f"  2. Not enough template rows for batch expenses")
        print(f"  3. Row insertion failed during export")
        print(f"  4. Rows hidden incorrectly")
        return False


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Verify Excel export for expense items')
    parser.add_argument('--excel', type=str, help='Path to exported Excel file (optional)')
    parser.add_argument('--year', type=int, default=2024, help='Year to verify (default: 2024)')
    parser.add_argument('--db-only', action='store_true', help='Only check database, skip Excel')

    args = parser.parse_args()

    if args.db_only:
        count_expenses_in_db(args.year)
    else:
        if not args.excel:
            print("Error: --excel is required unless --db-only is specified")
            print("Usage: python test_export_fix.py --excel path/to/file.xlsx --year 2024")
            sys.exit(1)

        verify_export(args.excel, args.year)
