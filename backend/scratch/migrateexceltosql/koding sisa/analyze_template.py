"""
Analyze Excel template structure
Usage: python scripts/analyze_template.py
"""
import os
import sys
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

def analyze_template(template_path):
    print(f"\n{'#'*70}")
    print(f"# ANALYZING TEMPLATE: {template_path}")
    print(f"{'#'*70}\n")

    if not os.path.exists(template_path):
        print(f"❌ File not found: {template_path}")
        return

    wb = load_workbook(template_path, data_only=True)

    # Find Revenue-Cost sheet
    sheet_name = next((name for name in wb.sheetnames if 'Revenue-Cost' in name), None)
    if not sheet_name:
        print(f"❌ No Revenue-Cost sheet found")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print(f"Max row: {ws.max_row}")
    print(f"Max column: {ws.max_column}")

    # Find Expense headers
    print(f"\n{'='*70}")
    print(f"EXPENSE BLOCK STRUCTURE:")
    print(f"{'='*70}")

    expense_headers = []
    for row_num in range(1, min(ws.max_row + 1, 300)):
        col_b = ws.cell(row=row_num, column=2).value
        if col_b and str(col_b).strip().lower().startswith('expense#'):
            expense_headers.append((row_num, str(col_b).strip()))

    print(f"\nFound {len(expense_headers)} expense headers:")
    for row_num, label in expense_headers:
        print(f"  {label} at row {row_num}")

    # Analyze each expense block
    print(f"\n{'='*70}")
    print(f"EXPENSE BLOCK DETAILS:")
    print(f"{'='*70}")

    for idx, (header_row, label) in enumerate(expense_headers):
        # Determine block range
        if idx + 1 < len(expense_headers):
            next_header = expense_headers[idx + 1][0]
            end_row = next_header - 1
        else:
            end_row = min(header_row + 100, ws.max_row)

        start_row = header_row + 1

        # Count different row types
        detail_rows = []
        subcategory_rows = []
        empty_rows = []

        for r in range(start_row, end_row + 1):
            col_b = ws.cell(row=r, column=2).value
            col_c = ws.cell(row=r, column=3).value
            col_d = ws.cell(row=r, column=4).value

            # Check if date-like
            is_date = False
            if col_b:
                if hasattr(col_b, 'year'):
                    is_date = True
                elif isinstance(col_b, str) and ('-' in col_b or '/' in col_b):
                    is_date = True

            if is_date:
                detail_rows.append(r)
            elif col_d and str(col_d).strip():
                subcategory_rows.append(r)
            else:
                empty_rows.append(r)

        print(f"\n{label} (Row {header_row}):")
        print(f"  Block range: {start_row} - {end_row} ({end_row - start_row + 1} rows)")
        print(f"  Detail rows (with date): {len(detail_rows)}")
        print(f"  Subcategory headers: {len(subcategory_rows)}")
        print(f"  Empty rows: {len(empty_rows)}")

        if detail_rows:
            print(f"  Detail row numbers: {detail_rows[:10]}{'...' if len(detail_rows) > 10 else ''}")
        if subcategory_rows:
            print(f"  Subcategory rows: {subcategory_rows[:10]}{'...' if len(subcategory_rows) > 10 else ''}")

    # Check specific rows around Expense#1
    print(f"\n{'='*70}")
    print(f"SAMPLE ROW DATA (around Expense#1):")
    print(f"{'='*70}")

    if expense_headers:
        first_header = expense_headers[0][0]
        print(f"\nRows {first_header} to {first_header + 30}:")
        print(f"{'Row':<5} | {'Col B (Date/Label)':<35} | {'Col C (Seq)':<12} | {'Col D (Desc)':<40}")
        print("-" * 95)

        for r in range(first_header, min(first_header + 31, ws.max_row + 1)):
            col_b = ws.cell(row=r, column=2).value
            col_c = ws.cell(row=r, column=3).value
            col_d = ws.cell(row=r, column=4).value

            col_b_str = str(col_b)[:33] if col_b else ''
            col_c_str = str(col_c)[:10] if col_c else ''
            col_d_str = str(col_d)[:38] if col_d else ''

            hidden = 'HIDDEN' if ws.row_dimensions[r].hidden else ''

            print(f"{r:<5} | {col_b_str:<35} | {col_c_str:<12} | {col_d_str:<40} {hidden}")

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Analyze Excel template structure')
    parser.add_argument('--template', type=str,
                       default=r'..\excel\Revenue-Cost_2024_cleaned_asli_cleaned.xlsx',
                       help='Path to template Excel file')
    args = parser.parse_args()

    template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), args.template))
    analyze_template(template_path)
