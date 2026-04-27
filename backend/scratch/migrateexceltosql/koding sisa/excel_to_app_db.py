#!/usr/bin/env python3
# cd "D:\2. Organize\1. Projects\MiniProjectKPI_EWI\backend\scripts"
# ..\venv\Scripts\python.exe .\excel_to_app_db.py --excel "D:\2. Organize\1. Projects\MiniProjectKPI_EWI\excel\Revenue-Cost_2024_cleaned_asli_cleaned.xlsx" --year 2024 --output-db "..\database.db"

# script untuk mengkonversi template tahunan excel menjadi database sqlite untuk aplikasi
#
# fitur:
# - membaca semua cell worksheet (termasuk formula dan nilai cache)
# - menyimpan metadata struktur workbook (sheet, merged range, cell)
# - memetakan bagian excel ke tabel aplikasi: users, categories, settlements,
#   expenses, revenues, taxes

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sqlite3
from typing import Any

from werkzeug.security import generate_password_hash

from openpyxl import load_workbook


REVENUE_ROW_START = 8
REVENUE_ROW_END = 21
TAX_ROW_START = 27
TAX_ROW_END = 36

STANDARD_SUBCATEGORIES = [
    "Transportation",
    "Accommodation",
    "Allowance",
    "Meal",
    "Logistic",
    "Shipping",
    "Laundry",
    "Operation",
    "Trip",
    "Training",
    "Hand Tools",
    "IT Services",
    "Medical",
]

STANDARD_SUBCATEGORY_TO_PARENT_CODE = {
    "Transportation": "A",
    "Accommodation": "A",
    "Allowance": "A",
    "Meal": "A",
    "Logistic": "F",
    "Shipping": "A",
    "Laundry": "A",
    "Operation": "A",
    "Trip": "A",
    "Training": "A",
    "Hand Tools": "F",
    "Medical": "H",
}


FULL_MAPPING = {
    # Transportation
    "accomadation & transportation": "Transportation",
    "air tikect": "Transportation",
    "airplane": "Transportation",
    "land transportation": "Transportation",
    "train ticket": "Transportation",
    "transportasi": "Transportation",
    "transportasi darat": "Transportation",
    "transportasi tools": "Transportation",
    "transportation": "Transportation",
    "prabumulih field-alfa by si": "Transportation",

    # Accommodation
    "hotel": "Accommodation",
    "hotel and logistic": "Accommodation",
    "hotel and loundry": "Accommodation",

    # Allowance
    "allowance": "Allowance",
    "aloowance": "Allowance",
    "field bonus": "Allowance",
    "tunjangan lapangan untuk 5 days": "Allowance",
    "perdiem uang makan": "Allowance",

    # Meal
    "meal": "Meal",
    "meal allowance": "Meal",
    "meal on field site": "Meal",
    "meals at wellsite": "Meal",
    "meals on field site": "Meal",
    "entertaiment on field site": "Meal",

    # Logistic
    "logictic": "Logistic",
    "logistic": "Logistic",
    "logistik": "Logistic",

    # Shipping
    "shipping": "Shipping",
    "jne courier": "Shipping",
    "atk dan dokumen sent": "Shipping",

    # Laundry
    "laundry": "Laundry",
    "laundry (25 - 30 may 2024)": "Laundry",

    # Operation
    "operation": "Operation",
    "operation need": "Operation",

    # Trip
    "trip alan to tgb-033 pertamina cirebon": "Trip",
    "trip to kl - benchmark dtr": "Trip",
    "trip to meeting at cirebon": "Trip",
    "trip zurailey to tgb-033 pertamina cirebon": "Trip",

    # Training
    "training": "Training",
    "training course": "Training",
    "trip to bandung - upskilling event elnusa": "Training",

    # Hand Tools
    "hand tools": "Hand Tools",
    "buy spare part connector from pei-genesis": "Hand Tools",

    # IT Services
    "google domain and email services pt. exspan wireline indonesia": "IT Services",

    # Medical
    "medical": "Medical"
}

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert annual Excel template to app DB format")
    parser.add_argument("--excel", default=r"..\excel\Revenue-Cost_2024_cleaned_asli_cleaned.xlsx", help="Path to source Excel file")
    parser.add_argument(
        "--year",
        type=int,
        default=None,
        help="Override report year manually if Excel header/template year is unreliable",
    )
    parser.add_argument(
        "--only-dividends",
        action="store_true",
        help="Import only dividend rows for the selected year without replacing other report data",
    )
    parser.add_argument(
        "--output-db",
        default="database_new.db",
        help="Path to output sqlite DB (if omitted, auto-generate timestamped DB in --output-dir)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory for auto-generated DB file (default: <project>/data)",
    )
    parser.add_argument(
        "--with-excel-metadata",
        action="store_true",
        help="Store extra tables excel_sheets/excel_merged_ranges/excel_cells",
    )
    return parser.parse_args()


def default_output_dir() -> str:
    script_dir = os.path.abspath(os.path.dirname(__file__))
    backend_dir = os.path.abspath(os.path.join(script_dir, ".."))
    project_dir = os.path.abspath(os.path.join(backend_dir, ".."))
    return os.path.join(project_dir, "data")


def build_output_db_path(
    excel_path: str,
    output_db: str | None,
    output_dir: str | None,
    year: int,
) -> str:
    if output_db:
        return os.path.abspath(output_db)

    out_dir = os.path.abspath(output_dir) if output_dir else default_output_dir()
    os.makedirs(out_dir, exist_ok=True)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"dataset_{year}_{timestamp}.db"
    return os.path.join(out_dir, filename)


def to_iso_datetime(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min).isoformat(sep=" ", timespec="seconds")
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            parsed = dt.datetime.strptime(text[:19], fmt)
            return parsed.isoformat(sep=" ", timespec="seconds")
        except ValueError:
            continue
    return None


def to_iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            parsed = dt.datetime.strptime(text[:19], fmt)
            return parsed.date().isoformat()
        except ValueError:
            continue
    return None


def to_num(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    text = text.replace(",", "")
    if text.startswith("="):
        expr = text[1:].strip()
        # dukung formula numerik sederhana dari template, contoh: "=7*100"
        if expr and re.fullmatch(r"[0-9\.\+\-\*/\(\)\s]+", expr):
            try:
                result = eval(expr, {"__builtins__": {}}, {})
                if isinstance(result, (int, float)):
                    return float(result)
            except Exception:
                return default
        return default
    try:
        return float(text)
    except ValueError:
        return default


_CELL_REF_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")


def _eval_formula_expr(
    expr: str,
    ws_formula,
    ws_values,
    _depth: int = 0,
) -> float | None:
    # evaluasi ekspresi aritmatika sederhana bergaya excel menjadi float.
    #
    # dukungan:
    # - +, -, *, /, (), %, desimal
    # - referensi cell seperti f8, $n$19 (di sheet yang sama)
    if _depth > 8:
        return None

    work = expr.strip()
    if not work:
        return None

    # konversi persentase, contoh: 11% -> (11/100)
    work = re.sub(
        r"(\d+(?:\.\d+)?)\s*%",
        lambda m: str(float(m.group(1)) / 100.0),
        work,
    )

    def _ref_repl(match: re.Match) -> str:
        col_letters = match.group(1)
        row_num = int(match.group(2))
        from openpyxl.utils.cell import column_index_from_string
        col_num = column_index_from_string(col_letters)
        val = to_num_cell(ws_formula, ws_values, row_num, col_num, 0.0, _depth=_depth + 1)
        return str(val)

    work = _CELL_REF_RE.sub(_ref_repl, work)
    work = work.replace("^", "**")

    # hanya aritmatika aman
    if not re.fullmatch(r"[0-9\.\+\-\*/\(\)\s]+", work):
        return None
    try:
        result = eval(work, {"__builtins__": {}}, {})
    except Exception:
        return None
    if isinstance(result, (int, float)):
        return float(result)
    return None


def to_num_cell(
    ws_formula,
    ws_values,
    row_num: int,
    col_num: int,
    default: float = 0.0,
    _depth: int = 0,
) -> float:
    # baca nilai numerik dari cell worksheet, dengan fallback ke formula
    v_cached = ws_values.cell(row_num, col_num).value
    if isinstance(v_cached, (int, float)):
        return float(v_cached)

    v_formula = ws_formula.cell(row_num, col_num).value
    if isinstance(v_formula, (int, float)):
        return float(v_formula)

    text = to_text(v_formula)
    if not text:
        return default
    if text.startswith("="):
        parsed = _eval_formula_expr(text[1:], ws_formula, ws_values, _depth=_depth)
        return parsed if parsed is not None else default
    return to_num(text, default)


def to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_subcategory_alias_map() -> dict[str, str]:
    # mode cleaned: gunakan label subkategori standar secara langsung
    return {name.lower(): name for name in STANDARD_SUBCATEGORIES}


SUBCATEGORY_ALIAS_MAP = build_subcategory_alias_map()


def normalize_subcategory(value: str) -> str:
    text = to_text(value)
    if not text:
        return ""
    clean = text.lower()
    # pencocokan presisi dari subcategory_alias_map
    exact = SUBCATEGORY_ALIAS_MAP.get(clean)
    if exact:
        return exact
    # pencocokan presisi dari full_mapping
    mapped = FULL_MAPPING.get(clean)
    if mapped:
        return mapped
    # pencocokan parsial/awalan dari full_mapping
    for keyword, mapped_val in FULL_MAPPING.items():
        if clean.startswith(keyword) or keyword in clean:
            return mapped_val
    return text


def ensure_schema(conn: sqlite3.Connection, with_excel_metadata: bool) -> None:
    conn.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER NOT NULL,
            username VARCHAR(80) NOT NULL,
            password_hash VARCHAR(256) NOT NULL,
            full_name VARCHAR(150) NOT NULL,
            role VARCHAR(20) NOT NULL,
            created_at DATETIME,
            PRIMARY KEY (id),
            UNIQUE (username)
        );

        CREATE TABLE IF NOT EXISTS revenues (
            id INTEGER NOT NULL,
            invoice_date DATE NOT NULL,
            description VARCHAR(300) NOT NULL,
            invoice_value FLOAT NOT NULL,
            currency VARCHAR(10),
            currency_exchange FLOAT,
            invoice_number VARCHAR(50),
            client VARCHAR(150),
            receive_date DATE,
            amount_received FLOAT,
            ppn FLOAT,
            pph_23 FLOAT,
            transfer_fee FLOAT,
            remark TEXT,
            created_at DATETIME,
            PRIMARY KEY (id)
        );

        CREATE TABLE IF NOT EXISTS taxes (
            id INTEGER NOT NULL,
            date DATE NOT NULL,
            description VARCHAR(300) NOT NULL,
            transaction_value FLOAT NOT NULL,
            currency VARCHAR(10),
            currency_exchange FLOAT,
            ppn FLOAT,
            pph_21 FLOAT,
            pph_23 FLOAT,
            pph_26 FLOAT,
            created_at DATETIME,
            PRIMARY KEY (id)
        );

        CREATE TABLE IF NOT EXISTS dividends (
            id INTEGER NOT NULL,
            date DATE NOT NULL,
            name VARCHAR(150) NOT NULL,
            amount FLOAT NOT NULL,
            recipient_count INTEGER NOT NULL DEFAULT 1,
            tax_percentage FLOAT,
            created_at DATETIME,
            PRIMARY KEY (id)
        );

        CREATE TABLE IF NOT EXISTS dividend_settings (
            id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            profit_retained FLOAT NOT NULL DEFAULT 0.0,
            created_at DATETIME,
            PRIMARY KEY (id),
            UNIQUE (year)
        );

        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER NOT NULL,
            name VARCHAR(100) NOT NULL,
            code VARCHAR(10) NOT NULL,
            parent_id INTEGER,
            status VARCHAR(20),
            created_by INTEGER,
            PRIMARY KEY (id),
            UNIQUE (code),
            FOREIGN KEY(parent_id) REFERENCES categories (id),
            FOREIGN KEY(created_by) REFERENCES users (id)
        );

        CREATE TABLE IF NOT EXISTS advances (
            id INTEGER NOT NULL,
            title VARCHAR(200) NOT NULL,
            description TEXT,
            user_id INTEGER NOT NULL,
            status VARCHAR(20),
            notes TEXT,
            created_at DATETIME,
            updated_at DATETIME,
            approved_at DATETIME,
            PRIMARY KEY (id),
            FOREIGN KEY(user_id) REFERENCES users (id)
        );

        CREATE TABLE IF NOT EXISTS advance_items (
            id INTEGER NOT NULL,
            advance_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            description VARCHAR(300) NOT NULL,
            estimated_amount FLOAT NOT NULL,
            evidence_path VARCHAR(500),
            evidence_filename VARCHAR(200),
            created_at DATETIME,
            PRIMARY KEY (id),
            FOREIGN KEY(advance_id) REFERENCES advances (id),
            FOREIGN KEY(category_id) REFERENCES categories (id)
        );

        CREATE TABLE IF NOT EXISTS settlements (
            id INTEGER NOT NULL,
            title VARCHAR(200) NOT NULL,
            description TEXT,
            user_id INTEGER NOT NULL,
            settlement_type VARCHAR(10),
            status VARCHAR(20),
            created_at DATETIME,
            updated_at DATETIME,
            completed_at DATETIME,
            advance_id INTEGER,
            PRIMARY KEY (id),
            FOREIGN KEY(user_id) REFERENCES users (id),
            FOREIGN KEY(advance_id) REFERENCES advances (id)
        );

        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER NOT NULL,
            settlement_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            description VARCHAR(300) NOT NULL,
            amount FLOAT NOT NULL,
            date DATE NOT NULL,
            source VARCHAR(50),
            currency VARCHAR(10),
            currency_exchange FLOAT,
            evidence_path VARCHAR(500),
            evidence_filename VARCHAR(200),
            status VARCHAR(20),
            notes TEXT,
            created_at DATETIME,
            PRIMARY KEY (id),
            FOREIGN KEY(settlement_id) REFERENCES settlements (id),
            FOREIGN KEY(category_id) REFERENCES categories (id)
        );

        CREATE TABLE IF NOT EXISTS report_entry_tags (
            id INTEGER PRIMARY KEY,
            table_name TEXT NOT NULL,
            row_id INTEGER NOT NULL,
            report_year INTEGER NOT NULL,
            source_excel TEXT,
            imported_at DATETIME,
            UNIQUE(table_name, row_id, report_year)
        );
        """
    )
    if with_excel_metadata:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS excel_sheets (
                id INTEGER PRIMARY KEY,
                sheet_name TEXT NOT NULL,
                max_row INTEGER NOT NULL,
                max_col INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS excel_merged_ranges (
                id INTEGER PRIMARY KEY,
                sheet_id INTEGER NOT NULL,
                cell_range TEXT NOT NULL,
                FOREIGN KEY(sheet_id) REFERENCES excel_sheets(id)
            );

            CREATE TABLE IF NOT EXISTS excel_cells (
                id INTEGER PRIMARY KEY,
                sheet_id INTEGER NOT NULL,
                row_num INTEGER NOT NULL,
                col_num INTEGER NOT NULL,
                coordinate TEXT NOT NULL,
                raw_value TEXT,
                cached_value TEXT,
                value_type TEXT,
                is_formula INTEGER NOT NULL DEFAULT 0,
                formula TEXT,
                number_format TEXT,
                target_table TEXT,
                target_pk INTEGER,
                target_column TEXT,
                FOREIGN KEY(sheet_id) REFERENCES excel_sheets(id)
            );
            """
        )
    conn.commit()
    dividend_columns = {
        row[1]
        for row in conn.execute("PRAGMA table_info(dividends)").fetchall()
    }
    if "recipient_count" not in dividend_columns:
        conn.execute(
            "ALTER TABLE dividends ADD COLUMN recipient_count INTEGER NOT NULL DEFAULT 1"
        )
        conn.commit()


def tag_report_row(
    conn: sqlite3.Connection,
    table_name: str,
    row_id: int,
    report_year: int,
    source_excel: str,
) -> None:
    conn.execute(
        """
        INSERT OR IGNORE INTO report_entry_tags
            (table_name, row_id, report_year, source_excel, imported_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            table_name,
            row_id,
            int(report_year),
            source_excel,
            dt.datetime.now().isoformat(sep=" ", timespec="seconds"),
        ),
    )


def ensure_reference_data(conn: sqlite3.Connection, with_excel_metadata: bool) -> dict[str, int]:
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    if with_excel_metadata:
        conn.execute("DELETE FROM excel_cells")
        conn.execute("DELETE FROM excel_merged_ranges")
        conn.execute("DELETE FROM excel_sheets")
    else:
        conn.execute("DROP TABLE IF EXISTS excel_cells")
        conn.execute("DROP TABLE IF EXISTS excel_merged_ranges")
        conn.execute("DROP TABLE IF EXISTS excel_sheets")

    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if user_count == 0:
        users = [
            (1, "manager1", "manager12345", "Manager", "manager"),
            (2, "staff1", "staff12345", "Staff 1", "staff"),
            (3, "staff2", "staff67890", "Staff 2", "staff"),
            (4, "mitra1", "mitra12345", "Mitra Eksternal 1", "mitra_eks"),
        ]
        for user_id, username, plain_password, full_name, role in users:
            conn.execute(
                """
                INSERT INTO users (id, username, password_hash, full_name, role, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    username,
                    generate_password_hash(plain_password),
                    full_name,
                    role,
                    now,
                ),
            )

    existing_categories = conn.execute(
        "SELECT id, name, code FROM categories ORDER BY id"
    ).fetchall()
    if existing_categories:
        category_ids_by_code = {code: cat_id for cat_id, _, code in existing_categories}
        category_ids_by_name = {name.lower(): cat_id for cat_id, name, _ in existing_categories}
        return category_ids_by_code, category_ids_by_name

    initial_data = {
        "Biaya Operasi": [
            "Transportation", "Accommodation", "Allowance", "Meal",
            "Shipping", "Laundry", "Operation", "Trip", "Training",
            "Gaji", "Sales", "Project Operation", "Team Building", "Maintenance",
        ],
        "Biaya Research (R&D)": [
            "Pembuatan Alat",
        ],
        "Biaya Sewa Peralatan": [
            "Rental Tool",
        ],
        "Biaya Interpretasi Log Data": [
            "Data Processing", "Software License",
        ],
        "Administrasi": [
            "IT Services",
            "Biaya Bank",
        ],
        "Pembelian Barang": [
            "Logistic", "Hand Tools", "Sparepart",
        ],
        "Sewa Kantor": [
            "Sewa Ruangan",
        ],
        "Kesehatan": [
            "Medical",
        ],
        "Bisnis Dev": [
            "Modal Kerja",
        ],
    }

    import string
    letters = string.ascii_uppercase

    categories = []
    cat_id = 1
    parent_idx = 0

    for parent_name, subs in initial_data.items():
        if parent_idx < len(letters):
            code = letters[parent_idx]
        else:
            code = f"Z{parent_idx - len(letters) + 1}"

        parent_id = cat_id
        categories.append((parent_id, parent_name, code, None))
        cat_id += 1

        for i, sub_name in enumerate(subs, 1):
            sub_code = f"{code}{i}"
            categories.append((cat_id, sub_name, sub_code, parent_id))
            cat_id += 1

        parent_idx += 1
    conn.executemany(
        """
        INSERT INTO categories (id, name, code, parent_id, status, created_by)
        VALUES (?, ?, ?, ?, 'approved', 1)
        """,
        categories,
    )
    conn.commit()

    category_ids_by_code = {code: cat_id for cat_id, _, code, _ in categories}
    category_ids_by_name = {name.lower(): cat_id for cat_id, name, _, _ in categories}
    return category_ids_by_code, category_ids_by_name


def purge_year_data(
    conn: sqlite3.Connection,
    year: int,
    only_tables: set[str] | None = None,
) -> None:
    # pembersihan yang disarankan: berdasarkan tag dataset (report_year), independen dari tahun transaksi
    try:
        tagged = conn.execute(
            """
            SELECT table_name, row_id
            FROM report_entry_tags
            WHERE report_year = ?
            ORDER BY id
            """,
            (int(year),),
        ).fetchall()
    except sqlite3.Error:
        tagged = []

    if tagged:
        ids_by_table: dict[str, set[int]] = {}
        for table_name, row_id in tagged:
            if only_tables is not None and table_name not in only_tables:
                continue
            ids_by_table.setdefault(table_name, set()).add(int(row_id))

        if not ids_by_table:
            return

        # hapus tabel child lebih dulu untuk memenuhi konstrain fk
        delete_order = [
            "expenses",
            "advance_items",
            "revenues",
            "taxes",
            "dividends",
            "dividend_settings",
            "settlements",
            "advances",
            "users",
            "categories",
        ]
        ordered_tables = [t for t in delete_order if t in ids_by_table] + [
            t for t in ids_by_table.keys() if t not in delete_order
        ]

        for table_name in ordered_tables:
            ids = ids_by_table.get(table_name, set())
            if not ids:
                continue
            placeholders = ",".join("?" for _ in ids)
            conn.execute(
                f"DELETE FROM {table_name} WHERE id IN ({placeholders})",
                tuple(ids),
            )

        if only_tables is None:
            conn.execute("DELETE FROM report_entry_tags WHERE report_year = ?", (int(year),))
        else:
            placeholders = ",".join("?" for _ in only_tables)
            params = [int(year), *sorted(only_tables)]
            conn.execute(
                f"DELETE FROM report_entry_tags WHERE report_year = ? AND table_name IN ({placeholders})",
                params,
            )
        conn.commit()
        return

    # fallback legacy: berdasarkan tahun transaksi
    year_text = str(year)
    imported_settlement_ids = [
        row[0]
        for row in conn.execute(
            """
            SELECT DISTINCT e.settlement_id
            FROM expenses e
            JOIN settlements s ON s.id = e.settlement_id
            WHERE strftime('%Y', e.date) = ?
              AND s.description LIKE 'Imported from Sheet1 row %'
            """,
            (year_text,),
        ).fetchall()
    ]

    if only_tables is None or "revenues" in only_tables:
        conn.execute("DELETE FROM revenues WHERE strftime('%Y', invoice_date) = ?", (year_text,))
    if only_tables is None or "taxes" in only_tables:
        conn.execute("DELETE FROM taxes WHERE strftime('%Y', date) = ?", (year_text,))
    if only_tables is None or "dividends" in only_tables:
        conn.execute("DELETE FROM dividends WHERE strftime('%Y', date) = ?", (year_text,))
    if only_tables is None or "dividend_settings" in only_tables:
        conn.execute("DELETE FROM dividend_settings WHERE year = ?", (int(year),))
    if only_tables is None or "expenses" in only_tables:
        conn.execute("DELETE FROM expenses WHERE strftime('%Y', date) = ?", (year_text,))

    for settlement_id in imported_settlement_ids:
        still_used = conn.execute(
            "SELECT 1 FROM expenses WHERE settlement_id = ? LIMIT 1",
            (settlement_id,),
        ).fetchone()
        if not still_used:
            conn.execute("DELETE FROM settlements WHERE id = ?", (settlement_id,))

    conn.commit()


def store_excel_structure(conn: sqlite3.Connection, ws_formula, ws_values, with_excel_metadata: bool) -> int | None:
    if not with_excel_metadata:
        return None

    cur = conn.execute(
        "INSERT INTO excel_sheets (sheet_name, max_row, max_col) VALUES (?, ?, ?)",
        (ws_formula.title, ws_formula.max_row, ws_formula.max_column),
    )
    sheet_id = int(cur.lastrowid)

    conn.executemany(
        "INSERT INTO excel_merged_ranges (sheet_id, cell_range) VALUES (?, ?)",
        [(sheet_id, str(rng)) for rng in ws_formula.merged_cells.ranges],
    )

    cell_rows = []
    for r in range(1, ws_formula.max_row + 1):
        for c in range(1, ws_formula.max_column + 1):
            cell_f = ws_formula.cell(r, c)
            cell_v = ws_values.cell(r, c)
            raw_value = cell_f.value
            cached_value = cell_v.value
            is_formula = int(isinstance(raw_value, str) and raw_value.startswith("="))
            formula = raw_value if is_formula else None
            cell_rows.append(
                (
                    sheet_id,
                    r,
                    c,
                    cell_f.coordinate,
                    None if raw_value is None else str(raw_value),
                    None if cached_value is None else str(cached_value),
                    type(raw_value).__name__ if raw_value is not None else "NoneType",
                    is_formula,
                    formula,
                    cell_f.number_format,
                    None,
                    None,
                    None,
                )
            )

    conn.executemany(
        """
        INSERT INTO excel_cells (
            sheet_id, row_num, col_num, coordinate, raw_value, cached_value,
            value_type, is_formula, formula, number_format,
            target_table, target_pk, target_column
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        cell_rows,
    )
    conn.commit()
    return sheet_id


def extract_year_from_header(ws_formula, excel_path: str | None = None, workbook=None) -> int:
    candidates: list[str] = []

    if excel_path:
        candidates.append(os.path.basename(excel_path))

    if workbook is not None:
        candidates.extend(str(name) for name in workbook.sheetnames)

    candidates.append(to_text(ws_formula.cell(2, 4).value))

    for text in candidates:
        m = re.search(r"(20\d{2})", text or "")
        if m:
            return int(m.group(1))

    return dt.datetime.now().year


def tag_cell_mapping(
    conn: sqlite3.Connection,
    sheet_id: int | None,
    row_num: int,
    col_num: int,
    target_table: str,
    target_pk: int,
    target_column: str,
) -> None:
    if sheet_id is None:
        return
    conn.execute(
        """
        UPDATE excel_cells
        SET target_table = ?, target_pk = ?, target_column = ?
        WHERE sheet_id = ? AND row_num = ? AND col_num = ?
        """,
        (target_table, target_pk, target_column, sheet_id, row_num, col_num),
    )


def import_revenues(
    conn: sqlite3.Connection,
    sheet_id: int | None,
    ws_formula,
    ws_values,
    report_year: int,
    source_excel: str,
) -> int:
    count = 0
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    for r in range(REVENUE_ROW_START, REVENUE_ROW_END + 1):
        invoice_date = to_iso_date(ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value)
        description = to_text(ws_formula.cell(r, 4).value)
        invoice_value = to_num_cell(ws_formula, ws_values, r, 6)
        if not invoice_date or not description:
            continue
        currency = to_text(ws_formula.cell(r, 7).value) or "IDR"
        exchange = to_num_cell(ws_formula, ws_values, r, 8, 1.0) or 1.0
        invoice_number = to_text(ws_formula.cell(r, 9).value) or None
        client = to_text(ws_formula.cell(r, 10).value) or None
        receive_date = to_iso_date(ws_values.cell(r, 11).value or ws_formula.cell(r, 11).value)
        amount_received = to_num_cell(ws_formula, ws_values, r, 12)
        ppn = to_num_cell(ws_formula, ws_values, r, 13)
        pph_23 = to_num_cell(ws_formula, ws_values, r, 14)
        transfer_fee = to_num_cell(ws_formula, ws_values, r, 15)
        remark = to_text(ws_formula.cell(r, 16).value) or None

        cur = conn.execute(
            """
            INSERT INTO revenues (
                invoice_date, description, invoice_value, currency, currency_exchange,
                invoice_number, client, receive_date, amount_received,
                ppn, pph_23, transfer_fee, remark, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                invoice_date,
                description,
                invoice_value,
                currency,
                exchange,
                invoice_number,
                client,
                receive_date,
                amount_received,
                ppn,
                pph_23,
                transfer_fee,
                remark,
                now,
            ),
        )
        revenue_id = int(cur.lastrowid)
        tag_cell_mapping(conn, sheet_id, r, 2, "revenues", revenue_id, "invoice_date")
        tag_cell_mapping(conn, sheet_id, r, 4, "revenues", revenue_id, "description")
        tag_cell_mapping(conn, sheet_id, r, 6, "revenues", revenue_id, "invoice_value")
        tag_cell_mapping(conn, sheet_id, r, 13, "revenues", revenue_id, "ppn")
        tag_cell_mapping(conn, sheet_id, r, 14, "revenues", revenue_id, "pph_23")
        tag_cell_mapping(conn, sheet_id, r, 15, "revenues", revenue_id, "transfer_fee")
        tag_report_row(conn, "revenues", revenue_id, report_year, source_excel)
        count += 1

    conn.commit()
    return count


def import_taxes(
    conn: sqlite3.Connection,
    sheet_id: int | None,
    ws_formula,
    ws_values,
    report_year: int,
    source_excel: str,
) -> int:
    count = 0
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    for r in range(TAX_ROW_START, TAX_ROW_END + 1):
        tax_date = to_iso_date(ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value)
        description = to_text(ws_formula.cell(r, 4).value)
        transaction_value = to_num_cell(ws_formula, ws_values, r, 6)
        if not tax_date or not description:
            continue
        currency = to_text(ws_formula.cell(r, 7).value) or "IDR"
        exchange = to_num_cell(ws_formula, ws_values, r, 8, 1.0) or 1.0
        ppn = to_num_cell(ws_formula, ws_values, r, 10)
        pph_21 = to_num_cell(ws_formula, ws_values, r, 12)
        pph_23 = to_num_cell(ws_formula, ws_values, r, 14)
        pph_26 = to_num_cell(ws_formula, ws_values, r, 16)

        cur = conn.execute(
            """
            INSERT INTO taxes (
                date, description, transaction_value, currency, currency_exchange,
                ppn, pph_21, pph_23, pph_26, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                tax_date,
                description,
                transaction_value,
                currency,
                exchange,
                ppn,
                pph_21,
                pph_23,
                pph_26,
                now,
            ),
        )
        tax_id = int(cur.lastrowid)
        tag_cell_mapping(conn, sheet_id, r, 2, "taxes", tax_id, "date")
        tag_cell_mapping(conn, sheet_id, r, 4, "taxes", tax_id, "description")
        tag_cell_mapping(conn, sheet_id, r, 6, "taxes", tax_id, "transaction_value")
        tag_cell_mapping(conn, sheet_id, r, 10, "taxes", tax_id, "ppn")
        tag_cell_mapping(conn, sheet_id, r, 12, "taxes", tax_id, "pph_21")
        tag_cell_mapping(conn, sheet_id, r, 14, "taxes", tax_id, "pph_23")
        tag_cell_mapping(conn, sheet_id, r, 16, "taxes", tax_id, "pph_26")
        tag_report_row(conn, "taxes", tax_id, report_year, source_excel)
        count += 1
    conn.commit()
    return count


def import_dividends(
    conn: sqlite3.Connection,
    wb_values,
    report_year: int,
    source_excel: str,
) -> int:
    # dividend tidak lagi diimpor dari excel.
    # data dividen tahunan dikelola dari aplikasi menggunakan
    # profit ditahan + nama penerima.
    return 0


def build_expense_blocks(ws_formula, ws_values) -> dict:
    # kembalikan {'standalone_rows': [...], 'batch_blocks': [(seq, header_row, start_row, end_row), ...]}
    headers: list[tuple[int, int]] = []
    for r in range(1, ws_formula.max_row + 1):
        label = ws_formula.cell(r, 2).value
        if isinstance(label, str) and label.strip().lower().startswith("expense#"):
            digits = "".join(ch for ch in label if ch.isdigit())
            seq = int(digits) if digits else len(headers) + 1
            headers.append((seq, r))

    batch_blocks: list[tuple[int, int, int, int]] = []
    for i, (seq, header_row) in enumerate(headers):
        next_row = headers[i + 1][1] if i + 1 < len(headers) else ws_formula.max_row + 1
        batch_blocks.append((seq, header_row, header_row + 1, next_row - 1))

    # baris mandiri: antara akhir bagian tax dan header expense# pertama
    first_expense_row = headers[0][1] if headers else ws_formula.max_row + 1
    standalone_start = TAX_ROW_END + 1
    standalone_rows: list[int] = []
    for r in range(standalone_start, first_expense_row):
        date_val = ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value
        desc_val = to_text(ws_formula.cell(r, 4).value)
        amount_val = to_num_cell(ws_formula, ws_values, r, 6)
        if date_val and desc_val and amount_val > 0:
            standalone_rows.append(r)

    return {'standalone_rows': standalone_rows, 'batch_blocks': batch_blocks}


def detect_category_id(
    ws_formula,
    ws_values,
    row_num: int,
    category_ids_by_code: dict[str, int],
    category_ids_by_name: dict[str, int],
    description: str,
    subcategory: str = "",
    category_col: int | None = None,
) -> int:
    expense_col_to_code = {
        9: "A",
        10: "B",
        11: "C",
        12: "D",
        13: "E",
        14: "F",
        15: "G",
        16: "H",
        17: "I",
    }
    if category_col in expense_col_to_code:
        code = expense_col_to_code[category_col]
        return category_ids_by_code.get(code, category_ids_by_code.get("A", 1))

    description_lower = description.lower()
    if "bank" in description_lower:
        return category_ids_by_code.get("E", category_ids_by_code.get("A", 1))

    standard_sub = normalize_subcategory(subcategory)
    if not standard_sub:
        standard_sub = normalize_subcategory(description)

    sub_id = category_ids_by_name.get(standard_sub.lower())
    if sub_id:
        return sub_id

    # pencocokan fuzzy untuk keyword umum
    for keyword, cat_id in category_ids_by_name.items():
        if keyword in description_lower:
            return cat_id

    return category_ids_by_code.get("A", 1)


def detect_expense_value_column(ws_formula, ws_values, row_num: int) -> int | None:
    for col in range(9, 18):
        val = to_num_cell(ws_formula, ws_values, row_num, col)
        if val > 0:
            return col
    return None


def import_expenses(
    conn: sqlite3.Connection,
    sheet_id: int | None,
    ws_formula,
    ws_values,
    category_ids_by_code: dict[str, int],
    category_ids_by_name: dict[str, int],
    report_year: int,
    source_excel: str,
) -> tuple[int, int]:
    settlement_count = 0
    expense_count = 0
    now = dt.datetime.now().isoformat(sep=" ", timespec="seconds")
    result = build_expense_blocks(ws_formula, ws_values)
    standalone_rows = result['standalone_rows']
    batch_blocks = result['batch_blocks']

    for r in standalone_rows:
        date_iso = to_iso_date(ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value)
        description = to_text(ws_formula.cell(r, 4).value)
        amount = to_num_cell(ws_formula, ws_values, r, 6)
        category_value_col = detect_expense_value_column(ws_formula, ws_values, r)

        if amount <= 0 and category_value_col is not None:
            amount = to_num_cell(ws_formula, ws_values, r, category_value_col)

        if not date_iso or not description or amount <= 0:
            continue
        currency = to_text(ws_formula.cell(r, 7).value) or "IDR"
        rate = to_num_cell(ws_formula, ws_values, r, 8, 1.0) or 1.0
        source = to_text(ws_formula.cell(r, 5).value) or None
        category_id = detect_category_id(
            ws_formula,
            ws_values,
            r,
            category_ids_by_code,
            category_ids_by_name,
            description,
            category_col=category_value_col,
        )

        cur_settlement = conn.execute(
            """
            INSERT INTO settlements (
                title, description, user_id, settlement_type, status,
                created_at, updated_at, completed_at, advance_id
            ) VALUES (?, ?, 2, 'single', 'completed', ?, ?, ?, NULL)
            """,
            (
                description[:200],
                f"Imported from Sheet1 row {r}",
                now,
                now,
                now,
            ),
        )
        settlement_id = int(cur_settlement.lastrowid)
        settlement_count += 1
        tag_cell_mapping(conn, sheet_id, r, 4, "settlements", settlement_id, "title")
        tag_report_row(conn, "settlements", settlement_id, report_year, source_excel)

        cur_expense = conn.execute(
            """
            INSERT INTO expenses (
                settlement_id, category_id, description, amount, date,
                source, currency, currency_exchange, evidence_path,
                evidence_filename, status, notes, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, 'approved', ?, ?)
            """,
            (
                settlement_id,
                category_id,
                description,
                amount,
                date_iso,
                source,
                currency,
                rate,
                f"Imported from row {r}",
                now,
            ),
        )
        expense_id = int(cur_expense.lastrowid)
        tag_cell_mapping(conn, sheet_id, r, 2, "expenses", expense_id, "date")
        tag_cell_mapping(conn, sheet_id, r, 4, "expenses", expense_id, "description")
        tag_cell_mapping(conn, sheet_id, r, 6, "expenses", expense_id, "amount")
        tag_report_row(conn, "expenses", expense_id, report_year, source_excel)
        expense_count += 1

    for seq, header_row, start_row, end_row in batch_blocks:
        valid_rows = []
        current_subcategory = ""
        for r in range(start_row, end_row + 1):
            date_iso = to_iso_date(ws_values.cell(r, 2).value or ws_formula.cell(r, 2).value)
            description = to_text(ws_formula.cell(r, 4).value)
            amount = to_num_cell(ws_formula, ws_values, r, 6)
            currency = to_text(ws_formula.cell(r, 7).value) or "IDR"
            rate = to_num_cell(ws_formula, ws_values, r, 8, 1.0) or 1.0
            category_value_col = detect_expense_value_column(ws_formula, ws_values, r)

            if amount <= 0 and category_value_col is not None:
                amount = to_num_cell(ws_formula, ws_values, r, category_value_col)
                possible_curr = to_text(ws_formula.cell(r, category_value_col + 1).value)
                if (
                    possible_curr
                    and len(possible_curr) == 3
                    and possible_curr.isalpha()
                    and possible_curr.upper() != "IDR"
                ):
                    currency = possible_curr.upper()
                    rate = to_num_cell(ws_formula, ws_values, r, category_value_col + 2, 1.0) or 1.0

            if description and not date_iso and amount <= 0:
                normalized = normalize_subcategory(description)
                if normalized:
                    current_subcategory = normalized
                continue

            if not date_iso or not description or amount <= 0:
                continue
            source = to_text(ws_formula.cell(r, 5).value) or None
            category_id = detect_category_id(
                ws_formula,
                ws_values,
                r,
                category_ids_by_code,
                category_ids_by_name,
                description,
                current_subcategory,
                category_col=category_value_col,
            )
            valid_rows.append(
                (
                    r,
                    date_iso,
                    description,
                    amount,
                    currency,
                    rate,
                    source,
                    category_id,
                    current_subcategory,
                )
            )

        if not valid_rows:
            continue

        title = to_text(ws_formula.cell(header_row, 4).value)
        if not title or title.upper() == "N/A":
            title = valid_rows[0][2]
        title = title or f"Expense Group {seq}"

        cur_settlement = conn.execute(
            """
            INSERT INTO settlements (
                title, description, user_id, settlement_type, status,
                created_at, updated_at, completed_at, advance_id
            ) VALUES (?, ?, 2, 'batch', 'completed', ?, ?, ?, NULL)
            """,
            (
                title[:200],
                f"Imported from Sheet1 row {header_row}",
                now,
                now,
                now,
            ),
        )
        settlement_id = int(cur_settlement.lastrowid)
        settlement_count += 1
        tag_cell_mapping(conn, sheet_id, header_row, 4, "settlements", settlement_id, "title")
        tag_report_row(conn, "settlements", settlement_id, report_year, source_excel)

        for r, date_iso, description, amount, currency, rate, source, category_id, subcategory in valid_rows:
            note = f"Imported from row {r}"
            if subcategory:
                note += f" | Subcategory: {subcategory}"
            cur_expense = conn.execute(
                """
                INSERT INTO expenses (
                    settlement_id, category_id, description, amount, date,
                    source, currency, currency_exchange, evidence_path,
                    evidence_filename, status, notes, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, 'approved', ?, ?)
                """,
                (
                    settlement_id,
                    category_id,
                    description,
                    amount,
                    date_iso,
                    source,
                    currency,
                    rate,
                    note,
                    now,
                ),
            )
            expense_id = int(cur_expense.lastrowid)
            tag_cell_mapping(conn, sheet_id, r, 2, "expenses", expense_id, "date")
            tag_cell_mapping(conn, sheet_id, r, 4, "expenses", expense_id, "description")
            tag_cell_mapping(conn, sheet_id, r, 6, "expenses", expense_id, "amount")
            tag_report_row(conn, "expenses", expense_id, report_year, source_excel)
            expense_count += 1

    conn.commit()
    return settlement_count, expense_count


def main() -> None:
    args = parse_args()
    excel_path = os.path.abspath(args.excel)
    file_short = os.path.basename(excel_path)

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb_formula = load_workbook(excel_path, data_only=False)
    wb_values = load_workbook(excel_path, data_only=True)

    if "Revenue-Cost_2024" in wb_formula.sheetnames:
        ws_formula = wb_formula["Revenue-Cost_2024"]
        ws_values = wb_values["Revenue-Cost_2024"]
    else:
        ws_formula = wb_formula.active
        ws_values = wb_values[ws_formula.title]
    year = args.year or extract_year_from_header(
        ws_formula,
        excel_path=excel_path,
        workbook=wb_formula,
    )

    output_db = build_output_db_path(
        excel_path=excel_path,
        output_db=args.output_db,
        output_dir=args.output_dir,
        year=year,
    )
    os.makedirs(os.path.dirname(output_db), exist_ok=True)

    conn = sqlite3.connect(output_db)
    try:
        ensure_schema(conn, with_excel_metadata=args.with_excel_metadata)
        cat_code_map, cat_name_map = ensure_reference_data(conn, with_excel_metadata=args.with_excel_metadata)
        purge_year_data(
            conn,
            year,
            only_tables={"dividends", "dividend_settings"} if args.only_dividends else None,
        )
        sheet_id = store_excel_structure(conn, ws_formula, ws_values, with_excel_metadata=args.with_excel_metadata)
        revenues = 0
        taxes = 0
        settlements = 0
        expenses = 0

        if not args.only_dividends:
            revenues = import_revenues(conn, sheet_id, ws_formula, ws_values, year, file_short)
            taxes = import_taxes(conn, sheet_id, ws_formula, ws_values, year, file_short)
            settlements, expenses = import_expenses(
                conn,
                sheet_id,
                ws_formula,
                ws_values,
                cat_code_map,
                cat_name_map,
                year,
                file_short,
            )
        dividends = 0

        formula_count = 0
        cell_count = 0
        if args.with_excel_metadata:
            formula_count = conn.execute(
                "SELECT COUNT(*) FROM excel_cells WHERE is_formula = 1"
            ).fetchone()[0]
            cell_count = conn.execute("SELECT COUNT(*) FROM excel_cells").fetchone()[0]

        print("Conversion completed")
        print(f"Excel source      : {excel_path}")
        print(f"Output DB         : {output_db}")
        print(f"Detected year     : {year}")
        print(f"Metadata mode     : {'ON' if args.with_excel_metadata else 'OFF'}")
        if args.with_excel_metadata:
            print(f"Sheets stored     : 1")
            print(f"Cells stored      : {cell_count}")
            print(f"Formula cells     : {formula_count}")
        print(f"Revenues inserted : {revenues}")
        print(f"Taxes inserted    : {taxes}")
        print(f"Dividends inserted: {dividends}")
        print(f"Settlements ins.  : {settlements}")
        print(f"Expenses inserted : {expenses}")
        print("Default users:")
        print("- manager1 / manager12345")
        print("- staff1 / staff12345")
        print("- staff2 / staff67890")
        print("- mitra1 / mitra12345")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
