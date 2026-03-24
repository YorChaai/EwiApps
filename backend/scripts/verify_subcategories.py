#!/usr/bin/env python3
# verifikasi semua subkategori unik di tabel 3 (operation cost) pada file excel
import openpyxl

file_path = r'D:\2. Organize\1. Projects\MiniProjectKPI_EWI\excel\Revenue-Cost_2024_contoh.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

# tabel 3 dimulai setelah bagian tax. dari kode: tax_row_end = 36
# subkategori muncul di baris yang memiliki deskripsi (kolom 4) tapi tanpa tanggal (kolom 2) dan tanpa jumlah (kolom 6)

# pertama, cari header expense#
expense_headers = []
for r in range(1, ws.max_row + 1):
    label = ws.cell(r, 2).value
    if isinstance(label, str) and label.strip().lower().startswith("expense#"):
        expense_headers.append(r)

print(f"Found {len(expense_headers)} Expense# headers at rows: {expense_headers}")
print()

# sekarang scan di dalam setiap blok expense untuk baris subkategori
# teks subkategori -> list dari (nomor expense, baris)
subcategories = {}

for i, header_row in enumerate(expense_headers):
    next_row = expense_headers[i + 1] if i + 1 < len(expense_headers) else ws.max_row + 1
    expense_label = ws.cell(header_row, 2).value

    for r in range(header_row + 1, next_row):
        date_val = ws.cell(r, 2).value
        desc_val = ws.cell(r, 4).value
        amount_val = ws.cell(r, 6).value

        if desc_val and isinstance(desc_val, str) and desc_val.strip():
            desc_text = desc_val.strip()
            has_date = date_val is not None
            has_amount = amount_val is not None and (isinstance(amount_val, (int, float)) and float(amount_val) > 0)

            if not has_date and not has_amount:
                if desc_text not in subcategories:
                    subcategories[desc_text] = []
                subcategories[desc_text].append((expense_label, r))

print("=" * 80)
print(f"UNIQUE SUB-CATEGORIES FOUND: {len(subcategories)}")
print("=" * 80)
for idx, (subcat, locations) in enumerate(sorted(subcategories.items(), key=lambda x: x[0].lower()), 1):
    print(f"{idx:3d}. '{subcat}'")
    for exp_label, row in locations:
        print(f"      -> row {row} (under {exp_label})")

print()
print("=" * 80)
print("SORTED LIST (for mapping):")
print("=" * 80)
for subcat in sorted(subcategories.keys(), key=lambda x: x.lower()):
    print(f"  {subcat}")

wb.close()
