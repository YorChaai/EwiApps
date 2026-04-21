# shared helper/utility function untuk package reports

import re
import logging
from datetime import datetime
from typing import Optional, Any, List, Dict, Tuple
from copy import copy
from flask import current_app
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as openpyxl_get_column_letter

# Setup logging
logger = logging.getLogger(__name__)

# Default column constants
DEFAULT_START_COL = 2
DEFAULT_END_COL = 17


def _default_report_year() -> int:
    """Get default report year from config or current year."""
    return int(current_app.config.get('REPORT_DEFAULT_YEAR', datetime.now().year))


def _safe_set_cell(ws, row: int, col: int, value: Optional[Any]) -> None:
    """Set cell value only if not a merged cell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _get_top_left_cell(ws, row: int, col: int):
    """
    Get the top-left cell of a merged range.
    If cell is not merged, return the cell itself.
    """
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(row=min_row, column=min_col)
    return ws.cell(row=row, column=col)


def _safe_set_cell_with_merge(ws, row: int, col: int, value: Optional[Any]) -> None:
    """
    Set cell value, handling merged cells correctly.
    Writes to top-left cell of merged range.
    """
    cell = _get_top_left_cell(ws, row, col)
    cell.value = value if value is not None else ''


def _merge_description_cell(
    ws,
    row: int,
    description: Optional[str],
    col_start: int = 4,
    col_end: int = 5
) -> bool:
    """
    Merge columns D:E (or specified range) for description field.
    Handles unmerge first to avoid conflicts, then writes description.

    Returns:
        True if merge successful, False otherwise
    """
    try:
        ws.unmerge_cells(f'{_get_column_letter(col_start)}{row}:{_get_column_letter(col_end)}{row}')
    except Exception:
        pass

    try:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        _safe_set_cell_with_merge(ws, row, col_start, description or '')
        return True
    except Exception as e:
        logger.debug(f'Failed to merge description cell at row {row}: {e}')
        _safe_set_cell(ws, row, col_start, description or '')
        return False


def _get_column_letter(col_num: int) -> str:
    """Convert column number to letter (1->A, 2->B, etc.)."""
    return openpyxl_get_column_letter(col_num)


def _safe_set_number(
    ws,
    row: int,
    col: int,
    value: Optional[float],
    number_format: str = '#,##0'
) -> None:
    """Set numeric value with format, skip if merged cell."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    cell.number_format = number_format
def _unmerge_region(ws, start_row, end_row, start_col, end_col):
    """
    Find all merged ranges that overlap with the specified region and unmerge them.
    Restores borders after unmerge to avoid 'white gaps'.
    """
    ranges_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        overlaps = not (
            max_row < start_row or min_row > end_row or
            max_col < start_col or min_col > end_col
        )
        if overlaps:
            ranges_to_unmerge.append(merged_range)

    for r_range in ranges_to_unmerge:
        # Ambil border dari cell pertama (top-left) sebelum unmerge
        first_cell = ws.cell(row=r_range.min_row, column=r_range.min_col)
        original_border = copy(first_cell.border)

        try:
            ws.unmerge_cells(r_range.coord)
            # Terapkan kembali border ke semua cell di range tersebut
            for r in range(r_range.min_row, r_range.max_row + 1):
                for c in range(r_range.min_col, r_range.max_col + 1):
                    ws.cell(row=r, column=c).border = original_border
        except Exception as e:
            logger.debug(f'Warning: Failed to unmerge/restore {r_range}: {e}')


def _normalize_external_formula_refs(wb):
    # ubah referensi formula external seperti '[1]Revenue-Cost_2024'!A1 ke referensi sheet internal
    pat_quoted = re.compile(r"'\[\d+\]([^']+)'!")
    pat_plain = re.compile(r"\[\d+\]([A-Za-z0-9_\- ]+)!")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith('=')):
                    continue
                new_value = pat_quoted.sub(r"'\1'!", value)
                new_value = pat_plain.sub(r"\1!", new_value)
                if new_value != value:
                    cell.value = new_value


def _clear_range(ws, start_row, end_row, start_col=2, end_col=17):
    # Membersihkan nilai (value) sel tanpa merubah layout atau penggabungan (merge)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _clear_data_keep_formulas(ws, start_row, end_row, start_col=2, end_col=17):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                continue
            cell.value = None


def _clear_range_force(ws, start_row, end_row, start_col=2, end_col=17, reset_style=True):
    """
    ✅ FORCE CLEAR: Clear range including handling merged cells.
    If reset_style=True, also resets all styles (borders, fonts, etc).
    """
    # Step 1: Unmerge all cells in this range first
    ranges_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        overlaps = not (
            max_row < start_row or min_row > end_row or
            max_col < start_col or min_col > end_col
        )
        if overlaps:
            ranges_to_unmerge.append(str(merged_range))

    for merge_range in ranges_to_unmerge:
        try:
            ws.unmerge_cells(merge_range)
        except Exception:
            pass

    # Step 2: Clear all values and optionally reset ALL formatting
    if reset_style:
        empty_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
        empty_font = Font(bold=False, italic=False, color='000000', size=11, name='Calibri')
        empty_fill = PatternFill(fill_type=None)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            if reset_style:
                cell.number_format = 'General'
                cell.border = empty_border
                cell.font = empty_font
                cell.fill = empty_fill
                cell.alignment = Alignment(horizontal='general', vertical='bottom', wrap_text=False)


def _set_rows_hidden(ws, start_row, end_row, hidden):
    if start_row > end_row:
        return
    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].hidden = hidden


def _safe_text(value):
    if value is None:
        return ''
    return str(value)


def _extract_imported_row(notes):
    text = _safe_text(notes)
    match = re.search(r'Imported from row\s+(\d+)', text, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _extract_imported_sheet_row(text):
    raw = _safe_text(text)
    match = re.search(r'Imported from Sheet1 row\s+(\d+)', raw, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _is_date_like(value):
    if value is None:
        return False
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return True
    text = str(value).strip()
    if not text:
        return False
    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%b-%y', '%d/%m/%Y'):
        try:
            datetime.strptime(text[:19], fmt)
            return True
        except ValueError:
            continue
    return False


def _is_template_detail_data_row(ws, row_num):
    """
    Detect if a row is a detail data row (for expense items).

    A row is considered a detail row if:
    1. Column B has a date (primary check)
    2. AND Column C has a sequence number (secondary check)
    3. AND Column D has description text (tertiary check)

    CRITICAL: Subcategory header rows have text in column D but NO date in column B
    and NO sequence number in column C. They must NOT be treated as detail rows.
    """
    # Primary check: Column B has date
    col_b = ws.cell(row=row_num, column=2).value
    has_date = _is_date_like(col_b)

    # Secondary check: Column C has sequence number (1, 2, 3, etc.)
    col_c = ws.cell(row=row_num, column=3).value
    has_seq = isinstance(col_c, (int, float)) and col_c > 0

    # Tertiary check: Column D has description text
    col_d = ws.cell(row=row_num, column=4).value
    has_desc = isinstance(col_d, str) and col_d.strip()

    # ✅ CRITICAL FIX: A detail row MUST have both date AND sequence number
    # Subcategory headers only have description text, no date or sequence
    if has_date and has_seq:
        return True

    # Fallback: If has date and description but no sequence, still consider it a detail row
    # (for cases where sequence might be missing in template)
    if has_date and has_desc:
        return True

    return False


def _map_expense_category_index(expense, cat_names):
    if not expense or not expense.category:
        return 0
    # Cari kategori akar (root)
    curr = expense.category
    while curr.parent:
        curr = curr.parent

    root_name = curr.name
    if root_name in cat_names:
        return cat_names.index(root_name)
    return 0


def _map_expense_category_index_from_name(category_name, cat_names):
    if not category_name:
        return None
    name = str(category_name).strip().lower()
    for i, cn in enumerate(cat_names):
        if str(cn).strip().lower() == name:
            return i
    return None


def _pick_template_formula_col(ws, row_num, start_col=9, end_col=17):
    for c in range(start_col, end_col + 1):
        v = ws.cell(row=row_num, column=c).value
        if isinstance(v, str) and v.startswith('='):
            return c
    return None


def _write_expense_line(ws, row_num, seq_num, expense, cat_names):
    ws.cell(row=row_num, column=2).value = expense.date
    ws.cell(row=row_num, column=3).value = seq_num
    ws.cell(row=row_num, column=4).value = expense.description
    ws.cell(row=row_num, column=5).value = expense.settlement.title if expense.settlement else '-'
    ws.cell(row=row_num, column=6).value = expense.amount
    ws.cell(row=row_num, column=7).value = 'IDR'
    ws.cell(row=row_num, column=8).value = 1

    fallback_col = 9 + _map_expense_category_index(expense, cat_names)
    target_col = _pick_template_formula_col(ws, row_num) or fallback_col
    ws.cell(row=row_num, column=target_col).value = f'=$F{row_num}*$H{row_num}'


def _write_expense_detail_line(ws, row_num, seq_num, expense, cat_names):
    ws.cell(row=row_num, column=2).value = expense.date
    ws.cell(row=row_num, column=3).value = seq_num
    ws.cell(row=row_num, column=4).value = expense.description
    ws.cell(row=row_num, column=6).value = expense.amount
    ws.cell(row=row_num, column=7).value = 'IDR'
    ws.cell(row=row_num, column=8).value = 1
    target_col = _pick_template_formula_col(ws, row_num) or 9
    ws.cell(row=row_num, column=target_col).value = f'=$F{row_num}*$H{row_num}'


def _get_expense_blocks(ws):
    headers = []
    for row_num in range(1, ws.max_row + 1):
        label = ws.cell(row=row_num, column=2).value
        if isinstance(label, str) and label.strip().lower().startswith('expense#'):
            digits = ''.join(ch for ch in label if ch.isdigit())
            seq = int(digits) if digits else len(headers) + 1
            headers.append((seq, row_num))

    blocks = []
    for idx, (seq, header_row) in enumerate(headers):
        # Get next header row, but cap at reasonable distance (max 200 rows per block)
        if idx + 1 < len(headers):
            next_header = headers[idx + 1][1]
            # Limit block size to avoid overflow into next block
            end_row = min(next_header - 1, header_row + 200)
        else:
            # For last block, find actual end by scanning for empty rows
            end_row = header_row + 1
            for r in range(header_row + 1, min(ws.max_row + 1, header_row + 200)):
                row_has_data = any(
                    ws.cell(row=r, column=c).value is not None
                    for c in range(2, 8)
                )
                if row_has_data:
                    end_row = r
                else:
                    # Stop at first completely empty row
                    break

        start_row = header_row + 1
        blocks.append((seq, header_row, start_row, end_row))

    return blocks


def _parse_iso_date(date_str):
    """Parse ISO date string and return date object."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return None


def _format_date_dd_mmm_yy(date_value) -> str:
    """
    Format date to DD-Mon-YY format (e.g., 11-Dec-24).
    Accepts datetime object, date object, or ISO date string.
    """
    if not date_value:
        return ''
    try:
        if isinstance(date_value, (datetime, datetime)):
            return date_value.strftime('%d-%b-%y')
        elif isinstance(date_value, str):
            dt = datetime.strptime(date_value, '%Y-%m-%d')
            return dt.strftime('%d-%b-%y')
        return ''
    except (ValueError, TypeError) as e:
        logger.debug(f'Failed to format date: {date_value} - {e}')
        return ''


def _to_float(value, default=0.0):
    """Convert value to float with default fallback."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(',', '').strip())
    except Exception:
        return default


def _set_formula_with_format(ws, row: int, col: int, formula: str, number_format: str = '#,##0') -> None:
    """
    Set formula with number format in one call.
    Eliminates redundancy of setting formula and format separately.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = formula
    cell.number_format = number_format


def _set_date_with_format(ws, row: int, col: int, date_value, date_format: str = 'dd-mmm-yy') -> None:
    """
    Set date value with format in one call.
    Eliminates redundancy of setting date and format separately.
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = _format_date_dd_mmm_yy(date_value)
    cell.number_format = date_format


def _as_iso_date(value):
    text = _safe_text(value)
    if not text:
        return ''
    return text[:10]


def _idr_from_currency(amount, currency, exchange):
    curr = _safe_text(currency).upper() or 'IDR'
    rate = _to_float(exchange, default=1.0) or 1.0
    nominal = _to_float(amount)
    if curr != 'IDR':
        return nominal * rate
    return nominal


def _shorten(text, size):
    text = _safe_text(text)
    return text[:size] + ('...' if len(text) > size else '')


def _map_expense_column(category_name, cat_names):
    if not category_name:
        return 0

    # Clean input name
    name = str(category_name).strip().lower()

    # Clean category names list for matching
    clean_cat_names = [str(c).strip().lower() for c in cat_names]

    # 1. Exact match (case-insensitive)
    if name in clean_cat_names:
        return clean_cat_names.index(name)

    # 2. Multi-category match (e.g., "A , B" matches column "A" or "B")
    if " , " in category_name:
        parts = [p.strip().lower() for p in category_name.split(",")]
        for p in parts:
            if p in clean_cat_names:
                return clean_cat_names.index(p)

    # 3. Partial match (induk kategori)
    for i, cat in enumerate(clean_cat_names):
        if cat in name or name in cat:
            return i

    return 0


def _extract_batch_number(text):
    raw = _safe_text(text)
    match = re.search(r'\bbatch\s*#?\s*(\d+)\b', raw, flags=re.IGNORECASE)
    if match:
        try:
            return int(match.group(1))
        except ValueError:
            return 10**9
    return 10**9


def _is_batch_settlement(settlement_type, settlement_title):
    stype = _safe_text(settlement_type).strip().lower()
    if stype == 'batch':
        return True
    if stype == 'single':
        return False
    return 'batch' in _safe_text(settlement_title).strip().lower()


def _clean_settlement_title(title):
    text = _safe_text(title).strip()
    if not text:
        return 'Tanpa Settlement'
    text = re.sub(r'^\s*single\s*[-:]\s*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'^\s*batch\s*#?\s*\d+\s*[-:]\s*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'^\s*batch\s*[-:]\s*', '', text, flags=re.IGNORECASE)
    text = text.strip()
    return text or _safe_text(title).strip() or 'Tanpa Settlement'


def _extract_note_subcategory(notes):
    """Extract subcategory from notes like 'Subcategory: Value |'"""
    text = _safe_text(notes).strip()
    if not text:
        return ''
    match = re.search(r'\bSubcategory:\s*([^|]+)', text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return ''


def _expense_subcategory_label(expense):
    """
    Extract subcategory from expense and append Parent Code.
    Example: "Allowance (A)"
    """
    label = ""

    # 1. Database value
    subcategory_name = _safe_text(expense.get('subcategory_name')).strip()
    if subcategory_name and subcategory_name != '-':
        label = subcategory_name
    else:
        raw_desc = _safe_text(expense.get('description')).strip()
        # 2. Check [SubCategory] prefix
        prefixed = re.match(r'^\[(.*?)\]\s*(.*)$', raw_desc)
        if prefixed:
            label = _safe_text(prefixed.group(1)).strip()
        else:
            # 3. Check notes for subcategory (matching frontend)
            notes_sub = _extract_note_subcategory(expense.get('notes'))
            if notes_sub:
                label = notes_sub
            else:
                # 4. Keyword matching (fallback)
                desc = raw_desc.lower()
                if 'rental tool' in desc: label = 'Rental Tool'
                elif 'sales' in desc: label = 'Sales'
                elif 'gaji' in desc or 'bonus' in desc: label = 'Gaji'
                elif 'pembuatan alat' in desc or 'mesin retort' in desc: label = 'Pembuatan Alat'
                elif 'thr' in desc or 'allowance' in desc: label = 'Allowance'
                elif 'data processing' in desc: label = 'Data Processing'
                elif 'moving slickline' in desc or 'project lampu' in desc: label = 'Project Operation'
                elif 'sampling tool' in desc or 'sparepart' in desc or 'ups biaya import' in desc: label = 'Sparepart'
                elif 'repair esor' in desc: label = 'Maintenance'
                elif 'licence' in desc or 'license' in desc: label = 'Software License'
                elif 'handphone operational' in desc: label = 'Operation'
                elif 'sewa ruangan' in desc or 'virtual office' in desc: label = 'Sewa Ruangan'
                elif 'modal kerja' in desc: label = 'Modal Kerja'
                elif 'team building' in desc: label = 'Team Building'
                elif 'biaya transaksi bank' in desc: label = 'Biaya Bank'

    if not label:
        return ''

    # APPEND PARENT CODE (A, B, C...)
    parent_code = _safe_text(expense.get('category_code')).strip()
    if parent_code and parent_code != '-' and parent_code != 'None':
        return f"{label} ({parent_code})"

    return label


def _group_annual_expenses(expenses, year):
    def _date_key(item):
        return _parse_iso_date(item.get('date')) or datetime(year, 1, 1).date()

    grouped = {}
    for e in expenses:
        # PENTING: Update label subkategori langsung di dictionary agar UI Flutter membacanya
        subcat_label = _expense_subcategory_label(e)
        if subcat_label:
            e['subcategory_name'] = subcat_label

        sid = e.get('settlement_id')
        if sid is None:
            key = f"title::{_safe_text(e.get('settlement_title'))}::{subcat_label}"
        else:
            key = f"sid::{sid}::{subcat_label}"
        grouped.setdefault(key, []).append(e)

    def _group_sort_key(items):
        first = items[0] if items else {}
        subcat_name = _expense_subcategory_label(first).strip().lower()
        is_batch = _is_batch_settlement(
            first.get('settlement_type'),
            first.get('settlement_title'),
        )
        settlement_id = int(first.get('settlement_id') or 0)
        batch_no = _extract_batch_number(first.get('settlement_title')) if is_batch else 0
        min_date = min((_date_key(x) for x in items), default=datetime(year, 1, 1).date())

        # Primary sort by subcategory name alphabetically (A-Z)
        if is_batch:
            if batch_no < 10**9:
                return (subcat_name, 1, batch_no, settlement_id, min_date)
            return (subcat_name, 1, settlement_id, min_date)
        return (subcat_name, 0, min_date, settlement_id)

    groups = list(grouped.values())
    for items in groups:
        # Sort items within each group by subcategory A-Z, then date, then id
        items.sort(key=lambda x: (
            (_expense_subcategory_label(x) or '').lower(),
            _date_key(x),
            int(x.get('id') or 0)
        ))
    groups.sort(key=_group_sort_key)
    return groups
